import os
import ctypes
import sqlite3
import tempfile
import asyncio
import edge_tts

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font


import io
import sys
import webbrowser
import json
import atexit

# Safely wrap atexit.register to prevent "RuntimeError: can't register atexit after shutdown" in threads
_orig_register = atexit.register
def _safe_register(func, *args, **kwargs):
    try:
        return _orig_register(func, *args, **kwargs)
    except RuntimeError as e:
        if "atexit after shutdown" in str(e):
            return func
        raise
atexit.register = _safe_register

from flask import Flask, request, jsonify, send_file, send_from_directory, redirect, Response
from flask_cors import CORS
from models import db, Category, Product, Partner, Order, OrderDetail, CashVoucher, CustomerPrice, AppSetting, ComboItem, PrintTemplate, User, BankAccount, BankTransaction, Event, EventLog, InventoryAudit, InventoryAuditDetail, StockBatch, InventoryConversion, AccountingTemplate, AccountingMapping, RemoteScanQueue
from migrations_manager import ensure_schema, initialize_stock_batches
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload
from datetime import datetime, timedelta, timezone
import re
import unicodedata
from sqlalchemy import event, inspect, extract, and_, or_, func
from sqlalchemy.engine import Engine
from werkzeug.security import generate_password_hash, check_password_hash


def get_vn_time():
    """Returns naive datetime in Vietnam Timezone (UTC+7)"""
    utc_now = datetime.now(timezone.utc)
    vn_tz = timezone(timedelta(hours=7))
    return utc_now.astimezone(vn_tz).replace(tzinfo=None)

def remove_accents(s):
    if not s: return ""
    s = str(s).lower()
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
    s = re.sub(r'[ìíịỉĩ]', 'i', s)
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
    s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
    s = re.sub(r'[đ]', 'd', s)
    # Combine with normalization for other characters
    nfkd_form = unicodedata.normalize('NFKD', s)
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)])

def normalize_date_sqlite(date_str):
    if not date_str: return "9999-12-31"
    date_str = str(date_str).strip()
    try:
        if '/' in date_str:
            parts = date_str.split('/')
            if len(parts) == 3:
                day = parts[0].zfill(2)
                month = parts[1].zfill(2)
                year = parts[2]
                if len(year) == 2: year = "20" + year
                return f"{year}-{month}-{day}"
        if '-' in date_str:
            parts = date_str.split('-')
            if len(parts) == 3:
                if len(parts[0]) == 4: return date_str # YYYY-MM-DD
                day = parts[0].zfill(2)
                month = parts[1].zfill(2)
                year = parts[2]
                if len(year) == 2: year = "20" + year
                return f"{year}-{month}-{day}"
    except:
        pass
    return date_str
    
def normalize_uom(uom):
    if not uom or not isinstance(uom, str):
        return uom
    
    clean = uom.strip().lower()
    if not clean:
        return ""
        
    uom_map = {
        'kg': 'Kg', 'ki': 'Kg', 'kilogram': 'Kg', 'kilo': 'Kg',
        'l': 'Lít', 'lit': 'Lít',
        'm': 'Mét', 'met': 'Mét',
        'cai': 'Cái', 'chiec': 'Cái',
        'thung': 'Thùng',
        'hop': 'Hộp',
        'chai': 'Chai', 'lo': 'Chai',
        'goi': 'Gói',
        'vien': 'Viên', 'vi': 'Vỉ',
        'tui': 'Túi',
        'cuon': 'Cuộn'
    }

    no_accent = remove_accents(clean)
    
    if clean in uom_map:
        return uom_map[clean]
    if no_accent in uom_map:
        return uom_map[no_accent]
        
    return clean.capitalize()

def safe_float(val, default=None):
    if val is None or str(val).strip() == '':
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def safe_int(val, default=0):
    if val is None or str(val).strip() == '':
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default

@event.listens_for(Engine, "connect")
def set_sqlite_custom_func(dbapi_connection, connection_record):
    # Fix: Only register these functions for SQLite connections
    # When using PostgreSQL (Render/Supabase), this would cause an AttributeError
    cursor = dbapi_connection.cursor()
    try:
        if hasattr(dbapi_connection, 'create_function'):
             dbapi_connection.create_function("remove_accents", 1, remove_accents)
             dbapi_connection.create_function("normalize_date", 1, normalize_date_sqlite)
             # Tune memory & auto-cleanup unused freelist pages in SQLite
             cursor.execute("PRAGMA cache_size = -2000")
             cursor.execute("PRAGMA auto_vacuum = INCREMENTAL")
             cursor.execute("PRAGMA wal_autocheckpoint = 200")
    except Exception:
        # If it fails (e.g. Postgres connection object doesn't have create_function), just ignore
        pass
    finally:
        cursor.close()

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
        if os.path.exists(os.path.join(base_dir, '_internal')):
             return os.path.join(base_dir, '_internal', relative_path)
        return os.path.join(base_dir, relative_path)
    # Go up one level from backend/ to project root
    base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

# Determine base path safely
if getattr(sys, 'frozen', False):
    # Packaged production mode: Try to use the App folder
    # If not writable (e.g. C:\Program Files), fallback to LocalAppData
    import platform
    current_dir = os.path.dirname(sys.executable)
    
    # Traverse up to find if we are in the development workspace (e.g. during tauri dev sidecar run)
    dev_workspace = None
    temp_dir = current_dir
    for _ in range(10):
        if os.path.exists(os.path.join(temp_dir, 'backend')) and os.path.exists(os.path.join(temp_dir, 'frontend')):
            dev_workspace = temp_dir
            break
        parent = os.path.dirname(temp_dir)
        if parent == temp_dir:
            break
        temp_dir = parent
        
    if dev_workspace:
        app_dir = dev_workspace
    else:
        # Safely traverse up past 'bin', 'resources', '_internal', 'src-tauri', 'target', 'release', 'debug' folders to find the main AppDir
        for _ in range(5):
            parent = os.path.dirname(current_dir)
            folder_name = os.path.basename(current_dir).lower()
            if folder_name in ['bin', 'resources', '_internal', 'src-tauri', 'target', 'release', 'debug']:
                current_dir = parent
            else:
                break
        app_dir = current_dir
    
    is_writable = False
    try:
        test_file = os.path.join(app_dir, ".write_test")
        with open(test_file, "w") as f:
            f.write("test")
        os.remove(test_file)
        is_writable = True
    except Exception:
        is_writable = False
        
    if is_writable:
        BASE_DIR = app_dir
    else:
        if platform.system() == "Windows":
            BASE_DIR = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")), "LyangPOS")
        else:
            BASE_DIR = os.path.expanduser("~/.lyangpos")
else:
    # Development mode
    try:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    except NameError:
        BASE_DIR = os.getcwd()

os.makedirs(BASE_DIR, exist_ok=True)

def get_storage_path(relative_path):
    path = os.path.join(BASE_DIR, relative_path)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    return path


static_folder_path = resource_path(os.path.join('frontend', 'dist'))
app = Flask(__name__, static_folder=static_folder_path, static_url_path='/')
app.secret_key = os.environ.get('SECRET_KEY', 'dev_key_super_secret_change_me_in_prod')

# Configure Permissive CORS for Safari (iOS/Mac) and Web/Mobile Client Access
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)

@app.after_request
def after_request_cors(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Terminal-Id, X-Requested-With, Accept, Origin'
    response.headers['Access-Control-Allow-Methods'] = 'GET, PUT, POST, DELETE, OPTIONS'
    return response

@app.before_request
def handle_options_preflight():
    if request.method == 'OPTIONS':
        response = Response(status=200)
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Terminal-Id, X-Requested-With, Accept, Origin'
        response.headers['Access-Control-Allow-Methods'] = 'GET, PUT, POST, DELETE, OPTIONS'
        return response




try:
    myappid = 'com.lyang.pos.v3.farmer.icon.v14' # increment version to refresh taskbar icon with new logo
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
except Exception:
    pass

# Port Configuration
CURRENT_PORT = 3579

# Logging setup (disabled file logging app_debug.log)
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s', handlers=[logging.NullHandler()])


# Database Configuration - Strictly Local SQLite with Smart Seed Auto-Preservation
db_file_path = get_storage_path(os.path.join("instance", "easypos.db"))

# Check if target db_file_path is missing or empty (0 bytes)
if not os.path.exists(db_file_path) or os.path.getsize(db_file_path) == 0:
    import shutil
    local_appdata = os.environ.get("LOCALAPPDATA", "")
    possible_seeds = [
        os.path.join(os.getcwd(), "instance", "easypos.db"),
        resource_path(os.path.join("instance", "easypos.db")),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "instance", "easypos.db"),
        os.path.join(local_appdata, "LyangPOS", "instance", "easypos.db") if local_appdata else "",
        os.path.join(local_appdata, "Programs", "LyangPOS", "instance", "easypos.db") if local_appdata else "",
    ]
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        possible_seeds.extend([
            os.path.join(exe_dir, "instance", "easypos.db"),
            os.path.join(os.path.dirname(exe_dir), "instance", "easypos.db")
        ])

    # Find the largest non-empty valid seed database
    best_seed = None
    max_size = 0
    for seed in possible_seeds:
        if seed and os.path.exists(seed) and os.path.abspath(seed) != os.path.abspath(db_file_path):
            try:
                size = os.path.getsize(seed)
                if size > max_size:
                    max_size = size
                    best_seed = seed
            except Exception:
                pass

    if best_seed:
        try:
            os.makedirs(os.path.dirname(db_file_path), exist_ok=True)
            shutil.copy2(best_seed, db_file_path)
            print(f"[DB INIT] Preserved existing database ({max_size} bytes) from {best_seed} -> {db_file_path}")
        except Exception as copy_err:
            print(f"[DB INIT] Failed to copy seed db: {copy_err}")

normalized_db_path = db_file_path.replace('\\', '/')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{normalized_db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# CORS Configuration: Permissive configuration for all origins, methods and headers to support Tauri LAN clients
CORS(app, resources={r"/*": {
    "origins": "*",
    "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    "allow_headers": ["*"]
}})

@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Private-Network'] = 'true'
    return response



ACTIVE_POS_TERMINALS = {}

@app.route('/api/ping', methods=['GET'])
def ping():
    return jsonify({'status': 'ok', 'message': 'pong'})

# =====================================
@app.route('/api/pos/terminal-state', methods=['POST'])
def update_pos_terminal_state():
    """Receive live cart / screen state heartbeat from a POS terminal."""
    try:
        data = request.json or {}
        terminal_id = data.get('terminal_id') or request.remote_addr or 'POS-DEFAULT'
        terminal_name = data.get('terminal_name') or f"Máy POS ({terminal_id})"
        user_name = data.get('user_name') or 'Thu ngân'
        
        now = datetime.now()
        
        # Check if there is an inspector update pending for this terminal
        existing = ACTIVE_POS_TERMINALS.get(terminal_id, {})
        if existing.get('remote_updated', False):
            # Clear flag and send sync command
            existing['remote_updated'] = False
            existing['last_active'] = now.isoformat()
            existing['ip_address'] = request.remote_addr
            action = existing.pop('action', None)
            return jsonify({
                'status': 'remote_sync',
                'action': action,
                'cart': existing.get('cart') or [],
                'partner': existing.get('partner'),
                'partner_name': existing.get('partner_name'),
                'payment_method': existing.get('payment_method'),
                'amount_paid': existing.get('amount_paid', 0),
                'cash_given': existing.get('cash_given', 0),
                'note': existing.get('note') or '',
                'total_amount': existing.get('total_amount', 0),
                'total_items': existing.get('total_items', 0)
            })

        ACTIVE_POS_TERMINALS[terminal_id] = {
            'terminal_id': terminal_id,
            'terminal_name': terminal_name,
            'user_name': user_name,
            'ip_address': request.remote_addr,
            'last_active': now.isoformat(),
            'cart': data.get('cart') or [],
            'partner': data.get('partner') or None,
            'partner_name': data.get('partner_name') or 'Khách lẻ',
            'payment_method': data.get('payment_method') or 'Cash',
            'amount_paid': data.get('amount_paid', 0),
            'cash_given': data.get('cash_given', 0),
            'note': data.get('note') or '',
            'total_amount': data.get('total_amount', 0),
            'total_items': data.get('total_items', 0),
            'status': data.get('status', 'active'),
            'current_page': data.get('current_page', 'POS'),
            'remote_updated': False
        }
        
        # Cleanup terminals inactive for > 3 minutes
        cutoff = now - timedelta(minutes=3)
        stale_ids = [tid for tid, tinfo in ACTIVE_POS_TERMINALS.items() 
                      if datetime.fromisoformat(tinfo['last_active']) < cutoff]
        for tid in stale_ids:
            ACTIVE_POS_TERMINALS.pop(tid, None)

        return jsonify({'status': 'ok', 'terminal_id': terminal_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/pos/terminal-state/edit-cart', methods=['POST'])
def edit_pos_terminal_cart():
    """Allow inspector terminal to modify the target terminal's cart / partner / state."""
    try:
        data = request.json or {}
        terminal_id = data.get('terminal_id')
        
        if not terminal_id or terminal_id not in ACTIVE_POS_TERMINALS:
            return jsonify({'error': 'Terminal not found'}), 404
            
        term = ACTIVE_POS_TERMINALS[terminal_id]
        
        if 'cart' in data:
            cart = data.get('cart') or []
            term['cart'] = cart
            term['total_items'] = sum(float(c.get('quantity') or 1) for c in cart)
            term['total_amount'] = sum(float(c.get('quantity') or 1) * float(c.get('price') or c.get('sale_price') or 0) for c in cart)
            
        if 'partner' in data:
            term['partner'] = data.get('partner')
            term['partner_name'] = data.get('partner_name') or (data.get('partner').get('name') if data.get('partner') else 'Khách lẻ')
            
        if 'payment_method' in data:
            term['payment_method'] = data.get('payment_method') or 'Cash'
            
        if 'amount_paid' in data:
            term['amount_paid'] = data.get('amount_paid', 0)
            
        if 'cash_given' in data:
            term['cash_given'] = data.get('cash_given', 0)
            
        if 'note' in data:
            term['note'] = data.get('note', '')
            
        term['remote_updated'] = True
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/pos/terminal-state/action', methods=['POST'])
def trigger_pos_terminal_action():
    """Allow inspector terminal to send an action (like save_order) to the target terminal."""
    try:
        data = request.json or {}
        terminal_id = data.get('terminal_id')
        action = data.get('action')
        
        if not terminal_id or terminal_id not in ACTIVE_POS_TERMINALS:
            return jsonify({'error': 'Terminal not found'}), 404
            
        term = ACTIVE_POS_TERMINALS[terminal_id]
        term['action'] = action
        term['remote_updated'] = True
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/pos/terminals', methods=['GET'])
def get_pos_terminals():
    """Return all active POS workstations and their live carts."""
    try:
        now = datetime.now()
        cutoff = now - timedelta(minutes=3)
        
        # Clean stale terminals
        stale_ids = [tid for tid, tinfo in ACTIVE_POS_TERMINALS.items() 
                     if datetime.fromisoformat(tinfo['last_active']) < cutoff]
        for tid in stale_ids:
            ACTIVE_POS_TERMINALS.pop(tid, None)

        terminals_list = list(ACTIVE_POS_TERMINALS.values())
        terminals_list.sort(key=lambda x: x['last_active'], reverse=True)
        
        import socket
        server_ips = ['127.0.0.1', '::1', 'localhost']
        try:
            hostname = socket.gethostname()
            ips = socket.gethostbyname_ex(hostname)[2]
            for ip in ips:
                if ip not in server_ips:
                    server_ips.append(ip)
        except Exception:
            pass
            
        return jsonify({
            'terminals': terminals_list, 
            'count': len(terminals_list),
            'request_ip': request.remote_addr,
            'server_ips': server_ips
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pos/terminals/<path:terminal_id>', methods=['DELETE'])
def delete_pos_terminal(terminal_id):
    """Remove a terminal from active list."""
    try:
        if terminal_id in ACTIVE_POS_TERMINALS:
            ACTIVE_POS_TERMINALS.pop(terminal_id, None)
            return jsonify({'status': 'ok', 'message': f'Terminal {terminal_id} deleted'})
        return jsonify({'status': 'not_found', 'message': 'Terminal not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pos/terminals/clear', methods=['POST', 'DELETE'])
def clear_all_pos_terminals():
    """Clear all active POS workstations."""
    try:
        ACTIVE_POS_TERMINALS.clear()
        return jsonify({'status': 'ok', 'message': 'All terminals cleared'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/weather', methods=['GET'])
def get_weather():
    import requests
    lat = request.args.get('latitude')
    lon = request.args.get('longitude')
    city = request.args.get('city')
    
    if not lat or not lon:
        geocoded = False
        if city and city.strip() and city.lower() != 'vị trí của tôi':
            try:
                geocode_url = "https://geocoding-api.open-meteo.com/v1/search"
                g_params = {"name": city.strip(), "count": 1, "language": "vi", "format": "json"}
                geores = requests.get(geocode_url, params=g_params, timeout=4).json()
                results = geores.get('results', [])
                if results:
                    lat = results[0].get('latitude')
                    lon = results[0].get('longitude')
                    city = results[0].get('name')
                    geocoded = True
            except Exception as geo_err:
                app.logger.error(f"Geocoding failed for {city}: {geo_err}")
        
        if not geocoded:
            # Get coordinates via IP from backend host
            try:
                r = requests.get('https://freeipapi.com/api/json', timeout=4)
                data = r.json()
                lat = data.get('latitude')
                lon = data.get('longitude')
                city = data.get('cityName')
            except Exception:
                try:
                    r = requests.get('https://ipapi.co/json/', timeout=4)
                    data = r.json()
                    lat = data.get('latitude')
                    lon = data.get('longitude')
                    city = data.get('city')
                except Exception:
                    # Default fallback
                    lat = 21.0285
                    lon = 105.8542
                    city = 'Hà Nội'
                
    try:
        weather_url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&current_weather=true"
        r = requests.get(weather_url, timeout=5)
        weather_data = r.json()
        current = weather_data.get('current_weather', {})
        temp = round(current.get('temperature', 28))
        weathercode = current.get('weathercode', 0)
        
        desc = 'Trời quang'
        if weathercode <= 3: desc = 'Ít mây'
        elif weathercode <= 48: desc = 'Sương mù'
        elif weathercode <= 67: desc = 'Mưa nhẹ'
        elif weathercode <= 82: desc = 'Mưa rào'
        elif weathercode <= 99: desc = 'Dông sét'
        
        return jsonify({
            'status': 'success',
            'latitude': lat,
            'longitude': lon,
            'city': city or 'Vị trí của tôi',
            'temp': temp,
            'desc': desc,
            'weathercode': weathercode
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

os.makedirs(get_storage_path("instance"), exist_ok=True)
db.init_app(app)
with app.app_context():
    is_pythonanywhere = 'pythonanywhere' in os.environ.get('SERVER_SOFTWARE', '').lower() or os.path.exists('/home/0848189789')
    if not is_pythonanywhere:
        db.create_all()
        # --- New Unified Auto-Migrations ---
        try:
            ensure_schema(db.engine)
            initialize_stock_batches(db.engine)
        except Exception as e:
            app.logger.error(f"Startup migration error: {e}")


import time
import threading
LAST_HEARTBEAT = time.time()

# --- Packing Display State ---
PACKING_DISPLAY_STATE = {
    "state_id": int(time.time() * 1000),
    "type": "CLEAR",
    "orders": [],
    "heldInvoices": []
}

@app.route('/api/packing/sync', methods=['GET', 'POST'])
def packing_sync():
    global PACKING_DISPLAY_STATE
    if request.method == 'POST':
        data = request.json
        if data:
            import time
            # Preserve existing fields if they are not in the new payload
            new_state = PACKING_DISPLAY_STATE.copy()
            new_state.update(data)
            new_state['state_id'] = int(time.time() * 1000)
            PACKING_DISPLAY_STATE = new_state
            
            # Sync active cart items to ACTIVE_POS_TERMINALS for PosMirror monitoring
            try:
                orders = data.get('orders') or []
                terminal_id = request.headers.get('X-Terminal-Id') or request.remote_addr or 'POS-DEFAULT'
                if orders and len(orders) > 0:
                    order = orders[0]
                    items = order.get('items') or []
                    partner_name = order.get('customer_name') or 'Khách lẻ'
                    cart = []
                    for i in items:
                        qty = float(i.get('quantity') or 1)
                        price = float(i.get('price') or i.get('sale_price') or 0)
                        cost = float(i.get('cost_price') or i.get('capital_price') or (price * 0.75))
                        cart.append({
                            'id': i.get('id'),
                            'product_id': i.get('product_id'),
                            'name': i.get('name') or i.get('product_name') or '',
                            'product_name': i.get('name') or i.get('product_name') or '',
                            'quantity': qty,
                            'unit': i.get('unit') or i.get('product_unit') or 'Cái',
                            'price': price,
                            'sale_price': price,
                            'cost_price': cost,
                            'code': i.get('code') or i.get('sku') or '',
                            'secondary_unit': i.get('secondary_unit'),
                            'secondary_qty': i.get('secondary_qty'),
                            'multiplier': i.get('multiplier')
                        })
                    
                    total_items = sum(int(c['quantity']) for c in cart)
                    total_amount = sum(int(c['quantity']) * float(c['price']) for c in cart)
                    now = datetime.now()
                    
                    existing = ACTIVE_POS_TERMINALS.get(terminal_id, {})
                    ACTIVE_POS_TERMINALS[terminal_id] = {
                        'terminal_id': terminal_id,
                        'terminal_name': existing.get('terminal_name') or f"Máy POS ({request.remote_addr})",
                        'user_name': existing.get('user_name') or 'Thu ngân',
                        'ip_address': request.remote_addr,
                        'last_active': now.isoformat(),
                        'cart': cart,
                        'partner_name': partner_name,
                        'total_amount': total_amount,
                        'total_items': total_items,
                        'status': 'active',
                        'current_page': existing.get('current_page') or '/pos'
                    }
                elif data.get('type') == 'CLEAR':
                    if terminal_id in ACTIVE_POS_TERMINALS:
                        ACTIVE_POS_TERMINALS[terminal_id]['cart'] = []
                        ACTIVE_POS_TERMINALS[terminal_id]['total_amount'] = 0
                        ACTIVE_POS_TERMINALS[terminal_id]['total_items'] = 0
            except Exception as sync_err:
                app.logger.error(f"Error updating POS terminal state in packing_sync: {sync_err}")

        return jsonify({"status": "ok", "state_id": PACKING_DISPLAY_STATE["state_id"]})
    
    return jsonify(PACKING_DISPLAY_STATE)

@app.route('/api/open-external-chrome', methods=['POST'])
def open_external_chrome():
    data = request.json or {}
    url = data.get('url')
    if not url:
        return jsonify({'error': 'URL is required'}), 400
        
    chrome_paths = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe")
    ]
    opened = False
    for path in chrome_paths:
        if os.path.exists(path):
            import subprocess
            app.logger.info(f"API request: Found Chrome at {path}, opening URL in app mode: {url}")
            subprocess.Popen([path, f"--app={url}", "--start-maximized"])
            opened = True
            break
            
    if opened:
        return jsonify({'status': 'ok', 'message': 'Chrome opened in app mode'})
    else:
        # Fallback to webbrowser
        import webbrowser
        webbrowser.open(url)
        return jsonify({'status': 'fallback', 'message': 'Chrome not found, opened in default browser'})

LAST_HEARTBEAT = time.time()

@app.route('/api/heartbeat', methods=['POST'])
def heartbeat():
    global LAST_HEARTBEAT
    LAST_HEARTBEAT = time.time()
    return jsonify({'status': 'ok'})

def monitor_heartbeat():
    while True:
        time.sleep(5)
        # If no heartbeat for more than 90 seconds, exit
        if time.time() - LAST_HEARTBEAT > 90:
            app.logger.info("No heartbeat for 90s. Shutting down...")
            os._exit(0)

# Only start monitor if frozen (production) - Disabled to prevent backend from auto-shutting down
# if getattr(sys, 'frozen', False):
#     import threading
#     heartbeat_thread = threading.Thread(target=monitor_heartbeat)
#     heartbeat_thread.daemon = True
#     heartbeat_thread.start()

import shutil

def backup_monitor():
    db_path = get_storage_path(os.path.join("instance", "easypos.db"))
    backup_dir = get_storage_path("backups")
    os.makedirs(backup_dir, exist_ok=True)
    
    last_mtime = None
    if os.path.exists(db_path):
        last_mtime = os.path.getmtime(db_path)
    
    app.logger.info("Backup service started.")
    
    while True:
        time.sleep(300) # 5 minutes
        if os.path.exists(db_path):
            current_mtime = os.path.getmtime(db_path)
            if last_mtime is None or current_mtime > last_mtime:
                from datetime import datetime
                time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_filename = f"lyangpos_backup_{time_str}.db"
                backup_path = os.path.join(backup_dir, backup_filename)
                try:
                    shutil.copy2(db_path, backup_path)
                    app.logger.info(f"Auto-backup created: {backup_filename}")
                    last_mtime = current_mtime
                    
                    # Rolling backup: Keep only 3 latest files
                    import glob
                    backup_files = glob.glob(os.path.join(backup_dir, "lyangpos_backup_*.db"))
                    backup_files.sort(key=os.path.getmtime)
                    if len(backup_files) > 3:
                        files_to_delete = backup_files[:-3]
                        for f in files_to_delete:
                            try:
                                os.remove(f)
                                app.logger.info(f"Deleted old backup: {os.path.basename(f)}")
                            except Exception as e:
                                app.logger.error(f"Failed to delete old backup {f}: {e}")
                                
                except Exception as e:
                    app.logger.error(f"Auto-backup failed: {e}")

backup_thread = threading.Thread(target=backup_monitor)
backup_thread.daemon = True
backup_thread.start()

try:
    import tkinter as tk
    from tkinter import ttk
    from PIL import Image, ImageTk, ImageOps, ImageFilter
except ImportError:
    tk = None

class SplashScreen:
    def __init__(self, port):
        self.port = port
        
        # If no GUI available, still monitor server to open browser
        if os.environ.get('NO_GUI') or os.environ.get('HEADLESS') or tk is None:
            app.logger.warning("GUI components missing or disabled. Browser will open in background.")
            threading.Thread(target=self.wait_for_server_headless, daemon=True).start()
            self.root = None
            return

        try:
            self.root = tk.Tk()
        except Exception as e:
            app.logger.error(f"Failed to initialize Tk: {e}")
            self.root = None
            threading.Thread(target=self.wait_for_server_headless, daemon=True).start()
            return

        # Properties
        width, height = 550, 400
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"{width}x{height}+{(sw-width)//2}+{(sh-height)//2}")
        self.root.overrideredirect(True)
        self.root.attributes("-topmost", True)
        
        # Colors: Luxury/Agri Blend
        self.c_bg = "#ffffff"
        self.c_bg_sub = "#faf8f3"
        self.c_primary = "#2d5016"
        self.c_accent = "#d4a574"
        self.c_text_muted = "#888888"

        self.canvas = tk.Canvas(self.root, width=width, height=height, bg=self.c_bg, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        # Draw a subtle background gradient manually
        for i in range(height):
            ratio = i / height
            r = int(255 * (1-ratio) + 250 * ratio)
            g = int(255 * (1-ratio) + 248 * ratio)
            b = int(255 * (1-ratio) + 243 * ratio)
            self.canvas.create_line(0, i, width, i, fill=f'#{r:02x}{g:02x}{b:02x}')

        # Load & Process Logo
        try:
            logo_path = resource_path('frontend/src/assets/logo.png')
            if not os.path.exists(logo_path):
                logo_path = resource_path('logo.png')
            
            self.raw_img = Image.open(logo_path).convert("RGBA")
            self.raw_img.thumbnail((220, 220), Image.Resampling.LANCZOS)
            
            # Simple soft shadow
            self.canvas.create_oval(width//2-70, height//2-40, width//2+70, height//2-10, fill="#eeeeee", outline="", tags="shadow")
            
            self.logo_tk = ImageTk.PhotoImage(self.raw_img)
            self.logo_id = self.canvas.create_image(width//2, height//2 - 50, image=self.logo_tk)
        except Exception as e:
            print(f"Logo Error: {e}")
            self.canvas.create_text(width//2, height//2 - 50, text="LyangPOS", font=("Segoe UI", 36, "bold"), fill=self.c_primary)

        # Labels
        self.canvas.create_text(width//2, height//2 + 90, text="LyangPOS v3.0", font=("Segoe UI", 16, "bold"), fill=self.c_primary)
        self.canvas.create_text(width//2, height//2 + 115, text="PHẦN MỀM QUẢN LÝ VỤ MÙA THÔNG MINH", font=("Segoe UI", 8, "bold"), fill=self.c_accent)
        
        self.status_text = self.canvas.create_text(width//2, height - 35, text="Đang đồng bộ hóa dữ liệu...", font=("Segoe UI", 9), fill=self.c_text_muted)

        # Progress System
        self.bar_w = 300
        self.bar_x = (width - self.bar_w) // 2
        self.bar_y = height - 60
        self.canvas.create_rectangle(self.bar_x, self.bar_y, self.bar_x + self.bar_w, self.bar_y + 3, fill="#f0eeee", outline="")
        self.progress_bar = self.canvas.create_rectangle(self.bar_x, self.bar_y, self.bar_x, self.bar_y + 3, fill=self.c_primary, outline="")
        
        # State
        self.p_val = 0
        self.animate()
        self.root.after(1000, self.check_server)

    def animate(self):
        self.p_val = (self.p_val + 1) % 100
        cw = (self.p_val / 100) * self.bar_w
        self.canvas.coords(self.progress_bar, self.bar_x, self.bar_y, self.bar_x + cw, self.bar_y + 3)
        
        # Subtle status text pulse
        if self.p_val % 20 == 0:
            current_fill = self.canvas.itemcget(self.status_text, "fill")
            next_fill = "#aaaaaa" if current_fill == self.c_text_muted else self.c_text_muted
            self.canvas.itemconfig(self.status_text, fill=next_fill)

        self.root.after(30, self.animate)

    def check_server(self):
        import socket
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(0.3)
                if s.connect_ex(('127.0.0.1', self.port)) == 0:
                    self.canvas.itemconfig(self.status_text, text="Đã sẵn sàng! Đang khởi động trình duyệt...", fill=self.c_primary)
                    # Increased delay slightly for better transition feel
                    self.root.after(1000, self.open_browser)
                    return
        except:
            pass
        self.root.after(1000, self.check_server)

    def wait_for_server_headless(self):
        """Wait for server and open browser without Tkinter mainloop"""
        import socket
        import time
        app.logger.info(f"Waiting for server on port {self.port} (Headless mode)...")
        # Max wait 30 seconds
        for _ in range(30):
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.settimeout(1.0)
                    if s.connect_ex(('127.0.0.1', self.port)) == 0:
                        app.logger.info("Server is up! Opening browser...")
                        time.sleep(1) # Extra buffer
                        self.open_browser_cmd()
                        return
            except:
                pass
            time.sleep(1)
        app.logger.error("Server took too long to start, browser launch aborted.")

    def open_browser(self):
        self.open_browser_cmd()
        # Destroy splash after browser is triggered
        if self.root:
            self.root.after(1000, self.root.destroy)

    def open_browser_cmd(self):
        import webbrowser
        target_url = f"http://localhost:{self.port}"
        app.logger.info(f"Triggering browser for {target_url}")
        try:
            # Try to open in Chrome App mode for cleaner POS look
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe")
            ]
            opened = False
            for path in chrome_paths:
                if os.path.exists(path):
                    import subprocess
                    app.logger.info(f"Found Chrome at {path}, opening in app mode.")
                    subprocess.Popen([path, f"--app={target_url}", "--start-maximized"])
                    opened = True
                    break
            
            if not opened:
                app.logger.info("Chrome not found or failed, using default browser.")
                webbrowser.open(target_url)
        except Exception as e:
            app.logger.error(f"Browser launch error: {e}")
            webbrowser.open(target_url)

    def run(self):
        if self.root:
            self.root.mainloop()
        else:
            # If no GUI, just keep the main thread alive for a bit if needed
            # though the headless thread handles browser opening
            pass

def run_migrations():
    """Obsolete. Logic moved to migrations_manager.py and called during startup."""
    pass

# Run initial migration
run_migrations()

# --- Auth Routes ---
@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    display_name = data.get('display_name', username)
    
    # Role Logic: First user is always admin. Subsequent users need the secret to be admin.
    admin_secret = data.get('admin_secret')
    is_first_user = User.query.count() == 0
    
    if is_first_user:
        role = 'admin'
    elif admin_secret:
        if admin_secret == '0607@Nghia':
            role = 'admin'
        else:
            app.logger.warning(f"Failed admin registration attempt for user '{username}': Wrong secret.")
            return jsonify({'error': 'Mã bí mật cấp Admin không chính xác. Vui lòng kiểm tra lại!'}), 403
    else:
        # Default or requested role, but force to 'user' if they requested 'admin' without secret
        requested_role = data.get('role', 'user').lower()
        role = 'user' if requested_role == 'admin' else requested_role

    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400

    if User.query.filter_by(username=username).first():
        app.logger.info(f"Registration failed: User '{username}' already exists.")
        return jsonify({'error': 'Tài khoản này đã tồn tại trên hệ thống.'}), 400

    new_user = User(
        username=username,
        password_hash=generate_password_hash(password),
        display_name=display_name,
        role=role
    )
    db.session.add(new_user)
    db.session.commit()

    return jsonify({'message': 'User registered successfully', 'user': new_user.to_dict()}), 201

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')

    user = User.query.filter_by(username=username).first()

    if user and check_password_hash(user.password_hash, password):
        return jsonify({
            'message': 'Login successful',
            'user': user.to_dict()
        }), 200
    
    return jsonify({'error': 'Invalid username or password'}), 401

@app.route('/api/users', methods=['GET'])
def get_users():
    users = User.query.all()
    return jsonify([u.to_dict() for u in users])

@app.route('/api/users/<int:user_id>', methods=['PATCH'])
def update_user(user_id):
    user = User.query.get_or_404(user_id)
    data = request.json
    
    if 'role' in data:
        user.role = data['role']
    if 'display_name' in data:
        user.display_name = data['display_name']
    if 'password' in data and data['password']:
        user.password_hash = generate_password_hash(data['password'])
        
    db.session.commit()
    return jsonify(user.to_dict())

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    # Don't allow deleting the last admin or yourself? 
    # For now, just a simple delete.
    db.session.delete(user)
    db.session.commit()
    return jsonify({'message': 'User deleted successfully'})


# --- Font Management ---
# Ensure storage folders are outside of PyInstaller's temp directory

FONTS_FOLDER = get_storage_path(os.path.join("uploads", "fonts"))
os.makedirs(FONTS_FOLDER, exist_ok=True)

LOGO_FOLDER = get_storage_path(os.path.join("uploads", "logos"))
os.makedirs(LOGO_FOLDER, exist_ok=True)

@app.route('/uploads/fonts/<path:filename>')
def serve_font(filename):
    return send_from_directory(FONTS_FOLDER, filename)

@app.route('/uploads/logos/<path:filename>')
def serve_logo(filename):
    return send_from_directory(LOGO_FOLDER, filename)

@app.route('/api/fonts', methods=['GET'])
def list_fonts():
    if not os.path.exists(FONTS_FOLDER): return jsonify([])
    fonts = [f for f in os.listdir(FONTS_FOLDER) if f.lower().endswith(('.ttf', '.otf', '.woff', '.woff2'))]
    return jsonify(fonts)

@app.route('/api/fonts', methods=['POST'])
def upload_font():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file:
        filename = file.filename
        file.save(os.path.join(FONTS_FOLDER, filename))
        return jsonify({'message': 'Font uploaded successfully', 'filename': filename})

@app.route('/api/upload-logo', methods=['POST'])
def upload_logo():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file:
        import time
        filename = f"{int(time.time())}_{file.filename}"
        file.save(os.path.join(LOGO_FOLDER, filename))
        return jsonify({'url': f'/uploads/logos/{filename}'})

# --- Categories ---
@app.route('/api/categories', methods=['GET'])
def get_categories():
    categories = Category.query.all()
    return jsonify([c.to_dict() for c in categories])

@app.route('/api/categories', methods=['POST'])
def create_category():
    data = request.json
    if not data.get('name'):
        return jsonify({'error': 'Tên danh mục không được để trống'}), 400
    
    new_cat = Category(
        name=data['name'],
        icon=data.get('icon', 'Package')
    )
    db.session.add(new_cat)
    db.session.commit()
    return jsonify(new_cat.to_dict()), 201

@app.route('/api/categories/<int:cat_id>', methods=['PUT'])
def update_category(cat_id):
    cat = Category.query.get_or_404(cat_id)
    data = request.json
    if 'name' in data:
        cat.name = data['name']
    if 'icon' in data:
        cat.icon = data['icon']
    db.session.commit()
    return jsonify(cat.to_dict())

@app.route('/api/categories/<int:cat_id>', methods=['DELETE'])
def delete_category(cat_id):
    cat = Category.query.get_or_404(cat_id)
    # Kiểm tra xem có sản phẩm nào đang dùng danh mục này không
    product_count = Product.query.filter_by(category_id=cat_id).count()
    if product_count > 0:
        return jsonify({'error': f'Không thể xóa danh mục này vì đang có {product_count} sản phẩm thuộc loại này.'}), 400
    
    db.session.delete(cat)
    db.session.commit()
    return jsonify({'message': 'Đã xóa danh mục thành công'})

# --- Products ---
@app.route('/api/products/summary', methods=['GET'])
def get_product_summary():
    try:
        # Total active products
        total_products = Product.query.filter_by(is_active=True).count()
        
        # Out of stock
        out_of_stock = Product.query.filter(Product.is_active == True, Product.stock <= 0).count()
        
        # Warning (Low stock: e.g. stock <= min_stock or multiplier)
        low_stock = Product.query.filter(
            Product.is_active == True, 
            Product.stock > 0, 
            Product.stock <= db.case((Product.min_stock > 0, Product.min_stock), else_=db.func.coalesce(Product.multiplier, 1))
        ).count()
        
        # Near expiry (within 30 days)
        # Dates are stored as "DD/MM/YYYY" strings or YYYY-MM-DD
        from datetime import datetime, timedelta
        near_expiry_count = 0
        today = get_vn_time()
        thirty_days_later = today + timedelta(days=30)
        
        # This is a bit slow for many products but okay for now
        all_prods = Product.query.filter(Product.is_active == True, Product.expiry_date != None, Product.expiry_date != '').all()
        for p in all_prods:
            try:
                # Try multiple formats
                ep_date = None
                for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                    try:
                        ep_date = datetime.strptime(p.expiry_date, fmt)
                        break
                    except ValueError: continue
                
                if ep_date and ep_date <= thirty_days_later and ep_date >= (today - timedelta(days=365)): # Exclude very old
                    near_expiry_count += 1
            except: continue

        return jsonify({
            'total_products': total_products,
            'out_of_stock': out_of_stock,
            'low_stock': low_stock,
            'near_expiry': near_expiry_count
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/products', methods=['GET'])
def get_products():
    search = request.args.get('search', '').lower()
    filter_type = request.args.get('filterType', 'all')
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)
    
    # Critical Fix: Optimize N+1 query for combo_items & batches
    query = Product.query.options(joinedload(Product.combo_items), joinedload(Product.batches))
    
    # Inactive filtering: 
    # Default: Show only active
    # if filterType='inactive': show only inactive
    # if include_inactive=true: show all
    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
    
    if filter_type == 'inactive':
        query = query.filter(Product.is_active == False)
    elif not include_inactive:
        # Default behavior: POS and generic lookups should be active only
        query = query.filter(Product.is_active == True)
    
    if search:
        s_norm = remove_accents(search)
        query = query.filter(
            db.func.remove_accents(Product.name).ilike(f'%{s_norm}%') | 
            Product.code.ilike(f'%{search}%') |
            db.func.remove_accents(db.func.coalesce(Product.active_ingredient, '')).ilike(f'%{s_norm}%') |
            db.func.remove_accents(db.func.coalesce(Product.brand, '')).ilike(f'%{s_norm}%')
        )
    
    brand = request.args.get('brand')
    if brand:
        query = query.filter(Product.brand == brand)
    
    category_id = request.args.get('category_id')
    if category_id:
        query = query.filter(Product.category_id == category_id)
    
    if filter_type == 'out_of_stock':
        query = query.filter(Product.stock <= 0)
    elif filter_type == 'warning':
        query = query.filter(
            Product.is_active == True, 
            Product.stock > 0, 
            Product.stock <= db.case((Product.min_stock > 0, Product.min_stock), else_=db.func.coalesce(Product.multiplier, 1))
        )
    elif filter_type == 'expired':
        today = datetime.now().strftime('%Y-%m-%d')
        query = query.filter(Product.expiry_date != None, Product.expiry_date != '', db.func.normalize_date(Product.expiry_date) <= today)
    elif filter_type == 'near_expiry':
        from datetime import timedelta
        today = datetime.now()
        target_date = (today + timedelta(days=60)).strftime('%Y-%m-%d')
        today_str = today.strftime('%Y-%m-%d')
        query = query.filter(
            Product.expiry_date != None, 
            Product.expiry_date != '', 
            db.func.normalize_date(Product.expiry_date) > today_str,
            db.func.normalize_date(Product.expiry_date) <= target_date
        )
    elif filter_type == 'loss':
        query = query.filter(Product.sale_price < Product.cost_price, Product.is_combo == False)
    elif filter_type == 'unused':
        # Truly unused products: No sales, no audits, not in combos, AND stock is 0
        from models import OrderDetail, InventoryAuditDetail, ComboItem
        # Subqueries with explicit NULL filtering for safety with NOT IN
        used_in_orders = db.session.query(OrderDetail.product_id).filter(OrderDetail.product_id.isnot(None))
        used_in_audits = db.session.query(InventoryAuditDetail.product_id).filter(InventoryAuditDetail.product_id.isnot(None))
        used_as_combo_ingredient = db.session.query(ComboItem.product_id).filter(ComboItem.product_id.isnot(None))
        is_a_combo_with_items = db.session.query(ComboItem.combo_id).filter(ComboItem.combo_id.isnot(None))
        
        query = query.filter(
            ~Product.id.in_(used_in_orders),
            ~Product.id.in_(used_in_audits),
            ~Product.id.in_(used_as_combo_ingredient),
            ~Product.id.in_(is_a_combo_with_items),
            Product.stock == 0
        )

    sort_by = request.args.get('sort_by', 'name')
    sort_order = request.args.get('sort_order', 'asc')

    # Mapping frontend sort keys to DB columns
    sort_map = {
        'id': Product.id,
        'name': Product.name,
        'code': Product.code,
        'unit': Product.unit,
        'cost_price': Product.cost_price,
        'sale_price': Product.sale_price,
        'stock': Product.stock,
        'expiry_date': Product.expiry_date
    }
    
    sort_col = sort_map.get(sort_by, Product.name)
    if sort_order == 'desc':
        query = query.order_by(sort_col.desc())
    else:
        query = query.order_by(sort_col.asc())

    if page and limit:
        pagination = query.paginate(page=page, per_page=limit, error_out=False)
        products = pagination.items
        total = pagination.total
        pages = pagination.pages
        current_page = pagination.page
    else:
        products = query.all()
        total = len(products)
        pages = 1
        current_page = 1

    results = []
    for p in products:
        d = p.to_dict()
        if p.is_combo:
            d['combo_items'] = [i.to_dict() for i in p.combo_items]
        results.append(d)
        
    if page and limit:
        return jsonify({
            'items': results,
            'total': total,
            'pages': pages,
            'current_page': current_page
        })
    else:
        return jsonify(results)

@app.route('/api/products', methods=['POST'])
def create_product():
    data = request.json
    cat_id = data.get('category_id')
    if cat_id == '' or cat_id == 0:
        cat_id = None

    new_prod = Product(
        name=data['name'],
        code=data.get('code'),
        unit=normalize_uom(data.get('unit')),
        secondary_unit=normalize_uom(data.get('secondary_unit')),
        multiplier=safe_float(data.get('multiplier'), 1),
        cost_price=safe_float(data.get('cost_price'), 0),
        sale_price=safe_float(data.get('sale_price'), 0),
        stock=safe_float(data.get('stock'), 0),
        expiry_date=data.get('expiry_date'),
        active_ingredient=data.get('active_ingredient'),
        brand=data.get('brand'),
        is_combo=data.get('is_combo', False),
        is_active=data.get('is_active', True),
        category_id=cat_id,
        accounting_price=safe_float(data.get('accounting_price'), 0),
        accounting_stock=safe_float(data.get('accounting_stock'), 0),
        bulk_quantity=safe_float(data.get('bulk_quantity')),
        bulk_price=safe_float(data.get('bulk_price')),
        alias=data.get('alias'),
        min_stock=safe_float(data.get('min_stock'), 0)
    )
    db.session.add(new_prod)
    db.session.flush() # Get ID

    # Create a StockBatch if initial stock > 0 for FIFO tracking
    if new_prod.stock > 0:
        batch = StockBatch(
            product_id=new_prod.id,
            original_quantity=new_prod.stock,
            current_quantity=new_prod.stock,
            cost_price=new_prod.cost_price or 0,
            created_at=get_vn_time()
        )
        db.session.add(batch)

    if data.get('is_combo') and 'combo_items' in data:
        for item in data['combo_items']:
            ci = ComboItem(
                combo_id=new_prod.id,
                product_id=item['product_id'],
                quantity=item['quantity']
            )
            db.session.add(ci)

    db.session.commit()
    return jsonify(new_prod.to_dict()), 201

@app.route('/api/products/<int:id>', methods=['PUT'])
def update_product(id):
    prod = Product.query.get_or_404(id)
    data = request.json
    prod.name = data.get('name', prod.name)
    prod.code = data.get('code', prod.code)
    prod.unit = normalize_uom(data.get('unit', prod.unit))
    prod.secondary_unit = normalize_uom(data.get('secondary_unit', prod.secondary_unit))
    prod.multiplier = safe_float(data.get('multiplier'), prod.multiplier)
    prod.cost_price = safe_float(data.get('cost_price'), prod.cost_price)
    prod.sale_price = safe_float(data.get('sale_price'), prod.sale_price)
    prod.stock = safe_float(data.get('stock'), prod.stock)
    prod.expiry_date = data.get('expiry_date', prod.expiry_date)
    prod.active_ingredient = data.get('active_ingredient', prod.active_ingredient)
    prod.brand = data.get('brand', prod.brand)
    prod.is_combo = data.get('is_combo', prod.is_combo)
    prod.is_active = data.get('is_active', prod.is_active)
    prod.alias = data.get('alias', prod.alias)
    if 'min_stock' in data:
        prod.min_stock = safe_float(data['min_stock'], prod.min_stock or 0)
    
    cat_id = data.get('category_id', prod.category_id)
    if cat_id == '' or cat_id == 0:
        cat_id = None
    prod.category_id = cat_id
    
    prod.accounting_price = safe_float(data.get('accounting_price'), prod.accounting_price)
    prod.accounting_stock = safe_float(data.get('accounting_stock'), prod.accounting_stock)
    
    if 'bulk_quantity' in data:
        prod.bulk_quantity = safe_float(data['bulk_quantity'])
    if 'bulk_price' in data:
        prod.bulk_price = safe_float(data['bulk_price'])
    
    if 'combo_items' in data:
        # Update combo items
        ComboItem.query.filter_by(combo_id=prod.id).delete()
        for item in data['combo_items']:
            ci = ComboItem(
                combo_id=prod.id,
                product_id=item['product_id'],
                quantity=item['quantity']
            )
            db.session.add(ci)

    db.session.commit()
    # Include items in response if combo
    res = prod.to_dict()
    if prod.is_combo:
        res['combo_items'] = [i.to_dict() for i in prod.combo_items]
    return jsonify(res)

@app.route('/api/products/bulk-update', methods=['POST'])
def bulk_update_products():
    data = request.json
    
    # Mode 1: Individual updates (List of objects with ID and specific fields)
    individual_updates = data.get('individual_updates', [])
    if individual_updates:
        updated_count = 0
        for item in individual_updates:
            p_id = item.get('id')
            if not p_id: continue
            
            prod = Product.query.get(p_id)
            if not prod: continue
            
            # Update fields if present in item
            if 'name' in item: prod.name = item['name']
            if 'code' in item: prod.code = item['code']
            if 'unit' in item: prod.unit = normalize_uom(item['unit'])
            if 'secondary_unit' in item: prod.secondary_unit = normalize_uom(item['secondary_unit'])
            if 'multiplier' in item: prod.multiplier = item['multiplier']
            if 'cost_price' in item: prod.cost_price = item['cost_price']
            if 'sale_price' in item: prod.sale_price = item['sale_price']
            if 'stock' in item: prod.stock = item['stock']
            if 'expiry_date' in item: prod.expiry_date = item['expiry_date']
            if 'active_ingredient' in item: prod.active_ingredient = item['active_ingredient']
            if 'brand' in item: prod.brand = item['brand']
            if 'category_id' in item: prod.category_id = item['category_id']
            if 'is_active' in item: prod.is_active = item['is_active']
            if 'accounting_price' in item: prod.accounting_price = item['accounting_price']
            if 'accounting_stock' in item: prod.accounting_stock = item['accounting_stock']
            if 'min_stock' in item: prod.min_stock = safe_float(item['min_stock'], 0)
            
            updated_count += 1
            
        db.session.commit()
        return jsonify({'message': f'Đã cập nhật thành công {updated_count} sản phẩm'})

    # Mode 2: Bulk updates (Same updates for multiple products)
    product_ids = data.get('product_ids', [])
    updates = data.get('updates', {}) 
    sku_map = data.get('sku_map', {}) # { product_id: new_code }
    
    if not product_ids:
        return jsonify({'error': 'Không có sản phẩm nào được chọn'}), 400
        
    products = Product.query.filter(Product.id.in_(product_ids)).all()
    
    # Hỗ trợ sinh mã tự động nếu có prefix
    auto_prefix = data.get('auto_sku_prefix')
    counter = data.get('auto_sku_start', 1)
    
    for prod in products:
        # 1. Cập nhật nhóm/trạng thái chung
        if 'category_id' in updates:
            prod.category_id = updates['category_id']
        if 'is_active' in updates:
            prod.is_active = updates['is_active']
            
        # 2. Cập nhật mã hàng hóa (ưu tiên theo thứ tự: sku_map > auto_prefix > updates['code'])
        str_id = str(prod.id)
        if str_id in sku_map:
            prod.code = sku_map[str_id]
        elif auto_prefix:
            # Sinh mã kiểu T-001, T-002...
            prod.code = f"{auto_prefix}{str(counter).zfill(3)}"
            counter += 1
        elif 'code' in updates:
            prod.code = updates['code']
            
    db.session.commit()
    return jsonify({'message': f'Đã cập nhật thành công {len(products)} sản phẩm'})

@app.route('/api/products/bulk-accounting-update', methods=['POST'])
def bulk_accounting_update():
    data = request.json
    if not isinstance(data, list):
        return jsonify({'error': 'Dữ liệu không hợp lệ, yêu cầu một danh sách'}), 400
    
    try:
        # Sử dụng bulk_update_mappings để tối ưu hiệu năng
        # Mỗi item trong data cần có 'id', 'accounting_price', 'accounting_stock'
        db.session.bulk_update_mappings(Product, data)
        db.session.commit()
        return jsonify({
            'message': f'Đã cập nhật thành công {len(data)} sản phẩm',
            'updated_count': len(data)
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Bulk accounting update failed: {e}")
        return jsonify({'error': str(e)}), 400

@app.route('/api/products/<int:id>', methods=['DELETE'])
def delete_product(id):
    # Check if used in transactions
    in_use = OrderDetail.query.filter_by(product_id=id).first()
    if in_use:
        return jsonify({'error': 'Không thể xóa sản phẩm đã có lịch sử giao dịch'}), 400
    
    prod = Product.query.get_or_404(id)
    db.session.delete(prod)
    db.session.commit()
    return jsonify({'message': 'Deleted successfully'})

@app.route('/api/products/bulk-delete', methods=['POST'])
def bulk_delete_products():
    data = request.json
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'Không có sản phẩm nào được chọn'}), 400
    
    # Check if any ID is used in orders
    used_ids = [r[0] for r in db.session.query(OrderDetail.product_id).filter(OrderDetail.product_id.in_(ids)).all()]
    if used_ids:
        return jsonify({'error': f'Không thể xóa {len(used_ids)} sản phẩm đã phát sinh đơn hàng'}), 400
    
    # Check if used in combos (as ingredient)
    used_in_combos = [r[0] for r in db.session.query(ComboItem.product_id).filter(ComboItem.product_id.in_(ids)).all()]
    if used_in_combos:
        return jsonify({'error': f'Không thể xóa {len(used_in_combos)} sản phẩm đang được dùng trong combo'}), 400

    count = Product.query.filter(Product.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({'message': f'Đã xóa thành công {count} sản phẩm'})

@app.route('/api/products/brands', methods=['GET'])
def get_product_brands():
    brands = db.session.query(Product.brand).distinct().all()
    # Filter out None or empty strings and sort
    brand_list = sorted([b[0] for b in brands if b[0]])
    return jsonify(brand_list)



@app.route('/api/normalize-uom', methods=['POST'])
def normalize_uom_route():
    try:
        products = Product.query.all()
        updated_count = 0
        for p in products:
            changed = False
            new_unit = normalize_uom(p.unit)
            new_secondary = normalize_uom(p.secondary_unit)
            
            if new_unit != p.unit:
                p.unit = new_unit
                changed = True
                
            if new_secondary != p.secondary_unit:
                p.secondary_unit = new_secondary
                changed = True
            
            if changed:
                updated_count += 1
        
        if updated_count > 0:
            db.session.commit()
            
        return jsonify({
            'message': f'Chuẩn hóa hoàn tất! Đã cập nhật {updated_count} sản phẩm.',
            'updated_count': updated_count
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/products/import', methods=['POST'])
def import_products():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # Get headers
        headers = [str(cell.value).strip() for cell in ws[1]]
        
        def get_val(row, col_name, default=None):
            try:
                # First try exact match
                idx = headers.index(col_name)
                val = row[idx].value
                return val if val is not None else default
            except (ValueError, IndexError):
                # Try partial match (e.g. "Tồn kho" matches "Tồn kho (Đơn vị chính)")
                for i, h in enumerate(headers):
                    if col_name in h:
                        val = row[i].value
                        return val if val is not None else default
                return default

        count = 0
        # Iterate from row 2
        for row in ws.iter_rows(min_row=2):
            name = str(get_val(row, 'Tên sản phẩm', '')).strip()
            if not name or name == 'None' or name == '': continue
            
            prod = Product.query.filter_by(name=name).first()
            if not prod:
                prod = Product(name=name, unit=str(get_val(row, 'Đơn vị', 'Cái')))
                db.session.add(prod)
                db.session.flush() # Flush to get prod.id for StockBatch
            
            prod.unit = str(get_val(row, 'Đơn vị', prod.unit))
            sec_unit = get_val(row, 'Đơn vị phụ')
            if sec_unit is not None: prod.secondary_unit = str(sec_unit)
            
            multiplier = get_val(row, 'Quy cách')
            if multiplier is not None: prod.multiplier = float(multiplier)
            
            cost_price = get_val(row, 'Giá vốn')
            if cost_price is not None:
                prod.cost_price = float(cost_price)
                prod.latest_cost_price = float(cost_price)
            
            sale_price = get_val(row, 'Giá bán')
            if sale_price is not None: prod.sale_price = float(sale_price)
            
            stock = get_val(row, 'Tồn kho')
            if stock is not None:
                new_stock = int(float(stock))
                # If stock changed, create an adjustment batch for FIFO tracking and visibility
                existing_batch = StockBatch.query.filter_by(product_id=prod.id).order_by(StockBatch.created_at.desc()).first()
                if not existing_batch or new_stock != prod.stock:
                    # Create a batch for the difference or initial stock
                    # We treat the imported stock as a single batch
                    batch = StockBatch(
                        product_id=prod.id,
                        original_quantity=new_stock,
                        current_quantity=new_stock,
                        cost_price=prod.cost_price or 0,
                        created_at=get_vn_time()
                    )
                    db.session.add(batch)
                else:
                    # If the stock didn't change (e.g. both are 0), update the batch's cost price to match the newly imported price
                    existing_batch.cost_price = prod.cost_price or 0
                prod.stock = new_stock
            
            expiry = get_val(row, 'Hạn sử dụng')
            if expiry is not None: prod.expiry_date = str(expiry)
            
            ai = get_val(row, 'Hoạt chất')
            if ai is not None: prod.active_ingredient = str(ai)
            
            brand = get_val(row, 'Hãng')
            if brand is not None: prod.brand = str(brand)
            
            acc_price = get_val(row, 'Giá kế toán')
            if acc_price is not None: prod.accounting_price = float(acc_price)
            
            acc_stock = get_val(row, 'Tồn kế toán')
            if acc_stock is not None: prod.accounting_stock = float(acc_stock)
            
            is_active = get_val(row, 'Theo dõi')
            if is_active is not None:
                # Support "Có", "1", "True", "Yes"
                is_active_str = str(is_active).lower().strip()
                prod.is_active = is_active_str in ['1', 'true', 'yes', 'có', 'on']

            # New fields: Mã hàng and Loại sản phẩm
            code = get_val(row, 'Mã hàng')
            if code is not None: prod.code = str(code).strip()

            alias = get_val(row, 'Alias') or get_val(row, 'Tên alias')
            if alias is not None: prod.alias = str(alias).strip()
            
            cat_name = get_val(row, 'Loại sản phẩm')
            if cat_name:
                cat_name = str(cat_name).strip()
                cat = Category.query.filter_by(name=cat_name).first()
                if not cat:
                    cat = Category(name=cat_name)
                    db.session.add(cat)
                    db.session.flush() # Get ID
                prod.category_id = cat.id
            
            count += 1
            
        db.session.commit()
        return jsonify({'message': f'Imported {count} products successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/products/template', methods=['GET'])
def get_template():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    headers = [
        'Mã hàng', 'Tên sản phẩm', 'Loại sản phẩm', 'Hãng', 'Hoạt chất', 
        'Đơn vị', 'Đơn vị phụ', 'Quy cách (1 phụ = ? chính)', 
        'Giá vốn', 'Giá bán', 'Giá kế toán', 
        'Tồn kho (Đơn vị chính)', 'Tồn kế toán', 
        'Hạn sử dụng', 'Theo dõi (Có/Không)'
    ]
    ws.append(headers)
    ws.append(['SP001', 'Sản phẩm mẫu', 'Phân bón', 'Lyang Nghĩa', 'Abamectin', 'Chai', 'Thùng', 24, 100000, 150000, 110000, 10, 5, '31/12/2026', 'Có'])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='mau_nhap_san_pham.xlsx'
    )


# --- Product Orders ---
@app.route('/api/products/<int:id>/orders', methods=['GET'])
def get_product_orders(id):
    # Find all orders that contain this product
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    
    query = Order.query.join(OrderDetail).filter(OrderDetail.product_id == id)
    
    if year:
        query = query.filter(db.func.strftime('%Y', Order.date) == str(year))
    if month:
        query = query.filter(db.func.strftime('%m', Order.date) == str(month).zfill(2))
    if day:
        query = query.filter(db.func.strftime('%d', Order.date) == str(day).zfill(2))
        
    orders = query.order_by(Order.date.desc()).all()
    return jsonify([o.to_dict() for o in orders])

@app.route('/api/products/<int:id>/history', methods=['GET'])
def get_product_history(id):
    try:
        # Get all movements for this product
        # 1. From Order Details
        details = OrderDetail.query.filter_by(product_id=id).join(Order).order_by(Order.date.desc()).all()
        
        history = []
        for d in details:
            if not d.order: continue
            change_qty = 0
            type_label = ''
            
            if d.order.type == 'Sale':
                # Sale: negative quantity means customer returned items (stock increases)
                if (d.quantity or 0) < 0:
                    change_qty = -(d.quantity or 0)
                    type_label = 'Khách trả hàng'
                else:
                    change_qty = -(d.quantity or 0)
                    type_label = 'Bán hàng'
            elif d.order.type == 'Purchase':
                # Purchase: negative quantity means returned items to supplier (stock decreases)
                if (d.quantity or 0) < 0:
                    change_qty = d.quantity or 0
                    type_label = 'Trả hàng NCC'
                else:
                    change_qty = d.quantity or 0
                    type_label = 'Nhập hàng'
                
            # Handle Combos if necessary (Combos don't change stock directly? Wait, OrderDetail links to Product. 
            # If product is Combo, its stock decreases. But usually we track components.
            # If this request is for a Component Product, we must find Orders where this Component was part of a Combo.
            
            history.append({
                'date': d.order.date.isoformat(),
                'display_id': d.order.display_id,
                'order_id': d.order.id,
                'partner_name': d.order.partner.name if d.order.partner else ('Khách Lẻ' if d.order.type=='Sale' else 'NCC Vãng Lai'),
                'type': type_label,
                'quantity_change': change_qty,
                'price': d.price
            })

        # Also need to check if this product is part of any Combos that were sold
        # Find all combos that include this product
        parent_combos = ComboItem.query.filter_by(product_id=id).all()
        for ci in parent_combos:
            combo_prod_id = ci.combo_id
            # Find sales of this combo
            combo_details = OrderDetail.query.filter_by(product_id=combo_prod_id).join(Order).all()
            for cd in combo_details:
                qty_deducted = cd.quantity * ci.quantity # Total items used
                
                if cd.order.type == 'Sale':
                    type_label = 'Khách trả Combo' if (cd.quantity or 0) < 0 else 'Bán Combo'
                else:
                    type_label = 'Trả Combo NCC' if (cd.quantity or 0) < 0 else 'Nhập Combo'
                
                # If Sale, we lost stock
                change = -qty_deducted if cd.order.type == 'Sale' else qty_deducted
                
                history.append({
                    'date': cd.order.date.isoformat(),
                    'display_id': cd.order.display_id,
                    'order_id': cd.order.id,
                    'partner_name': cd.order.partner.name if cd.order.partner else 'Khách Lẻ',
                    'type': f"{type_label} ({cd.product.name})",
                    'quantity_change': change,
                    'price': 0 # Hard to attribute price
                })
        
        # Sort combined history by date desc
        history.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify(history)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# --- Combo Items ---
@app.route('/api/combos/<int:combo_id>/items', methods=['GET'])
def get_combo_items(combo_id):
    items = ComboItem.query.filter_by(combo_id=combo_id).all()
    return jsonify([i.to_dict() for i in items])

@app.route('/api/combos/<int:combo_id>/items', methods=['POST'])
def set_combo_items(combo_id):
    data = request.json # Expected list of {product_id, quantity}
    try:
        # Clear existing
        ComboItem.query.filter_by(combo_id=combo_id).delete()
        for item in data:
            new_item = ComboItem(
                combo_id=combo_id,
                product_id=item['product_id'],
                quantity=item['quantity']
            )
            db.session.add(new_item)
        db.session.commit()
        return jsonify({'message': 'Combo items updated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# --- Inventory Audits ---
@app.route('/api/inventory/products/search', methods=['GET'])
def search_inventory_products():
    search = request.args.get('search', '').lower()
    if not search:
        return jsonify([])
    
    s_norm = remove_accents(search)
    # Optimized search: ID, Name, Code, Unit, Stock
    products = Product.query.filter(
        (Product.is_active == True) & (
            db.func.remove_accents(Product.name).ilike(f'%{s_norm}%') | 
            Product.code.ilike(f'%{search}%')
        )
    ).limit(20).all()
    
    return jsonify([{
        'id': p.id,
        'name': p.name,
        'code': p.code,
        'unit': p.unit,
        'secondary_unit': p.secondary_unit,
        'multiplier': p.multiplier,
        'stock': p.stock,
        'cost_price': p.cost_price,
        'latest_audit': p.latest_audit.isoformat() if p.latest_audit else None
    } for p in products])

@app.route('/api/inventory/audit', methods=['POST'])
def create_inventory_audit():
    data = request.json
    # data: { note, status, items: [ { product_id, actual_stock, system_stock } ] }
    try:
        new_audit = InventoryAudit(
            note=data.get('note', ''),
            status=data.get('status', 'Completed')
        )
        db.session.add(new_audit)
        db.session.flush() # Get ID
        
        for item in data.get('items', []):
            product_id = item['product_id']
            actual_stock = item['actual_stock']
            system_stock = item.get('system_stock') # Frontend should send this for snapshot
            
            prod = db.session.get(Product, product_id)
            if not prod: continue
            
            if system_stock is None:
                system_stock = prod.stock
                
            discrepancy = actual_stock - system_stock
            
            audit_item = InventoryAuditDetail(
                audit_id=new_audit.id,
                product_id=product_id,
                system_stock=system_stock,
                actual_stock=actual_stock,
                discrepancy=discrepancy
            )
            db.session.add(audit_item)
            
            # Adjust StockBatches to maintain FIFO integrity and provide UI visibility
            if discrepancy > 0:
                # Add a positive adjustment batch
                new_batch = StockBatch(
                    product_id=product_id,
                    original_quantity=discrepancy,
                    current_quantity=discrepancy,
                    cost_price=prod.cost_price or 0,
                    created_at=get_vn_time()
                )
                db.session.add(new_batch)
            elif discrepancy < 0:
                # Subtract from existing batches (Oldest first)
                rem = abs(discrepancy)
                batches = StockBatch.query.filter(StockBatch.product_id == product_id, StockBatch.current_quantity > 0)\
                                          .order_by(StockBatch.created_at.asc()).all()
                for b in batches:
                    if rem <= 0: break
                    take = min(rem, b.current_quantity)
                    b.current_quantity -= take
                    rem -= take
                # Note: We don't error out if batches are insufficient, as prod.stock will be set anyway

            # Update product stock immediately
            prod.stock = actual_stock
            prod.latest_audit = get_vn_time()
            
        db.session.commit()
        return jsonify(new_audit.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/inventory/audits', methods=['GET'])
def get_inventory_audits():
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 20, type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')

    query = InventoryAudit.query

    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(InventoryAudit.date >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            # End date usually implies the end of the day if just YYYY-MM-DD
            if len(end_date) <= 10:
                end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
            else:
                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(InventoryAudit.date <= end_dt)
        except ValueError:
            pass
    
    if search:
        query = query.filter(InventoryAudit.note.ilike(f'%{search}%'))

    pagination = query.order_by(InventoryAudit.date.desc()).paginate(page=page, per_page=limit, error_out=False)
    return jsonify({
        'items': [a.to_dict() for a in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': pagination.page
    })

# --- Inventory Conversion Routes ---
@app.route('/api/inventory/convert', methods=['POST'])
def convert_inventory():
    data = request.json
    source_id = data.get('source_product_id')
    dest_id = data.get('dest_product_id')
    source_qty = float(data.get('source_qty', 0))
    multiplier = float(data.get('multiplier', 1))
    dest_qty_actual = float(data.get('dest_qty_actual', 0))
    note = data.get('note', '')
    user_id = data.get('user_id')
    # NEW: Manual fallback from frontend if needed
    cost_price_at_manual = data.get('cost_price_at_conversion')

    if not source_id or not dest_id or source_qty <= 0:
        return jsonify({'error': 'Thiếu thông tin sản phẩm hoặc số lượng không hợp lệ.'}), 400

    source_p = Product.query.get_or_404(source_id)
    dest_p = Product.query.get_or_404(dest_id)

    if source_p.stock < source_qty:
        return jsonify({'error': f'Sản phẩm nguồn "{source_p.name}" không đủ tồn kho (Còn {source_p.stock}).'}), 400

    try:
        # Transactional stock update
        # 1. Handle Source Product (FIFO Batch Subtraction)
        batches = StockBatch.query.filter(
            StockBatch.product_id == source_id,
            StockBatch.current_quantity > 0
        ).order_by(StockBatch.created_at.asc()).all()

        remaining_to_sub = source_qty
        total_source_cost = 0

        for batch in batches:
            if remaining_to_sub <= 0: break
            
            sub_qty = min(batch.current_quantity, remaining_to_sub)
            total_source_cost += sub_qty * batch.cost_price
            batch.current_quantity -= sub_qty
            remaining_to_sub -= sub_qty

        # Fallback 1: If batches are insufficient, use manual price (from frontend) OR default product cost
        if remaining_to_sub > 0:
            fallback_price = cost_price_at_manual if cost_price_at_manual is not None else (source_p.cost_price or 0)
            total_source_cost += remaining_to_sub * fallback_price
            remaining_to_sub = 0
            
        # Fallback 2: Ultimate protection - even if batches found but they had 0 cost, and we have a manual price
        if total_source_cost == 0 and cost_price_at_manual:
             total_source_cost = source_qty * cost_price_at_manual

        # Update source stock
        source_p.stock -= source_qty

        # 2. Handle Destination Product
        # Calculate new cost price for destination based on source cost
        qty_for_cost = dest_qty_actual if dest_qty_actual > 0 else (source_qty * multiplier)
        new_dest_cost = total_source_cost / qty_for_cost

        dest_p.stock += dest_qty_actual
        
        # Create new batch for dest product
        new_batch = StockBatch(
            product_id=dest_id,
            original_quantity=dest_qty_actual,
            current_quantity=dest_qty_actual,
            cost_price=new_dest_cost,
            created_at=get_vn_time()
        )
        db.session.add(new_batch)

        # 3. Log the conversion
        c_at_c = total_source_cost / source_qty if source_qty > 0 else 0
        app.logger.info(f"--- Conversion Debug ---")
        app.logger.info(f"Source: {source_p.name} (ID: {source_id})")
        app.logger.info(f"Total Source Cost: {total_source_cost}")
        app.logger.info(f"Source Qty: {source_qty}")
        app.logger.info(f"Source P Cost Price (Default): {source_p.cost_price}")
        app.logger.info(f"Calculated cost_price_at_conversion: {c_at_c}")
        
        conversion = InventoryConversion(
            source_product_id=source_id,
            dest_product_id=dest_id,
            source_qty=source_qty,
            multiplier=multiplier,
            dest_qty_expected=source_qty * multiplier,
            dest_qty_actual=dest_qty_actual,
            cost_price_at_conversion=c_at_c,
            user_id=user_id,
            note=note
        )
        db.session.add(conversion)

        db.session.commit()
        return jsonify({
            'message': 'Xẻ lẻ thành công!',
            'conversion': conversion.to_dict(),
            'source_stock': source_p.stock,
            'dest_stock': dest_p.stock
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Conversion error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/inventory/conversions', methods=['GET'])
def get_conversions():
    # Show last 50 conversions
    conversions = InventoryConversion.query.order_by(InventoryConversion.date.desc()).limit(50).all()
    return jsonify([c.to_dict() for c in conversions])

@app.route('/api/inventory/conversions/<int:id>', methods=['DELETE'])
def delete_conversion(id):
    try:
        c = InventoryConversion.query.get_or_404(id)
        
        # Reverse stock changes
        from models import Product
        source_p = Product.query.get(c.source_product_id)
        dest_p = Product.query.get(c.dest_product_id)
        
        if source_p:
            source_p.stock += c.source_qty
        if dest_p:
            dest_p.stock -= c.dest_qty_actual
            
        db.session.delete(c)
        db.session.commit()
        return jsonify({'message': 'Đã xóa bản ghi và hoàn lại tồn kho thành công'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/inventory/conversions/<int:id>', methods=['PUT'])
def update_conversion(id):
    try:
        c = InventoryConversion.query.get_or_404(id)
        data = request.json
        
        from models import Product
        source_p = Product.query.get(c.source_product_id)
        dest_p = Product.query.get(c.dest_product_id)
        
        # 1. Reverse OLD stock
        if source_p:
            source_p.stock += c.source_qty
        if dest_p:
            dest_p.stock -= c.dest_qty_actual
            
        # 2. Apply NEW data
        new_source_qty = float(data.get('source_qty', c.source_qty))
        new_multiplier = float(data.get('multiplier', c.multiplier))
        new_dest_qty_actual = float(data.get('dest_qty_actual', c.dest_qty_actual))
        new_note = data.get('note', c.note)
        new_cost_price_at_conversion = data.get('cost_price_at_conversion', c.cost_price_at_conversion)
        
        # Update record
        c.source_qty = new_source_qty
        c.multiplier = new_multiplier
        c.dest_qty_expected = new_source_qty * new_multiplier
        c.dest_qty_actual = new_dest_qty_actual
        c.note = new_note
        c.cost_price_at_conversion = new_cost_price_at_conversion
        
        # Apply NEW stock
        if source_p:
            source_p.stock -= new_source_qty
        if dest_p:
            dest_p.stock += new_dest_qty_actual
            
        db.session.commit()
        return jsonify({
            'message': 'Cập nhật và cân bằng tồn kho thành công',
            'conversion': c.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# --- Partners ---
@app.route('/api/partners', methods=['GET'])
def get_partners():
    search = request.args.get('search', '').lower()
    partner_type = request.args.get('type', 'All')
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)
    
    query = Partner.query
    
    if search:
        s_norm = remove_accents(search)
        query = query.filter(db.func.remove_accents(Partner.name).ilike(f'%{s_norm}%') | Partner.phone.ilike(f'%{search}%'))
    
    if partner_type == 'Customer':
        query = query.filter(Partner.is_customer == True)
    elif partner_type == 'Supplier':
        query = query.filter(Partner.is_supplier == True)
    elif partner_type == 'Both':
        query = query.filter(Partner.is_customer == True, Partner.is_supplier == True)

    brand = request.args.get('brand')
    if brand:
        query = query.join(Order, Partner.id == Order.partner_id)\
                     .join(OrderDetail, Order.id == OrderDetail.order_id)\
                     .join(Product, OrderDetail.product_id == Product.id)\
                     .filter(Product.brand == brand)
        if partner_type == 'Customer':
            query = query.filter(Order.type == 'Sale')
        elif partner_type == 'Supplier':
            query = query.filter(Order.type == 'Purchase')
        query = query.distinct()

    sort_by = request.args.get('sort_by', 'name')
    sort_order = request.args.get('sort_order', 'asc')

    sort_map = {
        'id': Partner.id,
        'name': Partner.name,
        'type': Partner.type,
        'debt_balance': Partner.debt_balance,
        'phone': Partner.phone
    }

    sort_col = sort_map.get(sort_by, Partner.name)
    if sort_order == 'desc':
        query = query.order_by(sort_col.desc())
    else:
        query = query.order_by(sort_col.asc())

    if page and limit:
        pagination = query.paginate(page=page, per_page=limit, error_out=False)
        return jsonify({
            'items': [p.to_dict() for p in pagination.items],
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': pagination.page
        })
    else:
        partners = query.all()
        
        # Calculate yearly revenue for the current year
        current_year = str(datetime.now().year)
        partner_ids = [p.id for p in partners]
        revenue_rows = db.session.query(Order.partner_id, db.func.sum(Order.total_amount))\
            .filter(Order.partner_id.in_(partner_ids), Order.type == 'Sale')\
            .filter(db.func.strftime('%Y', Order.date) == current_year)\
            .group_by(Order.partner_id).all()
        rev_map = {r[0]: r[1] for r in revenue_rows}

        result = []
        for p in partners:
            pd = p.to_dict()
            pd['yearly_revenue'] = rev_map.get(p.id, 0)
            result.append(pd)
            
        return jsonify(result)

@app.route('/api/partners', methods=['POST'])
def create_partner():
    data = request.json
    new_partner = Partner(
        name=data['name'],
        is_customer=data.get('is_customer', True),
        is_supplier=data.get('is_supplier', False),
        cccd=data.get('cccd'),
        phone=data.get('phone'),
        address=data.get('address'),
        debt_balance=0 # Initially 0, will be updated by create_opening_balance_order + recalculate
    )
    # Sync 'type' for backward compatibility
    if new_partner.is_customer and new_partner.is_supplier:
        new_partner.type = 'Both'
    elif new_partner.is_supplier:
        new_partner.type = 'Supplier'
    else:
        new_partner.type = 'Customer'
        
    db.session.add(new_partner)
    db.session.commit()
    
    # Create Opening Balance Order if debt exists
    op_bal = float(data.get('opening_balance', data.get('debt_balance', 0)))
    if op_bal != 0:
        create_opening_balance_order(new_partner.id, op_bal, new_partner.type)
        recalculate_partner_debt_internal(new_partner.id)
        
    return jsonify(new_partner.to_dict()), 201

@app.route('/api/partners/<int:id>', methods=['PUT'])
def update_partner(id):
    try:
        partner = Partner.query.get_or_404(id)
        data = request.json
        
        partner.name = data.get('name', partner.name)
        partner.is_customer = data.get('is_customer', partner.is_customer)
        partner.is_supplier = data.get('is_supplier', partner.is_supplier)
        partner.cccd = data.get('cccd', partner.cccd)
        partner.phone = data.get('phone', partner.phone)
        partner.address = data.get('address', partner.address)
        
        # Sync 'type' for backward compatibility
        if partner.is_customer and partner.is_supplier:
            partner.type = 'Both'
        elif partner.is_supplier:
            partner.type = 'Supplier'
        else:
            partner.type = 'Customer'
        
        # Support both 'opening_balance' and 'debt_balance' (legacy) for updating nợ đầu kỳ
        opening_key = 'opening_balance' if 'opening_balance' in data else 'debt_balance'
        if opening_key in data:
            new_opening = float(data[opening_key])
            # Check if Opening Balance Order exists
            nodau = Order.query.filter_by(partner_id=id).filter(Order.display_id.in_(['#NODAU', 'NODAU'])).first()
            if nodau:
                if new_opening == 0:
                    db.session.delete(nodau)
                else:
                    nodau.total_amount = abs(new_opening)
                    nodau.type = 'Sale' if new_opening >= 0 else 'Purchase'
                    nodau.display_id = '#NODAU' # Normalize
            elif new_opening != 0:
                create_opening_balance_order(id, new_opening)
            
        db.session.commit()
        # Ensure consistency
        recalculate_partner_debt_internal(id)
        return jsonify(partner.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

def recalculate_partner_debt_internal(id):
    partner = Partner.query.get(id)
    if not partner: return
    # Sum of all Debt orders
    sale_debt = db.session.query(db.func.sum(Order.total_amount)).filter(Order.partner_id == id, Order.payment_method == 'Debt', Order.type == 'Sale').scalar() or 0
    purchase_debt = db.session.query(db.func.sum(Order.total_amount)).filter(Order.partner_id == id, Order.payment_method == 'Debt', Order.type == 'Purchase').scalar() or 0
    
    # Vouchers (Filter out auto-generated vouchers for Cash payment method, as they do not affect debt)
    cash_receipts = db.session.query(db.func.sum(CashVoucher.amount)).filter(CashVoucher.partner_id == id, CashVoucher.type == 'Receipt', CashVoucher.source != 'auto').scalar() or 0
    cash_payments = db.session.query(db.func.sum(CashVoucher.amount)).filter(CashVoucher.partner_id == id, CashVoucher.type == 'Payment', CashVoucher.source != 'auto').scalar() or 0
    debt_increases = db.session.query(db.func.sum(CashVoucher.amount)).filter(CashVoucher.partner_id == id, CashVoucher.type == 'DebtIncrease', CashVoucher.source != 'auto').scalar() or 0
    
    # Bank Transactions - Filter out non-debt payments
    bank_receipts = db.session.query(db.func.sum(BankTransaction.amount))\
        .outerjoin(Order, BankTransaction.order_id == Order.id)\
        .filter(BankTransaction.partner_id == id, BankTransaction.type == 'Deposit')\
        .filter(db.or_(Order.id == None, Order.payment_method == 'Debt'))\
        .scalar() or 0
        
    bank_payments = db.session.query(db.func.sum(BankTransaction.amount))\
        .outerjoin(Order, BankTransaction.order_id == Order.id)\
        .filter(BankTransaction.partner_id == id, BankTransaction.type == 'Withdrawal')\
        .filter(db.or_(Order.id == None, Order.payment_method == 'Debt'))\
        .scalar() or 0
    
    total_receipts = cash_receipts + bank_receipts
    total_payments = cash_payments + bank_payments + debt_increases
    
    partner.debt_balance = (sale_debt - purchase_debt) - (total_receipts - total_payments)
    db.session.commit()

def recalculate_product_cost_price(product_id):
    """
    Recalculates both Average Cost (cost_price) and Latest Purchase Price (latest_cost_price).
    Average Cost = Total Value of ALL active batches / Total Qty of ALL active batches.
    Latest Cost = cost_price of the most recent StockBatch (regardless of current_quantity).
    """
    prod = Product.query.get(product_id)
    if not prod or prod.is_combo: return
    
    # 1. Calculate Average Cost from ACTIVE batches
    active_batches = StockBatch.query.filter(StockBatch.product_id == product_id, StockBatch.current_quantity > 0).all()
    calculated_avg = 0
    if active_batches:
        total_qty = sum(b.current_quantity for b in active_batches)
        if total_qty > 0:
            total_value = sum(b.current_quantity * b.cost_price for b in active_batches)
            calculated_avg = total_value / total_qty

    # Fallback if no active batches or calculated average is 0
    if calculated_avg == 0:
        # Keep existing cost_price if it's already set and positive, otherwise look for latest nonzero batch
        if not prod.cost_price or prod.cost_price <= 0:
            latest_nonzero_batch = StockBatch.query.filter(StockBatch.product_id == product_id, StockBatch.cost_price > 0)\
                                                   .order_by(StockBatch.created_at.desc()).first()
            if latest_nonzero_batch:
                prod.cost_price = latest_nonzero_batch.cost_price
    else:
        prod.cost_price = calculated_avg

    # 2. Calculate Latest Purchase Price from THE MOST RECENT batch
    latest_batch = StockBatch.query.filter_by(product_id=product_id)\
                             .order_by(StockBatch.created_at.desc()).first()
    if latest_batch:
        # Update latest_cost_price only if the batch has a positive cost price,
        # or if we don't have a valid latest_cost_price already
        if latest_batch.cost_price and latest_batch.cost_price > 0:
            prod.latest_cost_price = latest_batch.cost_price
        elif not prod.latest_cost_price or prod.latest_cost_price <= 0:
            prod.latest_cost_price = 0
    else:
        # Do not force to 0 if they already have manually entered / imported values
        if not prod.cost_price:
            prod.cost_price = 0
        if not prod.latest_cost_price:
            prod.latest_cost_price = 0
    
    db.session.commit()

def adjust_negative_stock_backorder(product_id, incoming_qty, purchase_price):
    prod = Product.query.get(product_id)
    if not prod or prod.stock >= 0:
        return incoming_qty
    
    neg_stock = abs(prod.stock)
    resolve_qty = min(neg_stock, incoming_qty)
    
    # Update the cost price of the most recent sales that occurred when stock was negative
    details = OrderDetail.query.join(Order).filter(
        OrderDetail.product_id == product_id,
        Order.type == 'Sale'
    ).order_by(Order.date.desc()).all()
    
    rem = resolve_qty
    for detail in details:
        if rem <= 0:
            break
        qty = detail.quantity
        take = min(rem, qty)
        old_cost = detail.cost_price or 0
        detail.cost_price = (take * purchase_price + (qty - take) * old_cost) / qty
        rem -= take
        
    return incoming_qty - resolve_qty

@app.route('/api/partners/<int:id>/quick-debt', methods=['POST'])
def quick_debt(id):
    try:
        partner = Partner.query.get_or_404(id)
        data = request.json
        amount = float(data.get('amount', 0))
        date_str = data.get('date') # Expecting YYYY-MM-DD
        note = data.get('note', '')

        if not note:
            note = f"Ghi nợ nhanh (Sổ tay)"

        # Increment debt for customer
        partner.debt_balance += amount
        
        # Create a voucher with type "DebtIncrease"
        from models import utc_now
        dt = utc_now()
        if date_str:
            try:
                from datetime import datetime
                dt_obj = datetime.strptime(date_str, '%Y-%m-%d')
                dt = dt.replace(year=dt_obj.year, month=dt_obj.month, day=dt_obj.day)
            except Exception:
                pass

        voucher = CashVoucher(
            partner_id=id,
            amount=amount,
            note=note,
            type='DebtIncrease',
            source='quick_debt',
            date=dt
        )
        db.session.add(voucher)
        db.session.commit()
        
        # Enforce consistency
        recalculate_partner_debt_internal(id)
        
        return jsonify(partner.to_dict())

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/partners/<int:id>', methods=['DELETE'])
def delete_partner(id):
    try:
        # Check if has orders
        order_count = Order.query.filter_by(partner_id=id).count()
        if order_count > 0:
            return jsonify({'error': f'Không thể xóa đối tác vì đã có {order_count} hóa đơn/giao dịch. Hãy xóa các hóa đơn này trong Lịch sử trước.'}), 400
        
        # Check if has vouchers
        voucher_count = CashVoucher.query.filter_by(partner_id=id).count()
        if voucher_count > 0:
            return jsonify({'error': f'Không thể xóa đối tác vì đã có {voucher_count} phiếu thu/chi. Hãy xóa các phiếu này trong Quỹ tiền trước.'}), 400
        
        # Check if has bank transactions
        bank_count = BankTransaction.query.filter_by(partner_id=id).count()
        if bank_count > 0:
            return jsonify({'error': f'Không thể xóa đối tác vì đã có {bank_count} giao dịch ngân hàng. Hãy xóa các giao dịch này trong Quỹ tiền trước.'}), 400
        
        partner = Partner.query.get_or_404(id)
        db.session.delete(partner)
        db.session.commit()
        return jsonify({'message': 'Deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/partners/bulk-delete', methods=['POST'])
def bulk_delete_partners():
    data = request.json
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'No IDs provided'}), 400
    
    try:
        # Check if any have orders, vouchers, or bank transactions
        in_use_orders = Order.query.filter(Order.partner_id.in_(ids)).group_by(Order.partner_id).count()
        in_use_vouchers = CashVoucher.query.filter(CashVoucher.partner_id.in_(ids)).group_by(CashVoucher.partner_id).count()
        in_use_bank = BankTransaction.query.filter(BankTransaction.partner_id.in_(ids)).group_by(BankTransaction.partner_id).count()
        
        if in_use_orders > 0 or in_use_vouchers > 0 or in_use_bank > 0:
            return jsonify({'error': 'Có một số đối tác đã có lịch sử đơn hàng, phiếu thu chi, hoặc giao dịch ngân hàng, không thể xóa hàng loạt.'}), 400
            
        deleted = Partner.query.filter(Partner.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'message': f'Đã xóa {deleted} đối tác thành công'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/partners/import', methods=['POST'])
def import_partners():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        headers = [str(cell.value).strip() for cell in ws[1]]
        
        def get_val(row, col_name, default=None):
            try:
                idx = headers.index(col_name)
                val = row[idx].value
                return val if val is not None else default
            except (ValueError, IndexError):
                return default

        count = 0
        for row in ws.iter_rows(min_row=2):
            name = str(get_val(row, 'Tên đối tác', '')).strip()
            if not name or name == 'None' or name == '': continue
            
            p_type_raw = str(get_val(row, 'Loại', 'Customer')).strip()
            # Normalize type
            if 'khách' in p_type_raw.lower() and 'ncc' in p_type_raw.lower(): p_type = 'Both'
            elif 'khách' in p_type_raw.lower(): p_type = 'Customer'
            elif 'ncc' in p_type_raw.lower() or 'nhà cung cấp' in p_type_raw.lower(): p_type = 'Supplier'
            else: p_type = 'Customer'
            
            # Determine flags
            is_cust = True
            is_supp = False
            if p_type == 'Both':
                is_cust = True
                is_supp = True
            elif p_type == 'Supplier':
                is_cust = False
                is_supp = True
            elif p_type == 'Customer':
                is_cust = True
                is_supp = False

            partner = Partner.query.filter_by(name=name, type=p_type).first()
            if not partner:
                partner = Partner(name=name, type=p_type, is_customer=is_cust, is_supplier=is_supp)
                db.session.add(partner)
            else:
                # Update existing if needed
                partner.is_customer = is_cust
                partner.is_supplier = is_supp
            
            phone = get_val(row, 'Số điện thoại')
            if phone is not None: partner.phone = str(phone)
            
            address = get_val(row, 'Địa chỉ')
            if address is not None: partner.address = str(address)
            
            debt = get_val(row, 'Nợ đầu kỳ')
            if debt is not None: 
                d_val = float(debt)
                partner.debt_balance = d_val
                # Add to session first, flush to get ID, then we will handle opening order
                # But to handle it cleanly, wait until after main loop?
                # Actually, we can just create the Order object here and link it.
                # Since we haven't committed partner yet, we can't link validation might fail? 
                # No, SQLAlchemy handles uncommitted objects in session.
                # However, to avoid complexity, let's just flush.
                db.session.flush() # Ensure partner.id is set
                
                # Check if 'NODAU' exists
                existing_nodau = Order.query.filter_by(partner_id=partner.id).filter(Order.display_id.in_(['#NODAU', 'NODAU'])).first()
                if existing_nodau:
                    # Update it
                    is_positive = d_val > 0
                    existing_nodau.type = 'Sale' if is_positive else 'Purchase'
                    existing_nodau.total_amount = abs(d_val)
                    existing_nodau.display_id = '#NODAU' # Normalize
                elif d_val != 0:
                    # Create new
                    is_positive = d_val > 0
                    order_type = 'Sale' if is_positive else 'Purchase'
                    new_nodau = Order(
                        partner_id=partner.id,
                        type=order_type,
                        payment_method='Debt',
                        display_id='#NODAU',
                        total_amount=abs(d_val),
                        note='Nợ đầu kỳ (Import)',
                        amount_paid=0,
                        date=datetime.now()
                    )
                    db.session.add(new_nodau)
            
            count += 1
            
        db.session.commit()
        
        return jsonify({'message': f'Đã nhập {count} đối tác thành công! Lịch sử công nợ đầu kỳ đã được ghi nhận.'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

def create_opening_balance_order(partner_id, amount, p_type=None):
    # Ensure only ONE opening balance order exists
    existing = Order.query.filter_by(partner_id=partner_id).filter(Order.display_id.in_(['#NODAU', 'NODAU'])).first()
    
    is_positive = amount > 0
    order_type = 'Sale' if is_positive else 'Purchase'
    abs_amount = abs(amount)
    
    if existing:
        existing.type = order_type
        existing.total_amount = abs_amount
        existing.display_id = '#NODAU'
        db.session.commit()
        return existing

    order = Order(
        partner_id=partner_id,
        type=order_type,
        payment_method='Debt',
        display_id='#NODAU',
        total_amount=abs_amount,
        note='Nợ đầu kỳ',
        amount_paid=0,
        date=datetime.now()
    )
    db.session.add(order)
    db.session.commit()
    return order

@app.route('/api/partners/<int:id>/fix-opening-balance', methods=['POST'])
def fix_opening_balance(id):
    partner = Partner.query.get_or_404(id)
    data = request.json
    amount = data.get('amount') # The missing amount to record
    
    if not amount:
        return jsonify({'error': 'Amount required'}), 400
        
    try:
        create_opening_balance_order(id, float(amount))
        return jsonify({'message': 'Đã ghi nhận nợ đầu kỳ thành công'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/partners/template', methods=['GET'])
def get_partner_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Partners"
    
    headers = ['Tên đối tác', 'Loại', 'Số điện thoại', 'Địa chỉ', 'Nợ đầu kỳ']
    ws.append(headers)
    ws.append(['Nguyễn Văn A', 'Khách hàng', '0901234567', 'Hà Nội', 0])
    ws.append(['Công ty X', 'NCC', '0281234567', 'TP HCM', 1500000])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='mau_nhap_doi_tac.xlsx'
    )

@app.route('/api/partners/<int:id>/debt-cycles', methods=['GET'])
def get_partner_debt_cycles(id):
    # Fetch all orders, vouchers and bank transactions for the partner
    orders = Order.query.filter_by(partner_id=id).all()
    vouchers = CashVoucher.query.filter_by(partner_id=id).all()
    banks = BankTransaction.query.filter_by(partner_id=id).all()
    
    timeline = []
    for o in orders: timeline.append({'type': 'order', 'date': o.date, 'obj': o})
    for v in vouchers: timeline.append({'type': 'voucher', 'date': v.date, 'obj': v})
    for b in banks: timeline.append({'type': 'bank', 'date': b.date, 'obj': b})
    
    timeline.sort(key=lambda x: x['date'])
    
    cycles = []
    current_cycle = None
    balance = 0
    cycle_count = 0
    
    for item in timeline:
        prev_balance = balance
        
        if item['type'] == 'order':
            o = item['obj']
            if o.payment_method == 'Debt':
                if o.type == 'Sale': balance += o.total_amount
                else: balance -= o.total_amount
        elif item['type'] == 'voucher':
            v = item['obj']
            if v.type == 'Receipt': balance -= v.amount
            else: balance += v.amount
        elif item['type'] == 'bank':
            b = item['obj']
            if b.type == 'Deposit': balance -= b.amount
            else: balance += b.amount
        
        # Start cycle: absolute balance >= 1 but was < 1
        if abs(prev_balance) < 1 and abs(balance) >= 1 and current_cycle is None:
            cycle_count += 1
            current_cycle = {
                'id': cycle_count,
                'label': f"Chu kỳ {cycle_count} (Từ {item['date'].strftime('%d/%m/%y')})",
                'start_date': item['date'].isoformat(),
                'end_date': None,
                'status': 'Đang nợ'
            }
        
        # End cycle: balance returns to near 0
        if current_cycle and abs(balance) < 1:
            current_cycle['end_date'] = item['date'].isoformat()
            current_cycle['status'] = 'Đã tất toán'
            cycles.append(current_cycle)
            current_cycle = None
            
    if current_cycle:
        cycles.append(current_cycle)
        
    cycles.reverse()
    return jsonify(cycles)

@app.route('/api/partners/<int:id>/recalculate-debt', methods=['POST'])
def recalculate_partner_debt(id):
    partner = Partner.query.get_or_404(id)
    try:
        recalculate_partner_debt_internal(id)
        return jsonify({'message': 'Recalculated successfully', 'new_balance': partner.debt_balance})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

def calculate_partner_ledger(id, filter_type='all', start_date=None, end_date=None):
    partner = Partner.query.get_or_404(id)
    
    # 1. Fetch EVERYTHING for the partner to calculate historical running balance
    all_orders = Order.query.filter_by(partner_id=id).all()
    all_vouchers = CashVoucher.query.filter_by(partner_id=id).filter(CashVoucher.source != 'auto').all()
    all_banks = BankTransaction.query.filter_by(partner_id=id).all()
    
    timeline_raw = []
    for o in all_orders:
        display_id_clean = (o.display_id or '').strip()
        is_nodau = display_id_clean in ['#NODAU', 'NODAU'] or 'NODAU' in display_id_clean.upper()
        if is_nodau:
            desc = "Nợ đầu kỳ"
        else:
            clean_num = display_id_clean.lstrip('#') if display_id_clean else str(o.id)
            desc = f"Bán hàng - #{clean_num}" if o.type == 'Sale' else f"Nhập hàng - #{clean_num}"
            
        timeline_raw.append({
            'type': 'Order', 'date': o.date, 'is_debt': o.payment_method == 'Debt', 'obj': o,
            'desc': desc
        })
    for v in all_vouchers:
        timeline_raw.append({
            'type': 'Voucher', 'date': v.date, 'is_debt': True, 'obj': v,
            'desc': v.note or f"{'Phiếu thu' if v.type == 'Receipt' else 'Phiếu chi'} - #{v.id}"
        })
    for b in all_banks:
        # Check if Linked order is not 'Debt' - if so, ignore for debt ledger
        is_debt_payment = True
        if b.order_id:
            order = next((o for o in all_orders if o.id == b.order_id), None)
            if order and order.payment_method != 'Debt':
                is_debt_payment = False
                
        timeline_raw.append({
            'type': 'Bank', 'date': b.date, 'is_debt': is_debt_payment, 'obj': b,
            'desc': b.note or f"Chuyển khoản - {b.account.bank_name}"
        })
        
    timeline_raw.sort(key=lambda x: x['date'])
    
    s_date = datetime.fromisoformat(start_date) if start_date else None
    e_date = datetime.fromisoformat(end_date) if end_date else None
    
    ledger = []
    balance = 0
    opening_balance = 0
    
    # 2. Iterate through all items to calculate balance correctly
    for item in timeline_raw:
        inc = 0
        dec = 0
        
        if item['is_debt']:
            if item['type'] == 'Order':
                o = item['obj']
                if o.type == 'Sale': inc = o.total_amount; balance += o.total_amount
                else: dec = o.total_amount; balance -= o.total_amount
            elif item['type'] == 'Voucher':
                v = item['obj']
                if v.type == 'Receipt': dec = v.amount; balance -= v.amount
                else: inc = v.amount; balance += v.amount
            elif item['type'] == 'Bank':
                b = item['obj']
                if b.type == 'Deposit': dec = b.amount; balance -= b.amount
                else: inc = b.amount; balance += b.amount
        
        # Range check
        in_range = True
        if s_date and item['date'] < s_date: in_range = False; opening_balance = balance
        if e_date and item['date'] > e_date: in_range = False
        
        if not in_range: continue

        # Add initial "Opening Balance" row if we're filtering by date
        if s_date and not ledger and opening_balance != 0:
            ledger.append({
                'id': 0, 'date': s_date.isoformat(), 'ref_id': 'OPN', 'desc': 'Số dư trước kỳ lọc',
                'type': 'System', 'payment_method': '-', 'increase': 0, 'decrease': 0,
                'running_balance': opening_balance, 'details': [], 'user_name': 'Hệ thống', 'obj': {}
            })

        # Apply frontend filters (All/Debt/Cash)
        if filter_type == 'debt' and not item['is_debt']: continue
        if filter_type == 'cash' and item['is_debt']: continue
        
        is_nodau_order = item['type'] == 'Order' and getattr(item['obj'], 'display_id', '') in ['#NODAU', 'NODAU']
        row = {
            'id': item['obj'].id,
            'date': item['date'].isoformat(),
            'ref_id': 'NODAU' if is_nodau_order else f"{item['type'][:3].upper()}-{item['obj'].id}",
            'desc': item['desc'],
            'type': item['type'],
            'payment_method': '-' if is_nodau_order else getattr(item['obj'], 'payment_method', 'Bank' if item['type'] == 'Bank' else 'Cash'),
            'increase': inc,
            'decrease': dec,
            'running_balance': balance,
            'details': [d.to_dict() for d in getattr(item['obj'], 'details', [])] if item['type'] == 'Order' else [],
            'user_name': getattr(item['obj'], 'user_name', 'Hệ thống'),
            'obj': item['obj'].to_dict()
        }
        ledger.append(row)
        
    ledger.reverse()
    return ledger, balance, partner

@app.route('/api/partners/<int:id>/ledger', methods=['GET'])
def get_partner_ledger(id):
    try:
        filter_type = request.args.get('filter_type', 'all')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        ledger, balance, partner = calculate_partner_ledger(id, filter_type, start_date, end_date)
        return jsonify({
            'partner': partner.to_dict(),
            'ledger': ledger,
            'current_balance': balance
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/partners/<int:id>/ledger/export', methods=['GET'])
def export_partner_ledger(id):
    try:
        filter_type = request.args.get('filter_type', 'all')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        ledger, balance, partner = calculate_partner_ledger(id, filter_type, start_date, end_date)
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Ledger_{partner.name}"
        
        # Header Info
        ws.append(['BÁO CÁO CHI TIẾT CÔNG NỢ ĐỐI TÁC'])
        ws.append([f'Đối tác: {partner.name}'])
        ws.append([f'Số điện thoại: {partner.phone or "N/A"}'])
        ws.append([f'Thời gian xuất: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'])
        ws.append([])
        
        headers = ['Ngày', 'Mã tham chiếu', 'Nội dung', 'Loại', 'HT thanh toán', 'Số lượng', 'Đơn giá', 'Thành tiền', 'Ghi nợ (+)', 'Trả nợ (-)', 'Dư nợ cuối']
        ws.append(headers)
        
        # Style headers
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        
        # Helper for translations
        def translate_type(t):
            if not t: return ''
            m = {
                'order': 'Hóa đơn', 
                'voucher': 'Chứng từ', 
                'bank': 'Ngân hàng', 
                'system': 'Hệ thống',
                'purchase': 'Nhập hàng',
                'sale': 'Bán hàng'
            }
            return m.get(t.lower(), t)
            
        def translate_method(m):
            if not m: return '-'
            mapping = {
                'debt': 'Ghi nợ', 
                'cash': 'Tiền mặt', 
                'bank': 'Chuyển khoản'
            }
            return mapping.get(m.lower(), m)

        # Add data in chronological order for Excel
        data_to_export = ledger[::-1]
        for row in data_to_export:
            # Main transaction row
            ws.append([
                row['date'],
                row['ref_id'],
                row['desc'],
                translate_type(row['type']),
                translate_method(row['payment_method']),
                '', '', '', # Qty, Price, Total (empty for main row)
                row['increase'] if row['increase'] != 0 else '',
                row['decrease'] if row['decrease'] != 0 else '',
                row['running_balance']
            ])
            
            # If it's an order, add detail rows
            if row['type'] == 'Order' and row.get('details'):
                for d in row['details']:
                    ws.append([
                        '', # Date
                        '', # Ref
                        d['product_name'], # Removed prefix
                        'Chi tiết',
                        '', # Payment method
                        d['quantity'],
                        d['unit_price'],
                        d['total_price'],
                        '', '', '' # Debt columns empty for details
                    ])
                    # Style detail row a bit (dim it)
                    for cell in ws[ws.max_row]:
                        cell.font = Font(italic=True, color="666666", size=9)

        ws.append([])
        # Style total
        total_row = [''] * 7 + ['TỔNG DƯ NỢ HIỆN TẠI:', '', '', balance]
        ws.append(total_row)
        ws[ws.max_row][7].font = Font(bold=True)
        ws[ws.max_row][10].font = Font(bold=True)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"SoCai_{partner.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# --- Excel Export ---
@app.route('/api/products/export', methods=['GET'])
def export_products():
    wb = Workbook()
    ws = wb.active
    ws.title = "Products Export"
    
    # Headers MUST match the Import headers for consistency if possible, or be descriptive
    headers = [
        'Mã hàng', 'Tên sản phẩm', 'Loại sản phẩm', 'Hãng', 'Hoạt chất', 
        'Đơn vị', 'Đơn vị phụ', 'Quy cách (1 phụ = ? chính)', 
        'Giá vốn', 'Giá bán', 'Giá kế toán', 
        'Tồn kho (Đơn vị chính)', 'Tồn kế toán', 
        'Hạn sử dụng', 'Trạng thái'
    ]
    ws.append(headers)
    
    products = Product.query.order_by(Product.name).all()
    for p in products:
        ws.append([
            p.code or '',
            p.name,
            p.category.name if p.category else 'Chưa phân loại',
            p.brand or '',
            p.active_ingredient or '',
            p.unit or '',
            p.secondary_unit or '',
            p.multiplier or 1,
            p.cost_price or 0,
            p.sale_price or 0,
            p.accounting_price or 0,
            p.stock or 0,
            p.accounting_stock or 0,
            p.expiry_date or '',
            'Đang kinh doanh' if p.is_active else 'Ngừng kinh doanh'
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='danh_sach_san_pham.xlsx'
    )

@app.route('/api/partners/export', methods=['GET'])
def export_partners():
    wb = Workbook()
    ws = wb.active
    ws.title = "Partners Export"
    
    headers = ['Mã ĐT', 'Tên đối tác', 'Loại', 'Số điện thoại', 'CCCD', 'Địa chỉ', 'Nợ đầu kỳ', 'Nợ hiện tại']
    ws.append(headers)
    
    partners = Partner.query.order_by(Partner.name).all()
    for p in partners:
        pd = p.to_dict()
        p_type = 'Khách hàng'
        if p.is_supplier and p.is_customer: p_type = 'Khách & NCC'
        elif p.is_supplier: p_type = 'Nhà cung cấp'
        
        ws.append([
            pd['id'],
            pd['name'], 
            p_type, 
            pd['phone'] or '', 
            pd['cccd'] or '',
            pd['address'] or '', 
            pd['opening_balance'],
            pd['debt_balance']
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='danh_sach_doi_tac.xlsx'
    )


# --- Custom Prices (Wholesale) ---
@app.route('/api/custom-prices/bulk', methods=['POST'])
def bulk_save_custom_prices():
    data = request.json
    partner_id = data.get('partner_id')
    prices = data.get('prices', []) # Expected list of {product_id, price}
    
    if not partner_id:
        return jsonify({'error': 'Partner ID is required'}), 400
        
    synced_count = 0
    updated_count = 0
    
    for item in prices:
        prod_id = item.get('product_id')
        price = item.get('price')
        if prod_id is None: continue
        
        # Get general product price
        product = Product.query.get(prod_id)
        if not product: continue
        
        if price == product.sale_price:
            # If price matches general price, delete existing record if any
            existing = CustomerPrice.query.filter_by(partner_id=partner_id, product_id=prod_id).first()
            if existing:
                db.session.delete(existing)
                synced_count += 1
            continue

        cp = CustomerPrice.query.filter_by(partner_id=partner_id, product_id=prod_id).first()
        if not cp:
            cp = CustomerPrice(partner_id=partner_id, product_id=prod_id)
            db.session.add(cp)
        cp.price = price
        updated_count += 1
        
    db.session.commit()
    msg = f'Đã cập nhật {updated_count} giá riêng.'
    if synced_count > 0:
        msg += f' Đã đồng bộ {synced_count} sản phẩm theo giá chung.'
        
    return jsonify({
        'message': msg,
        'updated_count': updated_count,
        'synced_count': synced_count
    })

@app.route('/api/custom-prices/<int:partner_id>', methods=['GET'])
def get_custom_prices(partner_id):
    prices = CustomerPrice.query.filter_by(partner_id=partner_id).all()
    # Return as a dict for easy lookup: {product_id: price}
    return jsonify({p.product_id: p.price for p in prices})

@app.route('/api/custom-prices', methods=['POST'])
def save_custom_price():
    data = request.json
    partner_id = data['partner_id']
    product_id = data['product_id']
    price = data['price']
    
    product = Product.query.get(product_id)
    if product and price == product.sale_price:
        # If matches, delete and return synced status
        cp = CustomerPrice.query.filter_by(partner_id=partner_id, product_id=product_id).first()
        if cp:
            db.session.delete(cp)
            db.session.commit()
            return jsonify({'message': 'Đã đồng bộ về giá chung', 'synced': True})
        return jsonify({'message': 'Giá trùng với giá chung, không cần lưu riêng', 'synced': True})

    cp = CustomerPrice.query.filter_by(partner_id=partner_id, product_id=product_id).first()
    if not cp:
        cp = CustomerPrice(partner_id=partner_id, product_id=product_id)
        db.session.add(cp)
    
    cp.price = price
    db.session.commit()
    return jsonify(cp.to_dict())





@app.route('/api/custom-prices/cleanup', methods=['POST'])
def cleanup_custom_prices():
    # Find all custom prices that match their product's sale price
    custom_prices = CustomerPrice.query.all()
    deleted_count = 0
    
    for cp in custom_prices:
        if cp.product and cp.price == cp.product.sale_price:
            db.session.delete(cp)
            deleted_count += 1
            
    db.session.commit()
    return jsonify({
        'message': f'Đã dọn dẹp {deleted_count} bản ghi trùng khớp với giá chung.',
        'deleted_count': deleted_count
    })


# --- Sales / Purchases (Order) ---
@app.route('/api/orders', methods=['GET'])
def get_orders():
    order_type = request.args.get('type')
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    search_partner = request.args.get('search_partner', '')
    partner_id = request.args.get('partner_id', type=int)
    payment_method = request.args.get('payment_method')
    debt_cycle = request.args.get('debt_cycle', 'false').lower() == 'true'
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product_id = request.args.get('product_id', type=int)
    search_id = request.args.get('search_id', '')
    search_product = request.args.get('search_product', '')
    min_price = request.args.get('minPrice', type=float)
    max_price = request.args.get('maxPrice', type=float)
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)
    shipping_status = request.args.get('shipping_status')
    
    delivered_year = request.args.get('delivered_year')
    delivered_month = request.args.get('delivered_month')
    delivered_day = request.args.get('delivered_day')
    
    query = Order.query

    search = request.args.get('search', '')
    if search:
        s_norm = remove_accents(search)
        query = query.outerjoin(Partner).outerjoin(OrderDetail).outerjoin(Product).filter(
            or_(
                Order.display_id.ilike(f'%{search}%'),
                db.cast(Order.id, db.String).ilike(f'%{search}%'),
                db.func.remove_accents(db.func.coalesce(Partner.name, 'KHÁCH LẺ')).ilike(f'%{s_norm}%'),
                db.func.remove_accents(db.func.coalesce(Partner.phone, '')).ilike(f'%{search}%'),
                OrderDetail.product_name_override.ilike(f'%{search}%'),
                db.func.remove_accents(db.func.coalesce(Product.name, '')).ilike(f'%{s_norm}%')
            )
        ).distinct()

    if start_date:
        try:
            dt = datetime.fromisoformat(start_date)
            query = query.filter(Order.date >= dt)
        except: pass
    if end_date:
        try:
            dt = datetime.fromisoformat(end_date)
            if len(end_date) <= 10:
                dt = dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Order.date <= dt)
        except: pass

    if partner_id is not None:
        if partner_id == 0:
            query = query.filter(Order.partner_id == None)
        else:
            query = query.filter(Order.partner_id == partner_id)
        
        if debt_cycle:
            # Simple fallback for debt_cycle if dates not provided
            first_debt_order = Order.query.filter(
                Order.partner_id == (None if partner_id == 0 else partner_id),
                Order.payment_method == 'Debt',
                Order.display_id.notin_(['NODAU', '#NODAU'])
            ).order_by(Order.date.asc()).first()
            if first_debt_order:
                query = query.filter(Order.date >= first_debt_order.date)

    if order_type:
        query = query.filter(Order.type == order_type)
    
    if payment_method:
        query = query.filter(Order.payment_method == payment_method)

    is_consignment = request.args.get('is_consignment')
    if is_consignment:
        is_consignment_bool = is_consignment.lower() == 'true'
        query = query.filter(Order.is_consignment == is_consignment_bool)
    
    if product_id:
        query = query.join(OrderDetail).filter(OrderDetail.product_id == product_id).distinct()
    
    if search_partner:
        s_norm = remove_accents(search_partner)
        query = query.outerjoin(Partner).filter(db.func.remove_accents(db.func.coalesce(Partner.name, 'KHÁCH LẺ')).ilike(f'%{s_norm}%'))
    
    if search_id:
        query = query.filter(Order.display_id.ilike(f'%{search_id}%') | db.cast(Order.id, db.String).ilike(f'%{search_id}%'))
    
    if search_product:
        s_norm = remove_accents(search_product)
        query = query.join(OrderDetail).outerjoin(Product).filter(
            or_(
                OrderDetail.product_name_override.ilike(f'%{search_product}%'),
                db.func.remove_accents(db.func.coalesce(Product.name, '')).ilike(f'%{s_norm}%')
            )
        ).distinct()
    
    # Price filtering
    if min_price is not None:
        query = query.filter(Order.total_amount >= min_price)
    if max_price is not None:
        query = query.filter(Order.total_amount <= max_price)

    if shipping_status:
        if shipping_status == 'any':
            query = query.filter(Order.shipping_status != None)
        else:
            query = query.filter(Order.shipping_status == shipping_status)

    # Date filtering
    if year:
        query = query.filter(extract('year', Order.date) == int(year))
    if month:
        query = query.filter(extract('month', Order.date) == int(month))
    if day:
        query = query.filter(extract('day', Order.date) == int(day))
    
    # Delivered Date filtering (Revised for legacy support)
    if delivered_year or delivered_month or delivered_day:
        delivered_conds = []
        if delivered_year: delivered_conds.append(extract('year', Order.delivery_date) == int(delivered_year))
        if delivered_month: delivered_conds.append(extract('month', Order.delivery_date) == int(delivered_month))
        if delivered_day: delivered_conds.append(extract('day', Order.delivery_date) == int(delivered_day))
        
        legacy_conds = []
        if delivered_year: legacy_conds.append(extract('year', Order.date) == int(delivered_year))
        if delivered_month: legacy_conds.append(extract('month', Order.date) == int(delivered_month))
        if delivered_day: legacy_conds.append(extract('day', Order.date) == int(delivered_day))
        
        query = query.filter(
            or_(
                and_(*delivered_conds),
                and_(Order.delivery_date.is_(None), *legacy_conds)
            )
        )
    
    quarter = request.args.get('quarter', type=int)
    if quarter:
        if quarter == 1:
            query = query.filter(extract('month', Order.date).in_([1, 2, 3]))
        elif quarter == 2:
            query = query.filter(extract('month', Order.date).in_([4, 5, 6]))
        elif quarter == 3:
            query = query.filter(extract('month', Order.date).in_([7, 8, 9]))
        elif quarter == 4:
            query = query.filter(extract('month', Order.date).in_([10, 11, 12]))
        
    sort_by = request.args.get('sort_by', 'date')
    sort_order = request.args.get('sort_order', 'desc')

    # Mapping frontend sort keys to DB columns/expressions
    sort_map = {
        'id': Order.display_id,
        'date': Order.date,
        'partner_name': db.func.coalesce(Partner.name, 'KHÁCH LẺ'),
        'total_amount': Order.total_amount,
        'payment_method': Order.payment_method
    }

    if sort_by == 'partner_name':
        query = query.outerjoin(Partner)
        
    sort_col = sort_map.get(sort_by, Order.date)
    if sort_order == 'desc':
        query = query.order_by(sort_col.desc())
    else:
        query = query.order_by(sort_col.asc())

    if not partner_id and (not search_id or 'NODAU' not in search_id.upper()):
        query = query.filter(Order.display_id.notin_(['NODAU', '#NODAU']))

    if limit:
        # Flask-SQLAlchemy pagination (Default to page 1 if only limit provided)
        p_val = page if page else 1
        pagination = query.paginate(page=p_val, per_page=limit, error_out=False)
        return jsonify({
            'items': [o.to_dict() for o in pagination.items],
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': pagination.page
        })
    
    orders = query.all()
    return jsonify([o.to_dict() for o in orders])

@app.route('/api/history/active-filters', methods=['GET'])
def get_history_active_filters():
    order_type = request.args.get('type')
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    
    # Optimized Partner Query: Only partners with orders in this period
    partners_query = db.session.query(Partner).join(Order).filter(Order.partner_id == Partner.id)
    if order_type: partners_query = partners_query.filter(Order.type == order_type)
    if year and year != '':
        partners_query = partners_query.filter(extract('year', Order.date) == int(year))
    if month and month != '':
        partners_query = partners_query.filter(extract('month', Order.date) == int(month))
    if day and day != '':
        partners_query = partners_query.filter(extract('day', Order.date) == int(day))
    active_partners = partners_query.distinct().all()

    # Optimized Product Query: Only products with order details in this period
    products_query = db.session.query(Product).join(OrderDetail).join(Order).filter(
        OrderDetail.order_id == Order.id, 
        OrderDetail.product_id == Product.id
    )
    if order_type: products_query = products_query.filter(Order.type == order_type)
    if year and year != '':
        products_query = products_query.filter(extract('year', Order.date) == int(year))
    if month and month != '':
        products_query = products_query.filter(extract('month', Order.date) == int(month))
    if day and day != '':
        products_query = products_query.filter(extract('day', Order.date) == int(day))
    active_products = products_query.distinct().all()

    return jsonify({
        'partners': [p.to_dict() for p in active_partners],
        'products': [p.to_dict() for p in active_products]
    })

# --- REPORT AGGREGATION ENDPOINTS ---
@app.route('/api/reports/kpis', methods=['GET'])
def get_report_kpis():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        brand = request.args.get('brand', 'All')
        
        # Base query for Sale orders (excluding opening balance orders)
        order_query = Order.query.filter(Order.type == 'Sale', Order.display_id.notin_(['NODAU', '#NODAU']))
        
        if start_date:
            order_query = order_query.filter(Order.date >= datetime.fromisoformat(start_date))
        if end_date:
            dt_end = datetime.fromisoformat(end_date)
            if len(end_date) <= 10: dt_end = dt_end.replace(hour=23, minute=59, second=59)
            order_query = order_query.filter(Order.date <= dt_end)

        # Optimization: Get list of order IDs first or use subquery
        order_ids_subquery = order_query.with_entities(Order.id)
        
        # 1. Total Revenue & Order Count
        # If brand filtering is needed, we must join with OrderDetail
        if brand != 'All':
            stats = db.session.query(
                func.sum(OrderDetail.price * OrderDetail.quantity).label('total_revenue'),
                func.count(func.distinct(Order.id)).label('order_count')
            ).join(Order).join(Product, OrderDetail.product_id == Product.id)\
             .filter(Order.id.in_(order_ids_subquery), Product.brand == brand).first()
        else:
            stats = db.session.query(
                func.sum(Order.total_amount).label('total_revenue'),
                func.count(Order.id).label('order_count')
            ).filter(Order.id.in_(order_ids_subquery)).first()

        # 2. Total Profit & Product Count
        detail_query = db.session.query(
            func.sum((OrderDetail.price - func.coalesce(OrderDetail.cost_price, 0)) * OrderDetail.quantity).label('total_profit'),
            func.count(func.distinct(OrderDetail.product_id)).label('product_count'),
            func.sum(OrderDetail.quantity).label('total_qty')
        ).join(Order).filter(Order.id.in_(order_ids_subquery))
        
        if brand != 'All':
            detail_query = detail_query.join(Product, OrderDetail.product_id == Product.id).filter(Product.brand == brand)
            
        profit_stats = detail_query.first()

        return jsonify({
            'total_revenue': stats.total_revenue or 0,
            'total_profit': profit_stats.total_profit or 0,
            'order_count': stats.order_count or 0,
            'product_count': profit_stats.product_count or 0,
            'total_qty': profit_stats.total_qty or 0
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/reports/sales-chart', methods=['GET'])
def get_report_sales_chart():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = Order.query.filter(Order.type == 'Sale', Order.display_id.notin_(['NODAU', '#NODAU']))
        if start_date: query = query.filter(Order.date >= datetime.fromisoformat(start_date))
        if end_date:
            dt_end = datetime.fromisoformat(end_date)
            if len(end_date) <= 10: dt_end = dt_end.replace(hour=23, minute=59, second=59)
            query = query.filter(Order.date <= dt_end)
            
        # Group by Date (YYYY-MM-DD)
        # For SQLite, date() function is available
        date_func = func.date(Order.date)
        
        # Revenue by day
        revenue_data = db.session.query(
            date_func.label('day'),
            func.sum(Order.total_amount).label('revenue')
        ).filter(Order.id.in_(query.with_entities(Order.id)))\
         .group_by('day').order_by('day').all()
        
        # Profit by day (needs join)
        profit_data = db.session.query(
            func.date(Order.date).label('day'),
            func.sum((OrderDetail.price - func.coalesce(OrderDetail.cost_price, 0)) * OrderDetail.quantity).label('profit')
        ).join(Order).filter(Order.id.in_(query.with_entities(Order.id)))\
         .group_by('day').all()
         
        profit_map = {row.day: row.profit for row in profit_data}
        
        result = []
        for row in revenue_data:
            result.append({
                'date': row.day,
                'revenue': row.revenue,
                'profit': profit_map.get(row.day, 0)
            })
            
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/reports/purchase-chart', methods=['GET'])
def get_report_purchase_chart():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = Order.query.filter(Order.type == 'Purchase', Order.display_id.notin_(['NODAU', '#NODAU']))
        if start_date: query = query.filter(Order.date >= datetime.fromisoformat(start_date))
        if end_date:
            dt_end = datetime.fromisoformat(end_date)
            if len(end_date) <= 10: dt_end = dt_end.replace(hour=23, minute=59, second=59)
            query = query.filter(Order.date <= dt_end)
            
        date_func = func.date(Order.date)
        
        spending_data = db.session.query(
            date_func.label('day'),
            func.sum(Order.total_amount).label('spending')
        ).filter(Order.id.in_(query.with_entities(Order.id)))\
         .group_by('day').order_by('day').all()
        
        return jsonify([{
            'date': row.day,
            'spending': row.spending
        } for row in spending_data])
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/reports/product-sales', methods=['GET'])
def get_report_product_sales():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        brand = request.args.get('brand', 'All')
        limit = request.args.get('limit', 50, type=int)
        
        search = request.args.get('search')
        sort_by = request.args.get('sort_by', 'revenue')
        sort_order = request.args.get('sort_order', 'desc')
        
        order_query = Order.query.filter(Order.type == 'Sale', Order.display_id.notin_(['NODAU', '#NODAU']))
        if start_date: order_query = order_query.filter(Order.date >= datetime.fromisoformat(start_date))
        if end_date:
            dt_end = datetime.fromisoformat(end_date)
            if len(end_date) <= 10: dt_end = dt_end.replace(hour=23, minute=59, second=59)
            order_query = order_query.filter(Order.date <= dt_end)

        # Base Query
        query = db.session.query(
            OrderDetail.product_id,
            func.max(OrderDetail.product_name_override).label('name'), 
            func.sum(OrderDetail.quantity).label('qty'),
            func.sum(OrderDetail.price * OrderDetail.quantity).label('revenue'),
            func.sum((OrderDetail.price - func.coalesce(OrderDetail.cost_price, 0)) * OrderDetail.quantity).label('profit')
        ).join(Order).filter(Order.id.in_(order_query.with_entities(Order.id)))
        
        if brand != 'All':
            query = query.join(Product, OrderDetail.product_id == Product.id).filter(Product.brand == brand)
            
        if search:
            if brand == 'All': query = query.join(Product, OrderDetail.product_id == Product.id)
            s_norm = remove_accents(search)
            query = query.filter(
                or_(
                    db.func.remove_accents(Product.name).ilike(f'%{s_norm}%'),
                    Product.code.ilike(f'%{search}%')
                )
            )

        # Dynamic Sorting
        sort_col = func.sum(OrderDetail.price * OrderDetail.quantity) # default revenue
        if sort_by == 'name': sort_col = func.max(OrderDetail.product_name_override)
        elif sort_by == 'qty': sort_col = func.sum(OrderDetail.quantity)
        elif sort_by == 'profit': sort_col = func.sum((OrderDetail.price - func.coalesce(OrderDetail.cost_price, 0)) * OrderDetail.quantity)
        
        if sort_order == 'asc':
            query = query.group_by(OrderDetail.product_id).order_by(sort_col.asc())
        else:
            query = query.group_by(OrderDetail.product_id).order_by(sort_col.desc())

        report = query.limit(limit).all()
        
        # Get product unit/code for final result
        p_ids = [r.product_id for r in report if r.product_id]
        products = {p.id: p for p in Product.query.filter(Product.id.in_(p_ids)).all()}
        
        return jsonify([{
            'id': r.product_id,
            'name': r.name,
            'code': products.get(r.product_id).code if products.get(r.product_id) else '',
            'unit': products.get(r.product_id).unit if products.get(r.product_id) else '',
            'qty': r.qty,
            'revenue': r.revenue,
            'profit': r.profit
        } for r in report])
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/reports/partner-sales', methods=['GET'])
def get_report_partner_sales():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        search = request.args.get('search')
        sort_by = request.args.get('sort_by', 'total_revenue')
        sort_order = request.args.get('sort_order', 'desc')
        
        order_query = Order.query.filter(Order.type == 'Sale', Order.display_id.notin_(['NODAU', '#NODAU']), Order.partner_id != None)
        if start_date: order_query = order_query.filter(Order.date >= datetime.fromisoformat(start_date))
        if end_date:
            dt_end = datetime.fromisoformat(end_date)
            if len(end_date) <= 10: dt_end = dt_end.replace(hour=23, minute=59, second=59)
            order_query = order_query.filter(Order.date <= dt_end)

        if search:
            s_norm = remove_accents(search)
            order_query = order_query.join(Partner).filter(
                or_(
                    db.func.remove_accents(Partner.name).ilike(f'%{s_norm}%'),
                    Partner.phone.ilike(f'%{search}%')
                )
            )

        # Revenue and Order Count per Partner
        summary = db.session.query(
            Order.partner_id,
            func.count(Order.id).label('order_count'),
            func.sum(Order.total_amount).label('total_revenue')
        ).filter(Order.id.in_(order_query.with_entities(Order.id)))\
         .group_by(Order.partner_id).all()
         
        # Profit per Partner
        profit_data = db.session.query(
            Order.partner_id,
            func.sum((OrderDetail.price - func.coalesce(OrderDetail.cost_price, 0)) * OrderDetail.quantity).label('total_profit')
        ).join(Order).filter(Order.id.in_(order_query.with_entities(Order.id)))\
         .group_by(Order.partner_id).all()
        profit_map = {r.partner_id: r.total_profit for r in profit_data}
        
        # Debt changes per Partner
        debt_data = db.session.query(
            Order.partner_id,
            func.sum(Order.total_amount - func.coalesce(Order.amount_paid, 0)).label('debt_increase')
        ).filter(Order.id.in_(order_query.with_entities(Order.id)), Order.payment_method == 'Debt')\
         .group_by(Order.partner_id).all()
        debt_map = {r.partner_id: r.debt_increase for r in debt_data}

        p_ids = [s.partner_id for s in summary]
        partners = {p.id: p for p in Partner.query.filter(Partner.id.in_(p_ids)).all()}
        
        result = []
        for s in summary:
            p = partners.get(s.partner_id)
            result.append({
                'id': s.partner_id,
                'name': p.name if p else 'Khách lạ',
                'phone': p.phone if p else '',
                # Dual-casing to be absolutely safe for frontend
                'orderCount': int(s.order_count or 0),
                'order_count': int(s.order_count or 0),
                'totalRevenue': float(s.total_revenue or 0),
                'total_revenue': float(s.total_revenue or 0),
                'totalProfit': float(profit_map.get(s.partner_id, 0)),
                'debtIncrease': float(debt_map.get(s.partner_id, 0)),
                'debt_increase': float(debt_map.get(s.partner_id, 0)),
                'totalDebt': float(p.debt_balance or 0)
            })
            
        is_reverse = (sort_order == 'desc')
        # Default to totalRevenue if sort_by not in list
        valid_sorts = ['name', 'orderCount', 'totalRevenue', 'totalProfit', 'debtIncrease', 'totalDebt']
        if sort_by not in valid_sorts:
            sort_by = 'totalRevenue'
            
        return jsonify(sorted(result, key=lambda x: x.get(sort_by) or 0, reverse=is_reverse))
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/reports/inventory-flow', methods=['GET'])
def get_report_inventory_flow():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        brand = request.args.get('brand', 'All')
        search = request.args.get('search')
        sort_by = request.args.get('sort_by', 'exportQty')
        sort_order = request.args.get('sort_order', 'desc')
        
        order_query = Order.query.filter(Order.display_id.notin_(['NODAU', '#NODAU']))
        if start_date: order_query = order_query.filter(Order.date >= datetime.fromisoformat(start_date))
        if end_date:
            dt_end = datetime.fromisoformat(end_date)
            if len(end_date) <= 10: dt_end = dt_end.replace(hour=23, minute=59, second=59)
            order_query = order_query.filter(Order.date <= dt_end)

        # Import Data (Purchases)
        import_data = db.session.query(
            OrderDetail.product_id,
            func.sum(OrderDetail.quantity).label('qty'),
            func.sum(OrderDetail.price * OrderDetail.quantity).label('val')
        ).join(Order).filter(Order.id.in_(order_query.filter(Order.type == 'Purchase').with_entities(Order.id)))\
         .group_by(OrderDetail.product_id).all()
        
        # Export Data (Sales)
        export_data = db.session.query(
            OrderDetail.product_id,
            func.sum(OrderDetail.quantity).label('qty'),
            func.sum(OrderDetail.price * OrderDetail.quantity).label('val')
        ).join(Order).filter(Order.id.in_(order_query.filter(Order.type == 'Sale').with_entities(Order.id)))\
         .group_by(OrderDetail.product_id).all()

        import_map = {row.product_id: row for row in import_data}
        export_map = {row.product_id: row for row in export_data}

        # Get filtered products
        p_query = Product.query
        if brand != 'All': p_query = p_query.filter(Product.brand == brand)
        if search:
            s_norm = remove_accents(search)
            p_query = p_query.filter(
                or_(
                    db.func.remove_accents(Product.name).ilike(f'%{s_norm}%'),
                    Product.code.ilike(f'%{search}%')
                )
            )
        products = p_query.all()

        result = []
        for p in products:
            im = import_map.get(p.id)
            ex = export_map.get(p.id)
            if not im and not ex and p.stock == 0: continue

            result.append({
                'id': p.id,
                'code': p.code,
                'name': p.name,
                'unit': p.unit,
                'currentStock': p.stock,
                'importQty': im.qty if im else 0,
                'exportQty': ex.qty if ex else 0,
                'importVal': im.val if im else 0,
                'exportVal': ex.val if ex else 0,
                'openingStock': p.stock - (im.qty if im else 0) + (ex.qty if ex else 0)
            })

        is_reverse = (sort_order == 'desc')
        # Map frontend sort keys to result object keys if they differ
        valid_sorts = ['name', 'openingStock', 'importQty', 'exportQty', 'currentStock']
        if sort_by not in valid_sorts: sort_by = 'exportQty'
        
        return jsonify(sorted(result, key=lambda x: x.get(sort_by) or 0, reverse=is_reverse))
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/reports/brands', methods=['GET'])
def get_report_brands():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        sort_by = request.args.get('sort_by', 'revenue')
        sort_order = request.args.get('sort_order', 'desc')
        
        order_query = Order.query.filter(Order.type == 'Sale', Order.display_id.notin_(['NODAU', '#NODAU']))
        if start_date: order_query = order_query.filter(Order.date >= datetime.fromisoformat(start_date))
        if end_date:
            dt_end = datetime.fromisoformat(end_date)
            if len(end_date) <= 10: dt_end = dt_end.replace(hour=23, minute=59, second=59)
            order_query = order_query.filter(Order.date <= dt_end)

        # Dynamic Sorting
        sort_map = {
            'revenue': func.sum(OrderDetail.price * OrderDetail.quantity),
            'profit': func.sum((OrderDetail.price - func.coalesce(OrderDetail.cost_price, 0)) * OrderDetail.quantity),
            'qty': func.sum(OrderDetail.quantity),
            'name': Product.brand
        }
        sort_col = sort_map.get(sort_by, sort_map['revenue'])
        if sort_order == 'asc': sort_col = sort_col.asc()
        else: sort_col = sort_col.desc()

        report = db.session.query(
            Product.brand,
            func.sum(OrderDetail.price * OrderDetail.quantity).label('revenue'),
            func.sum((OrderDetail.price - func.coalesce(OrderDetail.cost_price, 0)) * OrderDetail.quantity).label('profit'),
            func.sum(OrderDetail.quantity).label('qty')
        ).join(OrderDetail, Product.id == OrderDetail.product_id)\
         .join(Order, OrderDetail.order_id == Order.id)\
         .filter(Order.id.in_(order_query.with_entities(Order.id)))\
         .group_by(Product.brand).order_by(sort_col).all()

        return jsonify([{
            'name': r.brand or 'Khác',
            'revenue': r.revenue,
            'profit': r.profit,
            'qty': r.qty
        } for r in report])
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/reports/purchase-sales', methods=['GET'])
def get_report_purchase_sales():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        search = request.args.get('search')
        sort_by = request.args.get('sort_by', 'totalImport')
        sort_order = request.args.get('sort_order', 'desc')
        
        order_query = Order.query.filter(Order.type == 'Purchase', Order.display_id.notin_(['NODAU', '#NODAU']))
        if start_date: order_query = order_query.filter(Order.date >= datetime.fromisoformat(start_date))
        if end_date:
            dt_end = datetime.fromisoformat(end_date)
            if len(end_date) <= 10: dt_end = dt_end.replace(hour=23, minute=59, second=59)
            order_query = order_query.filter(Order.date <= dt_end)
            
        if search:
            s_norm = remove_accents(search)
            order_query = order_query.outerjoin(Partner).filter(
                or_(
                    db.func.remove_accents(Partner.name).ilike(f'%{s_norm}%'),
                    Partner.phone.ilike(f'%{search}%')
                )
            )

        summary = db.session.query(
            Order.partner_id,
            func.count(Order.id).label('importCount'),
            func.sum(Order.total_amount).label('totalImport')
        ).filter(Order.id.in_(order_query.with_entities(Order.id)))\
         .group_by(Order.partner_id).all()
         
        p_ids = [s.partner_id for s in summary if s.partner_id]
        partners = {p.id: p for p in Partner.query.filter(Partner.id.in_(p_ids)).all()}
        
        result = []
        for s in summary:
            p = partners.get(s.partner_id)
            result.append({
                'id': s.partner_id,
                'name': p.name if p else 'Nhà cung cấp lạ',
                'phone': p.phone if p else '',
                'importCount': s.importCount,
                'totalImport': s.totalImport
            })
            
        is_reverse = (sort_order == 'desc')
        valid_sorts = ['name', 'importCount', 'totalImport']
        if sort_by not in valid_sorts: sort_by = 'totalImport'
        
        return jsonify(sorted(result, key=lambda x: x.get(sort_by) or 0, reverse=is_reverse))
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/orders/<int:order_id>/status', methods=['PATCH'])
def update_order_status(order_id):
    order = Order.query.get_or_404(order_id)
    data = request.json
    if 'status' in data:
        order.status = data['status']
        db.session.commit()
    return jsonify(order.to_dict())

@app.route('/api/orders/<int:order_id>/shipping-status', methods=['PATCH'])
def update_shipping_status(order_id):
    order = Order.query.get_or_404(order_id)
    data = request.json
    if 'shipping_status' in data:
        new_status = data['shipping_status']
        if new_status == 'Delivered' and order.shipping_status != 'Delivered':
            order.delivery_date = get_vn_time()
        elif new_status != 'Delivered':
            order.delivery_date = None
            
        if new_status == 'Shipping':
            for detail in order.details:
                detail.shipped_quantity = 0
                
        order.shipping_status = new_status
        db.session.commit()
    return jsonify(order.to_dict())

@app.route('/api/order-details/<int:detail_id>/shipped-quantity', methods=['PATCH'])
def update_detail_shipped_quantity(detail_id):
    detail = OrderDetail.query.get_or_404(detail_id)
    data = request.json
    new_qty = data.get('shipped_quantity', 0)
    
    # Update current line
    detail.shipped_quantity = new_qty
    
    # Check overall order status
    order = detail.order
    all_details = order.details
    all_shipped = True
    any_shipped = False
    
    for d in all_details:
        if d.shipped_quantity < d.quantity:
            all_shipped = False
        if d.shipped_quantity > 0:
            any_shipped = True
            
    if all_shipped:
        if order.shipping_status != 'Delivered':
            order.delivery_date = get_vn_time()
        order.shipping_status = 'Delivered'
    elif any_shipped:
        order.delivery_date = None
        order.shipping_status = 'Shipping'
    else:
        # If user resets all to 0
        order.delivery_date = None
        # Typically once marked as Shipping, it stays Shipping.
        pass
        
    db.session.commit()
    return jsonify({
        'detail': detail.to_dict(),
        'order_shipping_status': order.shipping_status
    })

@app.route('/api/orders/<int:order_id>/import-consignment', methods=['POST'])
def import_consignment(order_id):
    order = Order.query.get_or_404(order_id)
    if order.type != 'Purchase':
        return jsonify({'error': 'Đơn hàng không phải là đơn nhập hàng.'}), 400
    
    data = request.json
    # Expected body: { details: [{product_id, quantity}] }
    imports = data.get('details', [])
    if not imports:
        return jsonify({'error': 'Không có sản phẩm nào được chọn để nhập kho.'}), 400
        
    local_now = get_vn_time()
    date_str = local_now.strftime('%d/%m/%Y %H:%M')
    log_entries = []
    imported_product_ids = set()
    
    try:
        for item in imports:
            product_id = item.get('product_id')
            qty_to_import = float(item.get('quantity', 0))
            if qty_to_import <= 0:
                continue
                
            # Find the corresponding OrderDetail
            detail = next((d for d in order.details if d.product_id == product_id), None)
            if not detail:
                continue
                
            prod = Product.query.get(product_id)
            if not prod:
                continue
                
            # Calculate remaining consignment quantity
            remaining = detail.quantity - detail.shipped_quantity
            if remaining <= 0:
                continue
                
            # Limit import to remaining consignment quantity
            actual_import = min(qty_to_import, remaining)
            if actual_import <= 0:
                continue
                
            # Update detail and product stock
            detail.shipped_quantity += actual_import
            imported_product_ids.add(prod.id)
            
            # Increase stock (combo handles child stocks)
            if prod.is_combo:
                for ci in prod.combo_items:
                    child = Product.query.get(ci.product_id)
                    if child:
                        child.stock += (actual_import * ci.quantity)
                        # Create StockBatch for child
                        new_batch = StockBatch(
                            product_id=child.id,
                            purchase_order_id=order.id,
                            original_quantity=actual_import * ci.quantity,
                            current_quantity=actual_import * ci.quantity,
                            cost_price=detail.price / ci.quantity if ci.quantity > 0 else 0
                        )
                        db.session.add(new_batch)
            else:
                prod.stock += actual_import
                # Create StockBatch
                new_batch = StockBatch(
                    product_id=prod.id,
                    purchase_order_id=order.id,
                    original_quantity=actual_import,
                    current_quantity=actual_import,
                    cost_price=detail.price
                )
                db.session.add(new_batch)
                
            log_entries.append(f"Nhập {actual_import} {prod.unit or 'ĐV'} {prod.name}")
            
        if log_entries:
            # Append log to order note
            log_text = f"\n- [{date_str}] " + ", ".join(log_entries)
            if order.note:
                order.note += log_text
            else:
                order.note = log_text.strip()
                
            # Update shipping_status
            all_shipped = True
            any_shipped = False
            for d in order.details:
                if d.shipped_quantity < d.quantity:
                    all_shipped = False
                if d.shipped_quantity > 0:
                    any_shipped = True
                    
            if all_shipped:
                order.shipping_status = 'Delivered'
                order.delivery_date = local_now
            elif any_shipped:
                order.shipping_status = 'Shipping'
                
            for pid in imported_product_ids:
                recalculate_product_cost_price(pid)
                
            db.session.commit()
            return jsonify({
                'message': 'Đã nhập kho thành công!',
                'order': order.to_dict()
            }), 200
        else:
            return jsonify({'error': 'Không có sản phẩm hợp lệ nào được nhập.'}), 400
            
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error importing consignment: {e}")
        return jsonify({'error': f'Lỗi hệ thống: {str(e)}'}), 500

@app.route('/api/orders/duplicates', methods=['GET'])
def get_duplicate_orders():
    # Chỉ kiểm tra đơn Sale phát sinh trong CÙNG NGÀY hôm nay
    local_now = get_vn_time()
    start_date = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Chỉ lấy các đơn CHƯA được đánh dấu là "đã kiểm tra" (is_duplicate_checked == False)
    orders = Order.query.filter(
        Order.type == 'Sale',
        Order.date >= start_date,
        Order.display_id.notin_(['NODAU', '#NODAU']),
        Order.is_duplicate_checked == False
    ).all()
    
    groups = {}
    for o in orders:
        if not o.details: continue
        
        # Tạo fingerprint ổn định từ chi tiết sản phẩm
        details_list = []
        for d in o.details:
            details_list.append({
                'p_id': d.product_id,
                'p_name': (d.product_name_override or "").strip(),
                'qty': float(d.quantity),
                'price': float(d.price)
            })
        
        # Sắp xếp để đảm bảo thứ tự nhập không ảnh hưởng đến kết quả so sánh
        details_list.sort(key=lambda x: (str(x['p_id'] or ''), x['p_name'], x['qty'], x['price']))
        fingerprint = json.dumps(details_list, sort_keys=True)
        
        # Khóa nhóm: Fingerprint sản phẩm + Tổng tiền (làm tròn để tránh sai số float)
        group_key = (fingerprint, round(float(o.total_amount), 2))
        
        if group_key not in groups:
            groups[group_key] = []
        groups[group_key].append(o)
    
    duplicates = []
    for key, o_list in groups.items():
        if len(o_list) > 1:
            # Chuẩn hóa tên người tạo để tránh lỗi do khoảng trắng hoặc viết hoa/thường
            # Chỉ tính các đơn có người tạo rõ ràng (không trống)
            user_map = {}
            for o in o_list:
                creator = (o.created_by or "Chưa rõ").strip().lower()
                if creator not in user_map:
                    user_map[creator] = []
                user_map[creator].append(o)
            
            # Nếu có từ 2 danh tính người nhập khác nhau trở lên
            if len(user_map) > 1:
                duplicates.append([o.to_dict() for o in o_list])
    
    return jsonify(duplicates)

@app.route('/api/orders/<int:order_id>/check-duplicate', methods=['POST'])
def check_duplicate_order(order_id):
    order = Order.query.get_or_404(order_id)
    order.is_duplicate_checked = True
    db.session.commit()
    return jsonify({'status': 'success', 'message': f'Order {order_id} marked as checked'})

@app.route('/api/orders', methods=['POST'])
def create_order():
    data = request.json
    # Expected: { partner_id, type: 'Sale'|'Purchase', payment_method: 'Cash'|'Debt', details: [{product_id, quantity, price}] }
    
    try:
        # Custom Order ID Generation (N.DD/MM/YY)
        local_now = get_vn_time()
        today_str = local_now.strftime('%d/%m/%y')
        start_of_day = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = local_now.replace(hour=23, minute=59, second=59, microsecond=999999)
        order_type = data.get('type', 'Sale')
        count_today = Order.query.filter(Order.date >= start_of_day, Order.date <= end_of_day)\
                                 .filter(Order.type == order_type)\
                                 .filter(Order.display_id.notin_(['NODAU', '#NODAU']))\
                                 .count()
        display_id = f"{count_today + 1}.{today_str}"
        
        is_consignment_order = bool(data.get('is_consignment', False))
        new_order = Order(
            date=local_now,
            partner_id=data.get('partner_id'),
            type=data['type'],
            payment_method=data['payment_method'],
            display_id=display_id,
            total_amount=0, # will calc
            note=data.get('note'),
            amount_paid=data.get('amount_paid', 0),
            shipping_status=data.get('shipping_status'),
            shipping_address=data.get('shipping_address'),
            shipping_phone=data.get('shipping_phone'),
            cash_given=data.get('cash_given', 0),
            created_by=data.get('created_by'),
            is_consignment=is_consignment_order
        )
        
        total = 0
        affected_product_ids = []
        created_batches = []
        for item in data['details']:
            product_id = item.get('product_id')
            if product_id is None:
                # Custom item
                detail = OrderDetail(
                    product_id=None,
                    product_name_override=item.get('product_name') or item.get('name'),
                    quantity=item['quantity'],
                    price=item['price']
                )
                new_order.details.append(detail)
                total += item['quantity'] * item['price']
                continue

            prod = Product.query.get(product_id)
            if not prod:
                raise Exception(f"Product {product_id} not found")
            
            # Inventory Management & FIFO
            item_qty = float(item['quantity'])
            item_price = float(item['price'])
            
            if data['type'] == 'Sale':
                if item_qty < 0:
                    # Return transaction: do NOT run FIFO
                    prod.stock -= int(item_qty)
                    if prod.is_combo:
                        total_combo_cost = 0
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                total_combo_cost += ci.quantity * (child.cost_price or 0)
                        avg_cost = total_combo_cost
                    else:
                        avg_cost = prod.cost_price or 0
                else:
                    # FIFO Calculation for Profit Locking
                    if prod.is_combo:
                        total_combo_cost = 0
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                needed_qty = item_qty * ci.quantity
                                child.stock -= int(needed_qty)
                                
                                # FIFO for this child component
                                child_cost = 0
                                remaining_needed = needed_qty
                                batches = StockBatch.query.filter(StockBatch.product_id == child.id, StockBatch.current_quantity > 0)\
                                                          .order_by(StockBatch.created_at.asc()).all()
                                
                                for batch in batches:
                                    if remaining_needed <= 0: break
                                    take = min(remaining_needed, batch.current_quantity)
                                    batch.current_quantity -= take
                                    child_cost += take * batch.cost_price
                                    remaining_needed -= take
                                
                                if remaining_needed > 0:
                                    # Fallback to current cost price for missing stock
                                    child_cost += remaining_needed * (child.cost_price or 0)
                                
                                total_combo_cost += child_cost
                        avg_cost = total_combo_cost / item_qty if item_qty > 0 else 0
                    else:
                        # Simple Product FIFO
                        prod.stock -= int(item_qty)
                        
                        total_sale_cost = 0
                        remaining_needed = item_qty
                        batches = StockBatch.query.filter(StockBatch.product_id == prod.id, StockBatch.current_quantity > 0)\
                                                  .order_by(StockBatch.created_at.asc()).all()
                        
                        for batch in batches:
                            if remaining_needed <= 0: break
                            take = min(remaining_needed, batch.current_quantity)
                            batch.current_quantity -= take
                            total_sale_cost += take * batch.cost_price
                            remaining_needed -= take
                        
                        if remaining_needed > 0:
                            total_sale_cost += remaining_needed * (prod.cost_price or 0)
                        
                        avg_cost = total_sale_cost / item_qty if item_qty > 0 else 0
            
            elif data['type'] == 'Purchase':
                if is_consignment_order:
                    # Consignment order: stock is NOT increased and batch is NOT created initially!
                    pass
                else:
                    if prod.is_combo:
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                child_incoming = item_qty * ci.quantity
                                child_cost = item_price / ci.quantity if ci.quantity > 0 else 0
                                child_curr = adjust_negative_stock_backorder(child.id, child_incoming, child_cost)
                                child.stock += int(child_incoming)
                                # Create batch for child if we want to track components
                                # Usually we track components on purchase of components, but if buying combo, we split
                                new_batch = StockBatch(
                                    product_id=child.id,
                                    original_quantity=child_incoming,
                                    current_quantity=child_curr,
                                    cost_price=child_cost
                                )
                                db.session.add(new_batch)
                                created_batches.append(new_batch)
                    else:
                        curr_qty = adjust_negative_stock_backorder(prod.id, item_qty, item_price)
                        prod.stock += int(item_qty)
                        # Create new Stock Batch
                        new_batch = StockBatch(
                            product_id=prod.id,
                            original_quantity=item_qty,
                            current_quantity=curr_qty,
                            cost_price=item_price
                        )
                        db.session.add(new_batch)
                        created_batches.append(new_batch)
                        affected_product_ids.append(prod.id)
                
                avg_cost = item_price

            detail = OrderDetail(
                product_id=prod.id,
                product_name_override=item.get('product_name') or item.get('name'),
                quantity=item_qty,
                price=item_price,
                cost_price=avg_cost if data['type'] == 'Sale' else item_price,
                shipped_quantity=0 if (is_consignment_order or data.get('shipping_status')) else item_qty
            )
            new_order.details.append(detail)
            total += item_qty * item_price
        
        new_order.total_amount = total
        db.session.add(new_order)
        db.session.flush() # ID is now available
        
        # Explicitly link newly created batches to this purchase order
        for b in created_batches:
            b.purchase_order_id = new_order.id
        
        # Debt Management & Cash History connection
        if data.get('partner_id'):
            partner = Partner.query.get(data['partner_id'])
            if partner:
                new_order.old_debt = partner.debt_balance
                # NEW LOGIC: Only 'Debt' orders affect balance. 
                # Partial payments at POS do NOT reduce balance or create vouchers.
                if data.get('payment_method') == 'Debt':
                    if data['type'] == 'Sale':
                        partner.debt_balance += total
                        # If total < 0 (Return), amount_paid is what WE give back to CUSTOMER
                        upfront = float(data.get('amount_paid', 0))
                        if upfront > 0:
                            if total >= 0:
                                v_type = 'Receipt'
                                v_note = f"Thanh toán trước cho đơn {display_id}"
                                partner.debt_balance -= upfront
                            else:
                                v_type = 'Payment'
                                v_note = f"Chi trả tiền hàng cho đơn trả {display_id}"
                                partner.debt_balance += upfront
                            
                            v = CashVoucher(
                                partner_id=partner.id,
                                amount=upfront,
                                note=v_note,
                                type=v_type,
                                source='settlement',
                                order_id=new_order.id
                            )
                            db.session.add(v)
                    else:
                        partner.debt_balance -= total
                        # Purchase: upfront payment reduces the negative balance (we pay supplier)
                        # If total < 0 (Return), upfront is what SUPPLIER gives back to US
                        upfront = float(data.get('amount_paid', 0))
                        if upfront > 0:
                            if total >= 0:
                                v_type = 'Payment'
                                v_note = f"Thanh toán trước cho đơn nhập {display_id}"
                                partner.debt_balance += upfront
                            else:
                                v_type = 'Receipt'
                                v_note = f"Thu tiền hàng cho đơn nhập trả {display_id}"
                                partner.debt_balance -= upfront

                            v = CashVoucher(
                                partner_id=partner.id,
                                amount=upfront,
                                note=v_note,
                                type=v_type,
                                source='settlement',
                                order_id=new_order.id
                            )
                            db.session.add(v)
                # Manual vouchers in Fund tab are now the ONLY way to reduce debt.

        # --- Cash Payment Support (Sổ Tiền Mặt) ---
        if data.get('payment_method') == 'Cash':
            v_type = 'Receipt' if data['type'] == 'Sale' else 'Payment'
            v_note = f"Thu tiền bán lẻ - Đơn {display_id}" if data['type'] == 'Sale' else f"Chi tiền nhập hàng - Đơn {display_id}"
            if data.get('partner_id'):
                partner = Partner.query.get(data['partner_id'])
                if partner:
                    v_note = f"Thu tiền bán hàng - Đơn {display_id} ({partner.name})" if data['type'] == 'Sale' else f"Chi tiền nhập hàng - Đơn {display_id} ({partner.name})"
            
            v_amount = total
            if total < 0:
                v_amount = abs(total)
                v_type = 'Payment' if data['type'] == 'Sale' else 'Receipt'
                v_note = f"Chi trả tiền hàng trả - Đơn {display_id}" if data['type'] == 'Sale' else f"Thu tiền nhập hàng trả - Đơn {display_id}"

            v = CashVoucher(
                partner_id=data.get('partner_id'),
                amount=v_amount,
                note=v_note,
                type=v_type,
                source='auto',
                order_id=new_order.id
            )
            db.session.add(v)
            new_order.amount_paid = total

        # --- Bank Transaction Support ---
        if data.get('payment_method') == 'Transfer' and data.get('bank_account_id'):
            acc_id = int(data['bank_account_id'])
            bank_acc = BankAccount.query.get(acc_id)
            if bank_acc:
                upfront = float(data.get('amount_paid', 0))
                # For Transfer, if amount_paid is 0, we assume the whole total is transferred
                if upfront == 0:
                    upfront = total
                
                t_type = 'Deposit' if data['type'] == 'Sale' else 'Withdrawal'
                # If Sale and total < 0 (Return), it's a Withdrawal from bank
                if data['type'] == 'Sale' and total < 0:
                    t_type = 'Withdrawal'
                # If Purchase and total < 0 (Return), it's a Deposit to bank
                elif data['type'] == 'Purchase' and total < 0:
                    t_type = 'Deposit'

                bt = BankTransaction(
                    account_id=acc_id,
                    amount=abs(upfront),
                    type=t_type,
                    note=f"Thanh toán đơn {display_id}",
                    partner_id=data.get('partner_id'),
                    order_id=new_order.id
                )
                
                if t_type == 'Deposit':
                    bank_acc.balance += abs(upfront)
                else:
                    bank_acc.balance -= abs(upfront)
                
                db.session.add(bt)
                new_order.amount_paid = upfront # Update order state
        
        db.session.commit()
        
        # 4. Recalculate cost prices for products affected by this Purchase
        if new_order.type == 'Purchase':
            for pid in set(affected_product_ids):
                recalculate_product_cost_price(pid)
        
        # Final Recalculate
        if new_order.partner_id:
            recalculate_partner_debt_internal(new_order.partner_id)
            
        return jsonify(new_order.to_dict()), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/orders/<int:id>', methods=['GET'])
def get_order(id):
    order = Order.query.get_or_404(id)
    return jsonify(order.to_dict())

@app.route('/api/orders/<int:id>', methods=['DELETE'])
def delete_order(id):
    order = Order.query.get_or_404(id)
    partner_id = order.partner_id
    try:
        # 1. Reverse Inventory & FIFO Batches
        affected_product_ids = []
        if order.type == 'Purchase':
            # Find and delete batches created by this purchase
            batches_to_delete = StockBatch.query.filter_by(purchase_order_id=order.id).all()
            for b in batches_to_delete:
                affected_product_ids.append(b.product_id)
                db.session.delete(b)
        
        for detail in order.details:
            if detail.product_id is None:
                continue
            affected_product_ids.append(detail.product_id)
            prod = Product.query.get(detail.product_id)
            if prod:
                qty_to_restore = float(detail.quantity)
                if order.type == 'Sale':
                    # Sale Reversal: Restore quantity to batches in LIFO order (latest used first)
                    if prod.is_combo:
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                child_needed = qty_to_restore * ci.quantity
                                child.stock += int(child_needed)
                                
                                # Restore batches for child
                                c_batches = StockBatch.query.filter(StockBatch.product_id == child.id, StockBatch.current_quantity < StockBatch.original_quantity)\
                                                            .order_by(StockBatch.created_at.desc()).all()
                                for cb in c_batches:
                                    if child_needed <= 0: break
                                    can_add = cb.original_quantity - cb.current_quantity
                                    take = min(child_needed, can_add)
                                    cb.current_quantity += take
                                    child_needed -= take
                    else:
                        prod.stock += int(qty_to_restore)
                        # Restore batches for product
                        p_batches = StockBatch.query.filter(StockBatch.product_id == prod.id, StockBatch.current_quantity < StockBatch.original_quantity)\
                                                    .order_by(StockBatch.created_at.desc()).all()
                        for pb in p_batches:
                            if qty_to_restore <= 0: break
                            can_add = pb.original_quantity - pb.current_quantity
                            take = min(qty_to_restore, can_add)
                            pb.current_quantity += take
                            qty_to_restore -= take
                            
                elif order.type == 'Purchase':
                    # Purchase Reversal: Just subtract from stock (batches handled above)
                    qty_to_subtract = float(detail.shipped_quantity) if order.is_consignment else float(detail.quantity)
                    if prod.is_combo:
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                child.stock -= int(qty_to_subtract * ci.quantity)
                    else:
                        prod.stock -= int(qty_to_subtract)
        
        # 2. Reverse Partner Debt Logic (Handled by recalculate at end)
        
        # 3. Cleanup linked settlement vouchers
        linked_vouchers = CashVoucher.query.filter_by(order_id=order.id).all()
        for v in linked_vouchers:
            db.session.delete(v)
        
        # 4. Cleanup linked bank transactions
        linked_bank_ts = BankTransaction.query.filter_by(order_id=order.id).all()
        for bt in linked_bank_ts:
            bank_acc = BankAccount.query.get(bt.account_id)
            if bank_acc:
                if bt.type == 'Deposit':
                    bank_acc.balance -= bt.amount
                else:
                    bank_acc.balance += bt.amount
            db.session.delete(bt)

        db.session.delete(order)
        db.session.commit()
        
        # 5. Recalculate cost prices for products affected by this Purchase deletion
        if order.type == 'Purchase':
            for pid in set(affected_product_ids):
                recalculate_product_cost_price(pid)

        # Final Recalculate
        if partner_id:
            recalculate_partner_debt_internal(partner_id)
            
        return jsonify({'message': 'Order deleted and data reversed successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

def sync_order_amount_paid(order_id):
    """
    Recalculate order.amount_paid based on linked CashVouchers.
    Only applies to Debt orders, as Cash orders do not typically have vouchers.
    """
    order = Order.query.get(order_id)
    if not order: return

    # If it's a Debt order, the amount_paid should be the sum of all settlement vouchers
    if order.payment_method == 'Debt':
        vouchers = CashVoucher.query.filter_by(order_id=order_id).all()
        bank_txs = BankTransaction.query.filter_by(order_id=order_id).all()
        total_paid = 0
        
        # Cash Vouchers
        for v in vouchers:
            # For Sale: Receipt (+) adds to paid amount, Payment (-) refunds it
            if order.type == 'Sale':
                if v.type == 'Receipt': total_paid += v.amount
                elif v.type == 'Payment': total_paid -= v.amount
            # For Purchase: Payment (+) adds to paid amount, Receipt (-) refunds it
            elif order.type == 'Purchase':
                if v.type == 'Payment': total_paid += v.amount
                elif v.type == 'Receipt': total_paid -= v.amount
                
        for t in bank_txs:
            # For Sale: Deposit (+) adds to paid amount, Withdrawal (-) refunds it
            if order.type == 'Sale':
                if t.type == 'Deposit': total_paid += t.amount
                elif t.type == 'Withdrawal': total_paid -= t.amount
            # For Purchase: Withdrawal (+) adds to paid amount, Deposit (-) refunds it
            elif order.type == 'Purchase':
                if t.type == 'Withdrawal': total_paid += t.amount
                elif t.type == 'Deposit': total_paid -= t.amount
        
        order.amount_paid = total_paid
        db.session.commit()

@app.route('/api/orders/<int:id>', methods=['PUT'])
def update_order(id):
    order = Order.query.get_or_404(id)
    data = request.json
    try:
        # Special case for Opening Balance (#NODAU)
        if order.display_id in ['#NODAU', 'NODAU']:
            new_amount = float(data.get('total_amount', 0))
            order.total_amount = abs(new_amount)
            if 'type' in data:
                order.type = data['type']
            else:
                order.type = 'Sale' if new_amount >= 0 else 'Purchase'
            order.note = data.get('note', order.note)
            db.session.commit()
            if order.partner_id:
                recalculate_partner_debt_internal(order.partner_id)
            return jsonify(order.to_dict())

        # 1. Reverse Previous Inventory & FIFO Batches
        affected_product_ids = []
        if order.type == 'Sale':
            for detail in order.details:
                if detail.product_id is None:
                    continue
                prod = Product.query.get(detail.product_id)
                if prod:
                    qty_to_restore = float(detail.quantity)
                    if prod.is_combo:
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                child_needed = qty_to_restore * ci.quantity
                                child.stock += int(child_needed)
                                c_batches = StockBatch.query.filter(StockBatch.product_id == child.id, StockBatch.current_quantity < StockBatch.original_quantity)\
                                                            .order_by(StockBatch.created_at.desc()).all()
                                for cb in c_batches:
                                    if child_needed <= 0: break
                                    can_add = cb.original_quantity - cb.current_quantity
                                    take = min(child_needed, can_add)
                                    cb.current_quantity += take
                                    child_needed -= take
                    else:
                        prod.stock += int(qty_to_restore)
                        p_batches = StockBatch.query.filter(StockBatch.product_id == prod.id, StockBatch.current_quantity < StockBatch.original_quantity)\
                                                    .order_by(StockBatch.created_at.desc()).all()
                        for pb in p_batches:
                             if qty_to_restore <= 0: break
                             can_add = pb.original_quantity - pb.current_quantity
                             take = min(qty_to_restore, can_add)
                             pb.current_quantity += take
                             qty_to_restore -= take
                             
        elif order.type == 'Purchase':
            is_consignment = data.get('is_consignment', order.is_consignment)
            if is_consignment:
                # Consignment: remove all batches and reverse stock
                batches_to_delete = StockBatch.query.filter_by(purchase_order_id=order.id).all()
                for b in batches_to_delete:
                    affected_product_ids.append(b.product_id)
                    db.session.delete(b)
                for detail in order.details:
                    if detail.product_id is None:
                        continue
                    prod = Product.query.get(detail.product_id)
                    if prod:
                        qty_to_reverse = float(detail.shipped_quantity)
                        if prod.is_combo:
                            for ci in prod.combo_items:
                                child = Product.query.get(ci.product_id)
                                if child:
                                    child.stock -= int(qty_to_reverse * ci.quantity)
                        else:
                            prod.stock -= int(qty_to_reverse)
            else:
                # Non-consignment purchase: match existing batches with new details index-by-index to preserve sold quantities
                existing_batches = StockBatch.query.filter_by(purchase_order_id=order.id).order_by(StockBatch.id.asc()).all()
                old_batches_by_prod = {}
                for b in existing_batches:
                    old_batches_by_prod.setdefault(b.product_id, []).append(b)
                
                new_batch_reqs = []
                for item in data['details']:
                    product_id = item.get('product_id')
                    if product_id is None:
                        continue
                    item_qty = float(item['quantity'])
                    item_price = float(item['price'])
                    prod = Product.query.get(product_id)
                    if not prod:
                        continue
                    if prod.is_combo:
                        for ci in prod.combo_items:
                            new_batch_reqs.append({
                                'product_id': ci.product_id,
                                'qty': item_qty * ci.quantity,
                                'price': item_price / ci.quantity if ci.quantity > 0 else 0.0
                            })
                    else:
                        new_batch_reqs.append({
                            'product_id': prod.id,
                            'qty': item_qty,
                            'price': item_price
                        })
                        
                new_reqs_by_prod = {}
                for req in new_batch_reqs:
                    new_reqs_by_prod.setdefault(req['product_id'], []).append(req)
                    
                all_pids = set(old_batches_by_prod.keys()) | set(new_reqs_by_prod.keys())
                for pid in all_pids:
                    p = Product.query.get(pid)
                    if not p:
                        continue
                    old_list = old_batches_by_prod.get(pid, [])
                    new_list = new_reqs_by_prod.get(pid, [])
                    
                    for i in range(max(len(old_list), len(new_list))):
                        if i < len(old_list) and i < len(new_list):
                            b = old_list[i]
                            req = new_list[i]
                            
                            old_orig = b.original_quantity
                            old_curr = b.current_quantity
                            sold_qty = old_orig - old_curr
                            
                            b.cost_price = req['price']
                            b.original_quantity = req['qty']
                            b.current_quantity = max(0.0, req['qty'] - sold_qty)
                            
                            p.stock += int(req['qty'] - old_orig)
                            affected_product_ids.append(pid)
                        elif i < len(new_list):
                            req = new_list[i]
                            curr_qty = adjust_negative_stock_backorder(pid, req['qty'], req['price'])
                            new_batch = StockBatch(
                                product_id=pid,
                                purchase_order_id=order.id,
                                original_quantity=req['qty'],
                                current_quantity=curr_qty,
                                cost_price=req['price'],
                                created_at=order.date or get_vn_time()
                            )
                            db.session.add(new_batch)
                            p.stock += int(req['qty'])
                            affected_product_ids.append(pid)
                        else:
                            b = old_list[i]
                            p.stock -= int(b.original_quantity)
                            db.session.delete(b)
                            affected_product_ids.append(pid)
        
        # Reverse Partner Debt Impact & Clean linked vouchers
        old_debt = 0
        old_partner = None
        if order.partner_id:
            old_partner = Partner.query.get(order.partner_id)
            if old_partner:
                if order.payment_method == 'Debt':
                    if order.type == 'Sale':
                        old_partner.debt_balance -= order.total_amount
                    else:
                        old_partner.debt_balance += order.total_amount

        # Clean all linked vouchers (both settlement and auto)
        linked_vouchers = CashVoucher.query.filter_by(order_id=order.id).all()
        for v in linked_vouchers:
            if old_partner and order.payment_method == 'Debt' and v.source == 'settlement':
                if v.type == 'Receipt':
                    old_partner.debt_balance += v.amount
                else:
                    old_partner.debt_balance -= v.amount
            db.session.delete(v)

        if old_partner:
            old_debt = old_partner.debt_balance
        
        old_bank_ts = BankTransaction.query.filter_by(order_id=order.id).all()
        for bt in old_bank_ts:
            bank_acc = BankAccount.query.get(bt.account_id)
            if bank_acc:
                if bt.type == 'Deposit':
                    bank_acc.balance -= bt.amount
                else:
                    bank_acc.balance += bt.amount
            db.session.delete(bt)
            
        old_is_consignment = order.is_consignment
        old_shipped_map = {d.product_id: d.shipped_quantity for d in order.details if d.product_id is not None}
        order.details.clear()
        db.session.flush()
        
        order.partner_id = data.get('partner_id')
        order.payment_method = data['payment_method']
        order.note = data.get('note')
        order.amount_paid = data.get('amount_paid', 0)
        order.cash_given = data.get('cash_given', 0)
        if 'is_consignment' in data:
            order.is_consignment = data['is_consignment']
        
        if 'shipping_status' in data: order.shipping_status = data['shipping_status']
        if 'shipping_address' in data: order.shipping_address = data['shipping_address']
        if 'shipping_phone' in data: order.shipping_phone = data['shipping_phone']
        
        total = 0
        for item in data['details']:
            product_id = item.get('product_id')
            item_qty = float(item['quantity'])
            item_price = float(item['price'])
 
            if product_id is None:
                detail = OrderDetail(
                    order_id=order.id,
                    product_id=None,
                    product_name_override=item.get('product_name') or item.get('name'),
                    quantity=item_qty,
                    price=item_price
                )
                db.session.add(detail)
                total += item_qty * item_price
                continue
 
            prod = Product.query.get(product_id)
            if not prod:
                raise Exception(f"Product {product_id} not found")
            
            avg_cost = 0
            if data['type'] == 'Sale':
                if item_qty < 0:
                    prod.stock -= int(item_qty)
                    if prod.is_combo:
                        total_combo_cost = 0
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                total_combo_cost += ci.quantity * (child.cost_price or 0)
                        avg_cost = total_combo_cost
                    else:
                        avg_cost = prod.cost_price or 0
                else:
                    if prod.is_combo:
                        total_combo_cost = 0
                        for ci in prod.combo_items:
                            child = Product.query.get(ci.product_id)
                            if child:
                                needed_qty = item_qty * ci.quantity
                                child.stock -= int(needed_qty)
                                c_cost = 0
                                rem = needed_qty
                                batches = StockBatch.query.filter(StockBatch.product_id == child.id, StockBatch.current_quantity > 0)\
                                                          .order_by(StockBatch.created_at.asc()).all()
                                for b in batches:
                                    if rem <= 0: break
                                    take = min(rem, b.current_quantity)
                                    b.current_quantity -= take
                                    c_cost += take * b.cost_price
                                    rem -= take
                                if rem > 0: c_cost += rem * (child.cost_price or 0)
                                total_combo_cost += c_cost
                        avg_cost = total_combo_cost / item_qty if item_qty > 0 else 0
                    else:
                        prod.stock -= int(item_qty)
                        rem = item_qty
                        tot_cost = 0
                        batches = StockBatch.query.filter(StockBatch.product_id == prod.id, StockBatch.current_quantity > 0)\
                                                  .order_by(StockBatch.created_at.asc()).all()
                        for b in batches:
                            if rem <= 0: break
                            take = min(rem, b.current_quantity)
                            b.current_quantity -= take
                            tot_cost += take * b.cost_price
                            rem -= take
                        if rem > 0: tot_cost += rem * (prod.cost_price or 0)
                        avg_cost = tot_cost / item_qty if item_qty > 0 else 0
            
            elif data['type'] == 'Purchase':
                # Batch and stock updates are already handled in Step 2!
                avg_cost = item_price
            
            if (old_is_consignment and order.is_consignment) or order.shipping_status is not None:
                preserved_shipped = old_shipped_map.get(prod.id, 0.0)
                shipped_qty = min(preserved_shipped, item_qty)
            else:
                shipped_qty = item_qty
            
            detail = OrderDetail(
                order_id=order.id,
                product_id=prod.id,
                product_name_override=item.get('product_name') or item.get('name'),
                quantity=item_qty,
                price=item_price,
                cost_price=avg_cost if data['type'] == 'Sale' else item_price,
                shipped_quantity=shipped_qty
            )
            db.session.add(detail)
            total += item_qty * item_price
        
        order.total_amount = total
        
        if order.partner_id:
            partner = Partner.query.get(order.partner_id)
            if partner:
                if not (old_partner and old_partner.id == partner.id and order.old_debt is not None):
                    order.old_debt = partner.debt_balance
                if order.payment_method == 'Debt':
                    if order.type == 'Sale':
                        partner.debt_balance += total
                    else:
                        partner.debt_balance -= total
                    
                    upfront = float(data.get('amount_paid', 0))
                    if upfront > 0:
                        v_type = 'Receipt' if (order.type == 'Sale' and total >= 0) or (order.type == 'Purchase' and total < 0) else 'Payment'
                        v_note = f"Thanh toán cho đơn {order.display_id}"
                        if v_type == 'Receipt': partner.debt_balance -= upfront
                        else: partner.debt_balance += upfront
                        
                        v = CashVoucher(partner_id=partner.id, amount=upfront, note=v_note, type=v_type, source='settlement', order_id=order.id)
                        db.session.add(v)
        
        if data.get('payment_method') == 'Transfer' and data.get('bank_account_id'):
            acc_id = int(data['bank_account_id'])
            bank_acc = BankAccount.query.get(acc_id)
            if bank_acc:
                upfront = float(data.get('amount_paid', 0)) or total
                t_type = 'Deposit' if data['type'] == 'Sale' else 'Withdrawal'
                if data['type'] == 'Sale' and total < 0: t_type = 'Withdrawal'
                elif data['type'] == 'Purchase' and total < 0: t_type = 'Deposit'
                
                bt = BankTransaction(account_id=acc_id, amount=abs(upfront), type=t_type, note=f"Cập nhật đơn {order.display_id}", partner_id=data.get('partner_id'), order_id=order.id)
                if t_type == 'Deposit': bank_acc.balance += abs(upfront)
                else: bank_acc.balance -= abs(upfront)
                db.session.add(bt)
                order.amount_paid = upfront

        # --- Cash Payment Support (Sổ Tiền Mặt) ---
        if data.get('payment_method') == 'Cash':
            v_type = 'Receipt' if data['type'] == 'Sale' else 'Payment'
            v_note = f"Thu tiền bán lẻ - Đơn {order.display_id}" if data['type'] == 'Sale' else f"Chi tiền nhập hàng - Đơn {order.display_id}"
            if data.get('partner_id'):
                partner = Partner.query.get(data['partner_id'])
                if partner:
                    v_note = f"Thu tiền bán hàng - Đơn {order.display_id} ({partner.name})" if data['type'] == 'Sale' else f"Chi tiền nhập hàng - Đơn {order.display_id} ({partner.name})"
            
            v_amount = total
            if total < 0:
                v_amount = abs(total)
                v_type = 'Payment' if data['type'] == 'Sale' else 'Receipt'
                v_note = f"Chi trả tiền hàng trả - Đơn {order.display_id}" if data['type'] == 'Sale' else f"Thu tiền nhập hàng trả - Đơn {order.display_id}"

            v = CashVoucher(
                partner_id=data.get('partner_id'),
                amount=v_amount,
                note=v_note,
                type=v_type,
                source='auto',
                order_id=order.id
            )
            db.session.add(v)
            order.amount_paid = total

        db.session.commit()
        
        # 4. Recalculate cost prices for products affected by this Purchase update
        if order.type == 'Purchase':
            # Also add current products in case they are different from old ones
            for detail in order.details:
                if detail.product_id:
                    affected_product_ids.append(detail.product_id)
            
            for pid in set(affected_product_ids):
                recalculate_product_cost_price(pid)

        # Enforce consistency: amount_paid must match vouchers
        sync_order_amount_paid(order.id)
        
        # Final Recalculate to ensure everything is perfect
        if order.partner_id:
            recalculate_partner_debt_internal(order.partner_id)
            
        order_dict = order.to_dict()
        order_dict['old_debt'] = old_debt
        return jsonify(order_dict)
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# --- Bank Accounts ---
@app.route('/api/bank-accounts', methods=['GET'])
def get_bank_accounts():
    accounts = BankAccount.query.all()
    return jsonify([a.to_dict() for a in accounts])

@app.route('/api/bank-accounts', methods=['POST'])
def create_bank_account():
    data = request.json
    new_acc = BankAccount(
        bank_name=data['bank_name'],
        account_number=data['account_number'],
        account_holder=data.get('account_holder'),
        balance=data.get('balance', 0)
    )
    db.session.add(new_acc)
    db.session.commit()
    return jsonify(new_acc.to_dict()), 201

@app.route('/api/bank-accounts/<int:id>', methods=['PUT'])
def update_bank_account(id):
    acc = BankAccount.query.get_or_404(id)
    data = request.json
    acc.bank_name = data.get('bank_name', acc.bank_name)
    acc.account_number = data.get('account_number', acc.account_number)
    acc.account_holder = data.get('account_holder', acc.account_holder)
    if 'balance' in data:
        acc.balance = float(data['balance'])
    db.session.commit()
    return jsonify(acc.to_dict())

@app.route('/api/bank-accounts/<int:id>', methods=['DELETE'])
def delete_bank_account(id):
    acc = BankAccount.query.get_or_404(id)
    db.session.delete(acc)
    db.session.commit()
    return jsonify({'message': 'Deleted successfully'})

@app.route('/api/bank-transactions', methods=['POST'])
def create_bank_transaction():
    data = request.json
    account_id = data['account_id']
    amount = float(data['amount'])
    t_type = data['type'] # 'Deposit', 'Withdrawal', 'Transfer'
    note = data.get('note', '')
    partner_id = data.get('partner_id')
    order_id = data.get('order_id')
    
    acc = BankAccount.query.get_or_404(account_id)
    partner = Partner.query.get(partner_id) if partner_id else None
    
    transaction = BankTransaction(
        account_id=account_id,
        amount=amount,
        type=t_type,
        note=note,
        partner_id=partner_id,
        order_id=order_id
    )
    
    if t_type == 'Deposit':
        acc.balance += amount
        if partner:
            # Customer pays us via Bank -> Reduce Debt
            partner.debt_balance -= amount
    elif t_type == 'Withdrawal':
        acc.balance -= amount
        if partner:
            # We pay supplier via Bank -> Increase Debt (towards 0 or positive)
            partner.debt_balance += amount
    
    db.session.add(transaction)
    db.session.commit()
    
    # Enforce consistency
    if partner_id:
        recalculate_partner_debt_internal(partner_id)
        
    if order_id:
        sync_order_amount_paid(order_id)
        
    return jsonify(transaction.to_dict()), 201


@app.route('/api/bank-transactions', methods=['GET'])
def get_bank_transactions():
    account_id = request.args.get('account_id', type=int)
    partner_id = request.args.get('partner_id', type=int)
    
    query = BankTransaction.query
    if account_id:
        query = query.filter(BankTransaction.account_id == account_id)
    if partner_id:
        query = query.filter(BankTransaction.partner_id == partner_id)
    
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    quarter = request.args.get('quarter', type=int)

    if year:
        query = query.filter(extract('year', BankTransaction.date) == int(year))
    if month:
        query = query.filter(extract('month', BankTransaction.date) == int(month))
    if day:
        query = query.filter(extract('day', BankTransaction.date) == int(day))
    if quarter:
        if quarter == 1:
            query = query.filter(extract('month', BankTransaction.date).in_([1, 2, 3]))
        elif quarter == 2:
            query = query.filter(extract('month', BankTransaction.date).in_([4, 5, 6]))
        elif quarter == 3:
            query = query.filter(extract('month', BankTransaction.date).in_([7, 8, 9]))
        elif quarter == 4:
            query = query.filter(extract('month', BankTransaction.date).in_([10, 11, 12]))
        
    transactions = query.order_by(BankTransaction.date.desc()).all()
    return jsonify([t.to_dict() for t in transactions])

@app.route('/api/bank-transactions/<int:id>', methods=['DELETE'])
def delete_bank_transaction(id):
    tx = BankTransaction.query.get_or_404(id)
    acc = BankAccount.query.get(tx.account_id)
    partner = Partner.query.get(tx.partner_id) if tx.partner_id else None
    
    # Reverse balance changes
    if tx.type == 'Deposit':
        if acc:
            acc.balance -= tx.amount
        if partner:
            partner.debt_balance += tx.amount
    elif tx.type == 'Withdrawal':
        if acc:
            acc.balance += tx.amount
        if partner:
            partner.debt_balance -= tx.amount
            
    db.session.delete(tx)
    db.session.commit()
    
    if tx.partner_id:
        recalculate_partner_debt_internal(tx.partner_id)
    if tx.order_id:
        sync_order_amount_paid(tx.order_id)
        
    return jsonify({'message': 'Deleted successfully'})

# --- Dashboard ---

@app.route('/api/dashboard-stats', methods=['GET'])
def dashboard_stats():
    from datetime import datetime, timedelta
    from sqlalchemy import func, extract

    # --- 1. Overall Stats ---
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')

    # For revenue and profit, default to today if no filter is provided
    now_dt = get_vn_time()
    filter_year = year
    filter_month = month
    filter_day = day

    if not any([year, month, day]):
        filter_year = str(now_dt.year)
        filter_month = str(now_dt.month).zfill(2)
        filter_day = str(now_dt.day).zfill(2)

    # Determine the datetime ranges for current and previous period
    if filter_year and filter_month and filter_day:
        current_start = datetime(int(filter_year), int(filter_month), int(filter_day))
        current_end = current_start + timedelta(days=1)
        prev_start = current_start - timedelta(days=1)
        prev_end = current_start
    elif filter_year and filter_month:
        current_start = datetime(int(filter_year), int(filter_month), 1)
        if int(filter_month) == 12:
            current_end = datetime(int(filter_year) + 1, 1, 1)
        else:
            current_end = datetime(int(filter_year), int(filter_month) + 1, 1)
        
        if int(filter_month) == 1:
            prev_start = datetime(int(filter_year) - 1, 12, 1)
        else:
            prev_start = datetime(int(filter_year), int(filter_month) - 1, 1)
        prev_end = current_start
    elif filter_year:
        current_start = datetime(int(filter_year), 1, 1)
        current_end = datetime(int(filter_year) + 1, 1, 1)
        prev_start = datetime(int(filter_year) - 1, 1, 1)
        prev_end = current_start
    else:
        # Fallback to today vs yesterday
        current_start = datetime(now_dt.year, now_dt.month, now_dt.day)
        current_end = current_start + timedelta(days=1)
        prev_start = current_start - timedelta(days=1)
        prev_end = current_start

    # Helper function to query stats for range
    def get_stats_for_range(start_dt, end_dt):
        rev = db.session.query(func.sum(Order.total_amount))\
            .filter(Order.type == 'Sale', Order.date >= start_dt, Order.date < end_dt, Order.display_id.notin_(['NODAU', '#NODAU'])).scalar() or 0
        
        cash_rev = db.session.query(func.sum(Order.amount_paid))\
            .filter(Order.type == 'Sale', Order.date >= start_dt, Order.date < end_dt, Order.display_id.notin_(['NODAU', '#NODAU'])).scalar() or 0
        
        cost = db.session.query(
            func.sum(OrderDetail.quantity * func.coalesce(OrderDetail.cost_price, Product.cost_price, OrderDetail.price))
        ).join(Order, Order.id == OrderDetail.order_id)\
         .outerjoin(Product, OrderDetail.product_id == Product.id)\
         .filter(Order.type == 'Sale', Order.date >= start_dt, Order.date < end_dt, Order.display_id.notin_(['NODAU', '#NODAU'])).scalar() or 0
         
        prof = rev - cost
        return rev, cash_rev, prof

    # Get current period stats
    revenue, cash_revenue, profit = get_stats_for_range(current_start, current_end)
    debt_revenue = revenue - cash_revenue

    # Get previous period stats for trend comparison
    prev_rev, prev_cash_rev, prev_profit = get_stats_for_range(prev_start, prev_end)

    # Calculate trends
    if prev_rev > 0:
        revenue_trend = round(((revenue - prev_rev) / prev_rev) * 100)
    else:
        revenue_trend = 100 if revenue > 0 else 0

    if prev_profit > 0:
        profit_trend = round(((profit - prev_profit) / prev_profit) * 100)
    else:
        profit_trend = 100 if profit > 0 else 0
    
    # --- 2. Debt (Unfiltered - All Time) ---
    all_partners = Partner.query.filter(Partner.debt_balance != 0).all()
    # total_customer_debt: Sum of all positive balances (Receivables) for partners marked as customers
    total_customer_debt = sum(p.debt_balance for p in all_partners if p.is_customer and p.debt_balance > 0)
    # total_supplier_debt: Sum of all absolute negative balances (Payables) for partners marked as suppliers
    total_supplier_debt = sum(abs(p.debt_balance) for p in all_partners if p.is_supplier and p.debt_balance < 0)

    # Lists filtered by type AND appropriate balance sign
    customers_with_debt = sorted([p for p in all_partners if p.is_customer and p.debt_balance > 0], key=lambda x: x.debt_balance, reverse=True)[:10]
    suppliers_with_debt = sorted([p for p in all_partners if p.is_supplier and p.debt_balance < 0], key=lambda x: abs(x.debt_balance), reverse=True)[:10]

    # --- 3. 7-Day Revenue Chart ---
    today_dt = get_vn_time()
    seven_days_ago = today_dt - timedelta(days=7)
    
    # Daily Revenue
    daily_revs = db.session.query(
        db.func.date(Order.date).label('day'),
        func.sum(Order.total_amount).label('rev')
    ).filter(Order.type == 'Sale', Order.date >= seven_days_ago, Order.display_id.notin_(['NODAU', '#NODAU']))\
     .group_by(db.func.date(Order.date)).all()
    
    # Daily Cost (Locked)
    daily_costs = db.session.query(
        db.func.date(Order.date).label('day'),
        func.sum(OrderDetail.quantity * func.coalesce(OrderDetail.cost_price, Product.cost_price, OrderDetail.price)).label('cost')
    ).join(Order, Order.id == OrderDetail.order_id)\
     .outerjoin(Product, OrderDetail.product_id == Product.id)\
     .filter(Order.type == 'Sale', Order.date >= seven_days_ago, Order.display_id.notin_(['NODAU', '#NODAU']))\
     .group_by(db.func.date(Order.date)).all()
    
    rev_map = {str(d.day): d.rev for d in daily_revs}
    cost_map = {str(d.day): d.cost for d in daily_costs}
    
    last_7_days = [(today_dt - timedelta(days=i)).date() for i in range(6, -1, -1)] 
    chart_labels = [d.strftime('%d/%m') for d in last_7_days]
    chart_data = []
    chart_profit_data = []

    for d in last_7_days:
        r = rev_map.get(str(d), 0)
        c = cost_map.get(str(d), 0)
        chart_data.append(r)
        chart_profit_data.append(r - c)

    # --- 4. Product Warnings (Expiry & Low Stock) ---
    today = today_dt.date()
    near_expiry_count = 0
    expired_count = 0
    
    # Optimize stock warning: use SQL for counting
    low_stock_count = Product.query.filter(Product.is_active == True, Product.stock > 0, Product.stock <= Product.multiplier).count()
    
    products_with_expiry = Product.query.filter(Product.expiry_date != None).all()
    for p in products_with_expiry:
        # Expiry logic: Near expiry <= 90 days, Expired < 0 days
        try:
            # Handle multiple date formats
            exp_date = None
            date_str = str(p.expiry_date).strip()
            
            # Try parsing commonly used formats
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
                try:
                    exp_date = datetime.strptime(date_str, fmt).date()
                    break
                except ValueError:
                    continue
            
            if exp_date:
                delta = (exp_date - today).days
                
                if delta < 0:
                    expired_count += 1
                elif delta <= 60: # Changed from 90 to 60 as per user request
                    near_expiry_count += 1
        except (ValueError, TypeError):
            pass # Ignore invalid dates
        except (ValueError, TypeError):
            pass # Ignore invalid dates

    return jsonify({
        'revenue': revenue,
        'cash_revenue': cash_revenue,
        'debt_revenue': debt_revenue,
        'profit': profit,
        'revenue_trend': revenue_trend,
        'profit_trend': profit_trend,
        'customer_debt': total_customer_debt,
        'supplier_debt': total_supplier_debt,
        'customer_debt_list': [{'id': p.id, 'name': p.name, 'balance': p.debt_balance} for p in customers_with_debt],
        'supplier_debt_list': [{'id': p.id, 'name': p.name, 'balance': p.debt_balance} for p in suppliers_with_debt],
        'chart': {
            'labels': chart_labels,
            'data': chart_data,
            'profit_data': chart_profit_data
        },
        'expiry': {
            'near': near_expiry_count,
            'expired': expired_count
        },
        'low_stock': low_stock_count
    })


# ... (Keep the first sync_order_amount_paid at line 1885)

@app.route('/api/vouchers', methods=['POST'])
def create_voucher():
    data = request.json
    partner_id = data.get('partner_id')
    amount = float(data.get('amount', 0))
    note = data.get('note', '')
    v_type = data.get('type', 'Payment')
    custom_date_str = data.get('date')
    
    partner = Partner.query.get(partner_id) if partner_id else None
    
    voucher_date = None
    if custom_date_str:
        try:
            # Expecting YYYY-MM-DD
            voucher_date = datetime.strptime(custom_date_str, '%Y-%m-%d')
            # Keep current time if it's today, else set to start of day for historical
            today_str = datetime.now().strftime('%Y-%m-%d')
            if custom_date_str == today_str:
                voucher_date = get_vn_time()
        except ValueError:
            pass

    voucher = CashVoucher(
        partner_id=partner_id,
        amount=amount,
        note=note,
        type=v_type,
        source=data.get('source', 'manual'),
        order_id=data.get('order_id'),
        date=voucher_date if voucher_date else get_vn_time()
    )
    
    if partner and v_type in ['Payment', 'Receipt']:
        # Receipt: Reducing Customer Debt (balance -= amount)
        # Payment: Reducing Supplier Debt (balance += amount towards 0)
        if v_type == 'Receipt':
            partner.debt_balance -= amount
        else:
            partner.debt_balance += amount
        
    db.session.add(voucher)
    db.session.commit()
    
    # Enforce consistency
    if partner_id:
        recalculate_partner_debt_internal(partner_id)
        
    # Sync amount_paid if linked to an order
    if voucher.order_id:
        sync_order_amount_paid(voucher.order_id)
        
    return jsonify(voucher.to_dict()), 201


@app.route('/api/vouchers', methods=['GET'])
def get_vouchers():
    partner_id = request.args.get('partner_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = CashVoucher.query
    
    source = request.args.get('source')
    if source:
        query = query.filter(CashVoucher.source == source)
    
    if partner_id is not None:
        if partner_id == 0:
            query = query.filter(CashVoucher.partner_id == None)
        else:
            query = query.filter(CashVoucher.partner_id == partner_id)
    
    if start_date:
        try:
            query = query.filter(CashVoucher.date >= datetime.fromisoformat(start_date))
        except: pass
    if end_date:
        try:
            query = query.filter(CashVoucher.date <= datetime.fromisoformat(end_date))
        except: pass

    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    quarter = request.args.get('quarter', type=int)

    if year:
        query = query.filter(extract('year', CashVoucher.date) == int(year))
    if month:
        query = query.filter(extract('month', CashVoucher.date) == int(month))
    if day:
        query = query.filter(extract('day', CashVoucher.date) == int(day))
    
    if quarter:
        if quarter == 1:
            query = query.filter(extract('month', CashVoucher.date).in_([1, 2, 3]))
        elif quarter == 2:
            query = query.filter(extract('month', CashVoucher.date).in_([4, 5, 6]))
        elif quarter == 3:
            query = query.filter(extract('month', CashVoucher.date).in_([7, 8, 9]))
        elif quarter == 4:
            query = query.filter(extract('month', CashVoucher.date).in_([10, 11, 12]))
        
    vouchers = query.order_by(CashVoucher.date.desc()).all()
    return jsonify([v.to_dict() for v in vouchers])

@app.route('/api/vouchers/<int:id>', methods=['PUT', 'PATCH'])
def update_voucher(id):
    try:
        voucher = CashVoucher.query.get_or_404(id)
        data = request.json
        new_amount = float(data.get('amount', voucher.amount))
        new_note = data.get('note', voucher.note)
        new_type = data.get('type', voucher.type)
        date_str = data.get('date') # Expecting ISO or YYYY-MM-DD
        
        # 1. Reverse old debt change
        if voucher.partner_id:
            partner = Partner.query.get(voucher.partner_id)
            if partner:
                if voucher.type == 'Receipt':
                    partner.debt_balance += voucher.amount
                elif voucher.type == 'Payment':
                    partner.debt_balance -= voucher.amount
                elif voucher.type == 'DebtIncrease':
                    partner.debt_balance -= voucher.amount
        
        # 2. Update voucher data
        voucher.amount = new_amount
        voucher.note = new_note
        voucher.type = new_type
        if date_str:
            try:
                if 'T' in date_str:
                    from datetime import datetime
                    voucher.date = datetime.fromisoformat(date_str.replace('Z', '+00:00')).replace(tzinfo=None)
                else:
                    from datetime import datetime
                    dt_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    voucher.date = voucher.date.replace(year=dt_obj.year, month=dt_obj.month, day=dt_obj.day)
            except Exception as e:
                print(f"Error parsing date: {e}")
        
        # 3. Apply new debt change
        if voucher.partner_id:
            partner = Partner.query.get(voucher.partner_id)
            if partner:
                if voucher.type == 'Receipt':
                    partner.debt_balance -= voucher.amount
                elif voucher.type == 'Payment':
                    partner.debt_balance += voucher.amount
                elif voucher.type == 'DebtIncrease':
                    partner.debt_balance += voucher.amount
        
        db.session.commit()
        
        # Enforce consistency
        if voucher.partner_id:
            recalculate_partner_debt_internal(voucher.partner_id)
            
        # Sync order amount paid if linked
        if voucher.order_id:
            sync_order_amount_paid(voucher.order_id)
            
        return jsonify(voucher.to_dict())

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/vouchers/<int:id>', methods=['DELETE'])
def delete_voucher(id):
    try:
        voucher = CashVoucher.query.get_or_404(id)
        # Reverse debt change if applicable
        if voucher.partner_id:
            partner = Partner.query.get(voucher.partner_id)
            if partner:
                if voucher.type == 'Receipt':
                    partner.debt_balance += voucher.amount
                elif voucher.type == 'Payment':
                    partner.debt_balance -= voucher.amount
                elif voucher.type == 'DebtIncrease':
                    partner.debt_balance -= voucher.amount
        
        
        # REVERSION LOGIC: If this was a settlement voucher, revert the order to 'Pending'
        if voucher.source == 'settlement' and voucher.order_id:
            order = Order.query.get(voucher.order_id)
            if order:
                order.payment_method = 'Pending'
        partner_id = voucher.partner_id
        order_id = voucher.order_id
        db.session.delete(voucher)
        db.session.commit()
        
        # Enforce consistency
        if partner_id:
            recalculate_partner_debt_internal(partner_id)
            
        if order_id:
            sync_order_amount_paid(order_id)
            
        return jsonify({'message': 'Deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400



# --- Settings ---
@app.route('/api/print-templates', methods=['GET'])
def get_print_templates():
    module = request.args.get('module')
    if module:
        templates = PrintTemplate.query.filter_by(module=module).all()
    else:
        templates = PrintTemplate.query.all()
    return jsonify([t.to_dict() for t in templates])

@app.route('/api/settings', methods=['GET'])
def get_settings():
    # Return general app settings from AppSetting model
    settings = AppSetting.query.all()
    return jsonify({s.setting_key: s.setting_value for s in settings})

@app.route('/api/settings', methods=['POST'])
def save_setting():
    data = request.json
    for key, value in data.items():
        setting = AppSetting.query.filter_by(setting_key=key).first()
        if setting:
            setting.setting_value = str(value)
        else:
            new_setting = AppSetting(setting_key=key, setting_value=str(value))
            db.session.add(new_setting)
    db.session.commit()
    return jsonify({'message': 'Settings saved successfully'})

@app.route('/api/purchase/scan-invoice', methods=['POST'])
def scan_invoice():
    data = request.json or {}
    image_base64 = data.get('image') # Expecting format: "data:image/jpeg;base64,..." or raw base64
    images_base64 = data.get('images', [])
    custom_api_key = data.get('api_key')
    
    base64_list = []
    if image_base64:
        base64_list.append(image_base64)
    if isinstance(images_base64, list):
        base64_list.extend(images_base64)
        
    if not base64_list:
        return jsonify({'error': 'Không nhận được dữ liệu hình ảnh.'}), 400
        
    # Get Gemini API key
    api_key = custom_api_key
    if not api_key:
        # Try from database
        setting = AppSetting.query.filter_by(setting_key='gemini_api_key').first()
        if setting and setting.setting_value:
            api_key = setting.setting_value
            
    if not api_key:
        # Try from environment variable
        api_key = os.environ.get('GEMINI_API_KEY')
        
    if not api_key:
        return jsonify({'error': 'Chưa cấu hình Gemini API Key. Vui lòng cấu hình trong Cài đặt hoặc nhập trực tiếp.'}), 400

    # Build parts list for Gemini
    prompt = (
        "Hãy quét các ảnh hóa đơn, phiếu nhập hàng, toa thuốc, đơn đặt hàng, hoặc giấy viết tay danh sách mua hàng của khách này và trả về danh sách các sản phẩm dưới dạng JSON array. "
        "Mỗi đối tượng sản phẩm trong mảng JSON bắt buộc phải tuân thủ cấu trúc sau:\n"
        "{\n"
        "  \"product_name\": \"Tên sản phẩm (đầy đủ, rõ ràng, giữ lại quy cách/hàm lượng/thể tích nếu có như 240ml, 500EC, 100g...)\",\n"
        "  \"quantity\": số lượng mua hoặc nhập (kiểu số, mặc định là 1 nếu không ghi rõ),\n"
        "  \"price\": đơn giá nếu có ghi trên giấy hoặc hóa đơn (kiểu số, nếu là đơn khách đặt không ghi giá thì để là 0),\n"
        "  \"unit\": \"đơn vị tính nếu có (ví dụ: chai, gói, thùng, bao, can, kg, hộp...)\"\n"
        "}\n\n"
        "Chỉ trả về chuỗi JSON thô hợp lệ, không bọc trong ```json ... ``` hoặc bất kỳ ký tự nào khác. "
        "Nếu không phát hiện được sản phẩm nào, hãy trả về mảng rỗng []."
    )
    
    parts = [{"text": prompt}]
    
    for img_b64 in base64_list:
        # Extract base64 raw data
        if ',' in img_b64:
            header, base64_data = img_b64.split(',', 1)
            mime_type = 'image/jpeg'
            if 'image/png' in header:
                mime_type = 'image/png'
            elif 'image/webp' in header:
                mime_type = 'image/webp'
            elif 'image/gif' in header:
                mime_type = 'image/gif'
        else:
            base64_data = img_b64
            mime_type = 'image/jpeg'
            
        parts.append({
            "inlineData": {
                "mimeType": mime_type,
                "data": base64_data
            }
        })

    import requests
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3.5-flash-lite:generateContent?key={api_key}"
    
    payload = {
        "contents": [
            {
                "parts": parts
            }
        ],
        "generationConfig": {
            "responseMimeType": "application/json"
        }
    }

    try:
        response = requests.post(url, json=payload, headers={'Content-Type': 'application/json'}, timeout=60)
        if response.status_code != 200:
            return jsonify({'error': f"Lỗi từ Gemini API (Mã {response.status_code}): {response.text}"}), response.status_code
            
        res_data = response.json()
        try:
            candidates = res_data.get('candidates', [])
            if not candidates:
                return jsonify([])
            text_content = candidates[0]['content']['parts'][0]['text']
            
            text_content = text_content.strip()
            if text_content.startswith("```json"):
                text_content = text_content[7:]
            if text_content.startswith("```"):
                text_content = text_content[3:]
            if text_content.endswith("```"):
                text_content = text_content[:-3]
            text_content = text_content.strip()
            
            parsed_items = json.loads(text_content)
            return jsonify(parsed_items)
        except Exception as parse_err:
            return jsonify({'error': f"Không thể phân tích dữ liệu JSON trả về từ AI: {str(parse_err)}. Nội dung thô: {text_content}"}), 500
            
    except Exception as e:
        return jsonify({'error': f"Lỗi trong quá trình gửi yêu cầu đến Gemini: {str(e)}"}), 500

@app.route('/api/print-templates', methods=['POST'])
def create_print_template():
    data = request.json
    new_template = PrintTemplate(
        name=data['name'],
        module=data['module'],
        is_default=data.get('is_default', False),
        config=json.dumps(data.get('config', {})),
        content_config=json.dumps(data.get('content_config', {}))
    )
    
    if new_template.is_default:
        # Unset other defaults for this module
        PrintTemplate.query.filter_by(module=new_template.module).update({PrintTemplate.is_default: False})
        
    db.session.add(new_template)
    db.session.commit()
    return jsonify(new_template.to_dict()), 201

@app.route('/api/print-templates/<int:id>', methods=['PUT'])
def update_print_template(id):
    template = PrintTemplate.query.get_or_404(id)
    data = request.json
    template.name = data.get('name', template.name)
    template.module = data.get('module', template.module)
    template.is_default = data.get('is_default', template.is_default)
    if 'config' in data:
        template.config = json.dumps(data['config'])
    if 'content_config' in data:
        template.content_config = json.dumps(data['content_config'])
        
    if template.is_default:
        PrintTemplate.query.filter_by(module=template.module).filter(PrintTemplate.id != id).update({PrintTemplate.is_default: False})
        
    db.session.commit()
    return jsonify(template.to_dict())

@app.route('/api/print-templates/<int:id>', methods=['DELETE'])
def delete_print_template(id):
    template = PrintTemplate.query.get_or_404(id)
    module = template.module
    was_default = template.is_default
    db.session.delete(template)
    db.session.commit()
    
    if was_default:
        remaining = PrintTemplate.query.filter_by(module=module).first()
        if remaining:
            remaining.is_default = True
            db.session.commit()
            
    return jsonify({'message': 'Template deleted successfully'})

# --- Reports ---

@app.route('/api/reports/flattened-products', methods=['GET'])
def report_flattened_products():
    from datetime import datetime
    
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    brands_str = request.args.get('brands')
    categories_str = request.args.get('categories')
    products_str = request.args.get('products')
    group_by_product = request.args.get('group_by_product') == 'true'
    price_mode = request.args.get('price_mode', 'sale') # 'sale' or 'accounting'
    has_code_only = request.args.get('has_code') == 'true'
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))

    try:
        target_profit = float(request.args.get('target_profit', 0))
        profit_variance = float(request.args.get('profit_variance', 0))
    except:
        target_profit = 0
        profit_variance = 0
    
    query = OrderDetail.query.join(Order).options(
        joinedload(OrderDetail.product).joinedload(Product.combo_items).joinedload(ComboItem.product),
        joinedload(OrderDetail.product).joinedload(Product.category)
    ).filter(Order.type == 'Sale')
    
    try:
        if start_date_str:
            query = query.filter(Order.date >= datetime.fromisoformat(start_date_str.replace('Z', '+00:00')))
        if end_date_str:
            query = query.filter(Order.date <= datetime.fromisoformat(end_date_str.replace('Z', '+00:00')))
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
        
    try:
        order_details = query.all()
    except Exception as e:
        app.logger.error(f"Error fetching order details: {str(e)}")
        return jsonify({'error': 'Database error'}), 500
        
    brands = [b.strip() for b in brands_str.split(',') if b.strip()] if brands_str else []
    category_ids = [c.strip() for c in categories_str.split(',') if c.strip()] if categories_str else []
    search_products = [p.strip().lower() for p in products_str.split(',') if p.strip()] if products_str else []
    
    raw_results = []
    
    for d in order_details:
        try:
            base_product = d.product
            if not base_product:
                continue
                
            time_str = d.order.date.strftime('%d/%m/%Y %H:%M')
            order_date_iso = d.order.date.isoformat()
            
            # Helper to append with optional group/flatten logic
            def add_item(p, qty, sale_price):
                # Use accounting price if requested, otherwise fallback to sale_price provided
                display_price = (p.accounting_price or 0) if price_mode == 'accounting' else sale_price
                
                # Generated price for preview
                import random
                # Use a stable seed per product/params to avoid flickering on re-paginations
                seed_val = f"{p.id}_{target_profit}_{profit_variance}"
                rng = random.Random(seed_val)
                current_variance = rng.uniform(-profit_variance, profit_variance)
                cost = p.accounting_price or 0
                gen_price = round(cost * (1 + (target_profit + current_variance) / 100))

                raw_results.append({
                    'order_id': d.order.display_id,
                    'time': time_str,
                    'date_iso': order_date_iso,
                    'code': p.code or '',
                    'product_name': p.name,
                    'brand': p.brand or '',
                    'category_name': p.category.name if p.category else 'Chưa phân loại',
                    'quantity': qty,
                    'retail_price': display_price,
                    'generated_price': gen_price,
                    'total': qty * (gen_price if price_mode == 'accounting' else display_price),
                    'unit': p.unit or '',
                    'accounting_price': p.accounting_price or 0
                })
            
            # Flatten combos
            if base_product.is_combo and base_product.combo_items:
                for item in base_product.combo_items:
                    child_prod = item.product
                    if not child_prod: continue
                    
                    if brands:
                        prod_brand = child_prod.brand.strip() if child_prod.brand and child_prod.brand.strip() else '_no_brand_'
                        if prod_brand not in brands:
                            continue
                    
                    if category_ids:
                        cat_id = str(child_prod.category_id) if child_prod.category_id else '_no_category_'
                        if cat_id not in category_ids:
                            continue

                    if search_products:
                        pc = (child_prod.code or '').strip().lower()
                        pn = (child_prod.name or '').strip().lower()
                        # Strict match for code, partial match for name
                        if not any(sp == pc or sp in pn for sp in search_products):
                            continue

                    if has_code_only and not child_prod.code:
                        continue
                        
                    add_item(child_prod, d.quantity * item.quantity, child_prod.sale_price)
            else:
                if has_code_only and not base_product.code:
                    continue

                if brands:
                    prod_brand = base_product.brand.strip() if base_product.brand and base_product.brand.strip() else '_no_brand_'
                    if prod_brand not in brands:
                        continue
                        
                if category_ids:
                    cat_id = str(base_product.category_id) if base_product.category_id else '_no_category_'
                    if cat_id not in category_ids:
                        continue

                if search_products:
                    p_code = (base_product.code or '').strip().lower()
                    p_name = (base_product.name or '').strip().lower()
                    if not any(sp == p_code or sp in p_name for sp in search_products):
                        continue
                    
                add_item(base_product, d.quantity, d.price)
                    

        except Exception as e:
            app.logger.error(f"Error processing order detail {d.id}: {str(e)}")
            continue
            
    if group_by_product:
        # Aggregate by code + name + unit to be safe
        agg = {}
        for r in raw_results:
            key = (r['code'], r['product_name'], r['unit'])
            if key not in agg:
                agg[key] = {
                    'code': r['code'],
                    'product_name': r['product_name'],
                    'brand': r['brand'],
                    'category_name': r['category_name'],
                    'unit': r['unit'],
                    'accounting_price': r['accounting_price'],
                    'retail_price': r['retail_price'],
                    'generated_price': r['generated_price'],
                    'quantity': 0,
                    'total': 0
                }
            agg[key]['quantity'] += r['quantity']
            agg[key]['total'] += r['total']
        
        results = list(agg.values())
        # Sort by total value desc for summary report
        results.sort(key=lambda x: x['total'], reverse=True)
    else:
        results = raw_results
        # Sort results by time (desc) before slicing
        results.sort(key=lambda x: x['date_iso'] if 'date_iso' in x else '', reverse=True)
            
    total_count = len(results)
    
    # Calculate overall totals for the current filters
    overall_qty = sum(item['quantity'] for item in results)
    overall_total = sum(item['total'] for item in results)
    
    start_idx = (page - 1) * limit
    end_idx = start_idx + limit
    paginated_results = results[start_idx:end_idx]
    
    return jsonify({
        'items': paginated_results,
        'total': total_count,
        'page': page,
        'limit': limit,
        'pages': (total_count + limit - 1) // limit,
        'overall_totals': {
            'quantity': overall_qty,
            'total': overall_total
        }
    })

@app.route('/api/reports/flattened-products/export', methods=['GET'])
def export_flattened_products():
    from datetime import datetime
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font
    
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    brands_str = request.args.get('brands')
    categories_str = request.args.get('categories')
    products_str = request.args.get('products')
    group_by_product = request.args.get('group_by_product') == 'true'
    price_mode = request.args.get('price_mode', 'sale') # 'sale' or 'accounting'
    has_code_only = request.args.get('has_code') == 'true'
    
    try:
        target_profit = float(request.args.get('target_profit', 0))
        profit_variance = float(request.args.get('profit_variance', 0))
    except:
        target_profit = 0
        profit_variance = 0
    
    query = OrderDetail.query.join(Order).options(
        joinedload(OrderDetail.product).joinedload(Product.combo_items).joinedload(ComboItem.product),
        joinedload(OrderDetail.product).joinedload(Product.category)
    ).filter(Order.type == 'Sale')
    
    try:
        if start_date_str:
            query = query.filter(Order.date >= datetime.fromisoformat(start_date_str.replace('Z', '+00:00')))
        if end_date_str:
            query = query.filter(Order.date <= datetime.fromisoformat(end_date_str.replace('Z', '+00:00')))
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
        
    try:
        order_details = query.order_by(Order.date.desc()).all()
    except Exception as e:
        app.logger.error(f"Error fetching order details for export: {str(e)}")
        return jsonify({'error': 'Database error'}), 500
        
    brands = [b.strip() for b in brands_str.split(',') if b.strip()] if brands_str else []
    category_ids = [c.strip() for c in categories_str.split(',') if c.strip()] if categories_str else []
    search_products = [p.strip().lower() for p in products_str.split(',') if p.strip()] if products_str else []
    
    raw_results = []
    
    for d in order_details:
        base_product = d.product
        if not base_product:
            continue
            
        time_str = d.order.date.strftime('%d/%m/%Y %H:%M')
        
        def add_export_item(p, qty, sale_price):
            cat_name = p.category.name if p.category else 'Chưa phân loại'
            display_price = (p.accounting_price or 0) if price_mode == 'accounting' else sale_price
            
            # Generated price logic
            import random
            seed_val = f"{p.id}_{target_profit}_{profit_variance}"
            rng = random.Random(seed_val)
            current_variance = rng.uniform(-profit_variance, profit_variance)
            cost = p.accounting_price or 0
            gen_price = round(cost * (1 + (target_profit + current_variance) / 100))

            raw_results.append({
                'time': time_str,
                'order_id': f"#{d.order.display_id}",
                'code': p.code or '',
                'name': p.name,
                'brand': p.brand or '',
                'category': cat_name,
                'unit': p.unit or '',
                'accounting_price': p.accounting_price or 0,
                'generated_price': gen_price,
                'quantity': qty,
                'price': display_price,
                'total': qty * display_price
            })

        # Flatten combos
        if base_product.is_combo and base_product.combo_items:
            for item in base_product.combo_items:
                child_prod = item.product
                if not child_prod: continue
                
                if brands:
                    prod_brand = child_prod.brand.strip() if child_prod.brand and child_prod.brand.strip() else '_no_brand_'
                    if prod_brand not in brands:
                        continue
                
                if category_ids:
                    cat_id = str(child_prod.category_id) if child_prod.category_id else '_no_category_'
                    if cat_id not in category_ids:
                        continue

                if search_products:
                    p_code = (child_prod.code or '').strip().lower()
                    p_name = (child_prod.name or '').strip().lower()
                    if not any(sp in p_code or sp in p_name for sp in search_products):
                            continue
                
                if has_code_only and not child_prod.code:
                    continue
                    
                add_export_item(child_prod, d.quantity * item.quantity, child_prod.sale_price)
        else:
            # Regular product
            if has_code_only and not base_product.code:
                continue

            if brands:
                prod_brand = base_product.brand.strip() if base_product.brand and base_product.brand.strip() else '_no_brand_'
                if prod_brand not in brands:
                    continue
                    
            if category_ids:
                cat_id = str(base_product.category_id) if base_product.category_id else '_no_category_'
                if cat_id not in category_ids:
                    continue

            if search_products:
                p_code = (base_product.code or '').strip().lower()
                p_name = (base_product.name or '').strip().lower()
                if not any(sp in p_code or sp in p_name for sp in search_products):
                    continue
                
            add_export_item(base_product, d.quantity, d.price)

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo tổng hợp" if group_by_product else "Báo cáo chi tiết"
    
    if group_by_product:
        agg = {}
        for r in raw_results:
            key = (r['code'], r['name'], r['unit'])
            if key not in agg:
                agg[key] = {
                    'code': r['code'],
                    'name': r['name'],
                    'brand': r['brand'],
                    'category': r['category'],
                    'unit': r['unit'],
                    'accounting_price': r['accounting_price'],
                    'generated_price': r['generated_price'],
                    'price': r['price'],
                    'quantity': 0,
                    'total': 0
                }
            agg[key]['quantity'] += r['quantity']
            agg[key]['total'] += r['total']
            
        final_results = sorted(agg.values(), key=lambda x: x['total'], reverse=True)
        headers = ["Mã hàng", "Tên sản phẩm", "Hãng", "Phân loại", "ĐVT", "Đơn giá", "Tổng SL", "Tổng thành tiền"]
        if price_mode == 'accounting':
            headers[5] = "Giá KT"
            headers.insert(6, "Giá bán (Tính)")
        
        ws.append(headers)
        for r in final_results:
            row_data = [r['code'], r['name'], r['brand'], r['category'], r['unit'], r['accounting_price'] if price_mode == 'accounting' else r['price'], r['quantity'], r['total']]
            if price_mode == 'accounting':
                row_data.insert(6, r['generated_price'])
            ws.append(row_data)
    else:
        # Detail view
        headers = ["Thời gian", "Mã đơn", "Mã hàng", "Tên sản phẩm", "Hãng", "Phân loại", "ĐVT", "Số lượng", "Đơn giá", "Thành tiền"]
        if price_mode == 'accounting':
            headers[8] = "Giá KT"
            headers.insert(9, "Giá bán (Tính)")
            
        ws.append(headers)
        for r in raw_results:
            row_data = [r['time'], r['order_id'], r['code'], r['name'], r['brand'], r['category'], r['unit'], r['quantity'], r['price'], r['total']]
            if price_mode == 'accounting':
                row_data.insert(9, r['generated_price'])
            ws.append(row_data)

    # Styling
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Bao_cao_tong_hop_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx" if group_by_product else f"Bao_cao_chi_tiet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def scan_excel_headers(filepath):
    from openpyxl import load_workbook
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Mapping rules for auto-selection (optional but helpful)
        header_keywords = {
            'code': ['mã hàng', 'mã sp', 'mã sản phẩm', 'code', 'item code'],
            'name': ['tên hàng', 'tên sản phẩm', 'sản phẩm', 'product', 'item name', 'tên'],
            'unit': ['đvt', 'đơn vị', 'unit'],
            'quantity': ['số lượng', 'sl', 'qty', 'quantity'],
            'price': ['đơn giá', 'giá', 'price', 'rate', 'đơn giá kt'],
            'total': ['thành tiền', 'tổng tiền', 'amount', 'total'],
            'brand': ['hãng', 'thương hiệu', 'brand'],
            'category': ['phân loại', 'loại', 'category', 'nhóm hàng'],
            'display_id': ['mã đơn', 'order id', 'số hóa đơn'],
            'time': ['thời gian', 'ngày', 'date', 'time']
        }
        
        start_row = 1
        header_row_idx = 1
        
        # Step 1: Find Header Row (must have some content)
        for r_idx in range(1, 51):
            row_vals = [str(ws.cell(row=r_idx, column=c).value or '').strip() for c in range(1, min(ws.max_column + 1, 40))]
            # Criteria for header: at least 3 non-empty cells
            if len([v for v in row_vals if v]) >= 3:
                header_row_idx = r_idx
                start_row = r_idx + 1
                break
        
        # Step 2: Read ALL columns in that header row and their sample values from start_row
        detected_mappings = []
        for c_idx in range(1, min(ws.max_column + 1, 40)):
            header_cell = ws.cell(row=header_row_idx, column=c_idx)
            header_name = str(header_cell.value or '').strip()
            if not header_name: continue
            
            sample_cell = ws.cell(row=start_row, column=c_idx)
            sample_val = str(sample_cell.value or '').strip()
            col_letter = header_cell.column_letter
            
            # Default to static mapping with sample value
            mapping = {
                'column_letter': col_letter,
                'source_type': 'static',
                'source_value': sample_val,
                'header_name': header_name # Useful for UI display
            }
            
            # Auto-suggest field if header matches keywords
            clean_header = header_name.lower()
            for field, keywords in header_keywords.items():
                if any(kw == clean_header or (len(clean_header) > 3 and kw in clean_header) for kw in keywords):
                    mapping['source_type'] = 'field'
                    mapping['source_value'] = field
                    break
            
            detected_mappings.append(mapping)
            
        return start_row, detected_mappings
    except Exception as e:
        print(f"Error scanning layout: {e}")
        return 1, []

@app.route('/api/accounting/templates/upload', methods=['POST'])
def upload_accounting_template():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and file.filename.endswith('.xlsx'):
        filename = f"template_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        filepath = os.path.join('storage', 'templates', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        file.save(filepath)
        
        # Auto-detect headers
        detected_start_row, detected_mappings = scan_excel_headers(filepath)
        
        # Deactivate current templates
        AccountingTemplate.query.update({AccountingTemplate.is_active: False})
        
        new_template = AccountingTemplate(
            name=file.filename,
            file_path=filepath,
            start_row=detected_start_row
        )
        db.session.add(new_template)
        db.session.flush() # Get ID
        
        # Add auto-detected mappings
        for dm in detected_mappings:
            m = AccountingMapping(
                template_id=new_template.id,
                column_letter=dm['column_letter'],
                header_name=dm.get('header_name'),
                source_type=dm['source_type'],
                source_value=dm['source_value']
            )
            db.session.add(m)
            
        db.session.commit()
        return jsonify(new_template.to_dict())
    return jsonify({'error': 'Invalid file type'}), 400

@app.route('/api/accounting/templates', methods=['GET'])
def get_accounting_templates():
    templates = AccountingTemplate.query.order_by(AccountingTemplate.created_at.desc()).all()
    return jsonify([t.to_dict() for t in templates])

@app.route('/api/accounting/config', methods=['GET', 'POST'])
def handle_accounting_config():
    template = AccountingTemplate.query.filter_by(is_active=True).first()
    if not template and request.method == 'GET':
        return jsonify({'error': 'No active template found'}), 404
    
    if request.method == 'POST':
        data = request.json
        if not template:
             return jsonify({'error': 'Please upload a template first'}), 400
        
        if 'start_row' in data:
            template.start_row = int(data['start_row'])
            
        if 'mappings' in data:
            AccountingMapping.query.filter_by(template_id=template.id).delete()
            for col, config in data['mappings'].items():
                m = AccountingMapping(
                    template_id=template.id,
                    column_letter=col.upper(),
                    source_type=config['source_type'],
                    source_value=config['source_value'],
                    header_name=config.get('header_name')
                )
                db.session.add(m)
        db.session.commit()
        return jsonify(template.to_dict())
    
    return jsonify(template.to_dict())

@app.route('/api/accounting/source-fields', methods=['GET'])
def get_accounting_source_fields():
    fields = [
        {'id': 'index', 'name': 'Số thứ tự', 'type': 'index'},
        {'id': 'code', 'name': 'Mã hàng', 'type': 'field'},
        {'id': 'name', 'name': 'Tên sản phẩm', 'type': 'field'},
        {'id': 'brand', 'name': 'Hãng', 'type': 'field'},
        {'id': 'category', 'name': 'Phân loại', 'type': 'field'},
        {'id': 'unit', 'name': 'ĐVT', 'type': 'field'},
        {'id': 'quantity', 'name': 'Số lượng', 'type': 'field'},
        {'id': 'price', 'name': 'Đơn giá', 'type': 'field'},
        {'id': 'total', 'name': 'Thành tiền', 'type': 'field'},
        {'id': 'display_id', 'name': 'Thống kê: Mã đơn (Hệ thống)', 'type': 'field'},
        {'id': 'time', 'name': 'Thống kê: Thời gian (Hệ thống)', 'type': 'field'},
        {'id': 'manual_id', 'name': 'Nhập tay: Số hóa đơn', 'type': 'field'},
        {'id': 'manual_date', 'name': 'Nhập tay: Ngày hóa đơn', 'type': 'field'},
        {'id': 'generated_price', 'name': 'Giá bán (Tự động tính theo LN mục tiêu)', 'type': 'field'},
        {'id': 'static', 'name': 'Dữ liệu mẫu / Cố định', 'type': 'static'},
        {'id': 'static_first', 'name': 'Cố định (Chỉ dòng đầu)', 'type': 'static_first'},
        {'id': 'skip', 'name': 'Bỏ qua (Không ghi dữ liệu)', 'type': 'skip'}
    ]
    return jsonify(fields)

@app.route('/api/accounting/export', methods=['GET'])
def export_accounting_template():
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
    import io

    # 1. Get Active Template & Config
    template = AccountingTemplate.query.filter_by(is_active=True).first()
    if not template or not template.file_path or not os.path.exists(template.file_path):
        return jsonify({'error': 'Vui lòng upload phôi Excel trước khi xuất'}), 404
    
    # 2. Fetch Data
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    brands_str = request.args.get('brands')
    categories_str = request.args.get('categories')
    products_str = request.args.get('products')
    group_by_product = request.args.get('group_by_product') == 'true'
    price_mode = request.args.get('price_mode', 'sale')
    has_code_only = request.args.get('has_code') == 'true'
    manual_invoice_id = request.args.get('manual_id', '')
    manual_invoice_date = request.args.get('manual_date', '')
    
    # Smart Price Generation Params
    try:
        target_profit = float(request.args.get('target_profit', 0))
        profit_variance = float(request.args.get('profit_variance', 0))
    except:
        target_profit = 0
        profit_variance = 0
    
    query = OrderDetail.query.join(Order).options(
        joinedload(OrderDetail.product).joinedload(Product.combo_items).joinedload(ComboItem.product),
        joinedload(OrderDetail.product).joinedload(Product.category)
    ).filter(Order.type == 'Sale')
    
    try:
        if start_date_str:
            query = query.filter(Order.date >= datetime.fromisoformat(start_date_str.replace('Z', '+00:00')))
        if end_date_str:
            query = query.filter(Order.date <= datetime.fromisoformat(end_date_str.replace('Z', '+00:00')))
    except:
        pass
        
    order_details = query.order_by(Order.date.desc()).all()
    brands = [b.strip() for b in brands_str.split(',') if b.strip()] if brands_str else []
    category_ids = [c.strip() for c in categories_str.split(',') if c.strip()] if categories_str else []
    search_products = [p.strip().lower() for p in products_str.split(',') if p.strip()] if products_str else []
    
    raw_results = []
    
    # Pre-import random to use for stable pricing
    import random
    
    for d in order_details:
        base_product = d.product
        if not base_product: continue
        
        def add_item_to_list(p, qty, sale_price):
            cat_name = p.category.name if p.category else 'Chưa phân loại'
            display_price = (p.accounting_price or 0) if price_mode == 'accounting' else sale_price
            
            # Generate smart price if requested (Stable Random)
            seed_val = f"{p.id}_{target_profit}_{profit_variance}"
            rng = random.Random(seed_val)
            current_variance = rng.uniform(-profit_variance, profit_variance)
            cost = p.accounting_price or 0
            gen_price = round(cost * (1 + (target_profit + current_variance) / 100))

            raw_results.append({
                'time': d.order.date.strftime('%d/%m/%Y %H:%M'),
                'order_id': f"#{d.order.display_id}",
                'display_id': d.order.display_id,
                'code': p.code or '',
                'name': p.name,
                'brand': p.brand or '',
                'category': cat_name,
                'unit': p.unit or '',
                'accounting_price': p.accounting_price or 0,
                'generated_price': gen_price,
                'quantity': qty,
                'price': display_price,
                'total': qty * (gen_price if price_mode == 'accounting' else display_price)
            })

        if base_product.is_combo and base_product.combo_items:
            for item in base_product.combo_items:
                child = item.product
                if not child: continue
                # Apply filters to combo components
                child_brand = (child.brand or '_no_brand_').strip()
                if brands and child_brand not in brands: continue
                if category_ids and str(child.category_id or '_no_category_') not in category_ids: continue
                if search_products:
                    child_code = (child.code or '').strip().lower()
                    child_name = (child.name or '').strip().lower()
                    if not any(sp == child_code or sp in child_name for sp in search_products): continue

                if has_code_only and not child.code: continue
                add_item_to_list(child, d.quantity * item.quantity, child.sale_price)
        else:
            # Apply filters to regular product
            if has_code_only and not base_product.code: continue
            p_brand = (base_product.brand or '_no_brand_').strip()
            if brands and p_brand not in brands: continue
            if category_ids and str(base_product.category_id or '_no_category_') not in category_ids: continue
            if search_products:
                p_code = (base_product.code or '').strip().lower()
                p_name = (base_product.name or '').strip().lower()
                if not any(sp == p_code or sp in p_name for sp in search_products): continue
            add_item_to_list(base_product, d.quantity, d.price)

    results = []
    if group_by_product:
        agg = {}
        for r in raw_results:
            key = (r['code'], r['name'], r['unit'])
            if key not in agg:
                agg[key] = {**r, 'quantity': 0, 'total': 0}
            agg[key]['quantity'] += r['quantity']
            agg[key]['total'] += r['total']
        results = sorted(agg.values(), key=lambda x: x['total'], reverse=True)
    else:
        results = raw_results

    # 3. Populate Template
    try:
        from openpyxl.utils import column_index_from_string
        wb = load_workbook(template.file_path)
        ws = wb.active
        
        # Format Vietnamese date if provided
        formatted_manual_date = manual_invoice_date
        if manual_invoice_date:
            try:
                # manual_invoice_date is usually yyyy-mm-dd from frontend
                dt_obj = datetime.strptime(manual_invoice_date, '%Y-%m-%d')
                formatted_manual_date = dt_obj.strftime('%d/%m/%Y')
            except:
                pass

        # CLEAR EXISTING DATA starting from start_row down to a safe margin (e.g. 500 rows)
        # to remove any sample/dummy data left in the template
        for r in range(template.start_row, ws.max_row + 100):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).value = None

        current_row = template.start_row
        # SORT MAPPINGS BY EXCEL COLUMN ORDER
        mappings = sorted(template.mappings, key=lambda m: column_index_from_string(m.column_letter))
        
        for i, row_data in enumerate(results):
            for m in mappings:
                try:
                    if m.source_type == 'skip': continue 
                    
                    col_idx = column_index_from_string(m.column_letter)
                    val = ""
                    if m.source_type == 'index':
                        val = i + 1
                    elif m.source_type == 'field':
                        if m.source_value == 'manual_id':
                            val = manual_invoice_id
                        elif m.source_value == 'manual_date':
                            val = formatted_manual_date
                        elif m.source_value == 'generated_price':
                            val = row_data.get('generated_price', 0)
                        else:
                            val = row_data.get(m.source_value, "")
                    elif m.source_type == 'static':
                        val = m.source_value
                    elif m.source_type == 'static_first':
                        val = m.source_value if i == 0 else ""
                    
                    ws.cell(row=current_row, column=col_idx, value=val)
                except:
                    continue
            current_row += 1
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Bao_cao_ke_toan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        app.logger.error(f"Excel Template Export Error: {str(e)}")
        return jsonify({'error': f'Lỗi khi ghi file Excel: {str(e)}'}), 500

@app.route('/api/reports/unsold', methods=['GET'])
def report_unsold():
    search = request.args.get('search', '').lower()
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 20, type=int)
    
    from datetime import datetime, timedelta
    today = datetime.now()
    
    from sqlalchemy import func, asc
    subq = db.session.query(
        OrderDetail.product_id, 
        func.max(Order.date).label('last_sold_date')
    ).join(Order).filter(Order.type == 'Sale').group_by(OrderDetail.product_id).subquery()
    
    query = db.session.query(
        Product,
        subq.c.last_sold_date
    ).outerjoin(subq, Product.id == subq.c.product_id).filter(
        Product.is_active == True,
        Product.stock > 0
    )
    
    if search:
        query = query.filter(db.or_(
            Product.name.ilike(f'%{search}%'),
            Product.code.ilike(f'%{search}%')
        ))
    
    query = query.order_by(asc(subq.c.last_sold_date).nullsfirst(), Product.id.asc())
    
    total = query.count()
    if limit and limit > 0:
        items = query.offset((page - 1) * limit).limit(limit).all()
    else:
        items = query.all()
        
    unsold_list = []
    for prod, last_date in items:
        if last_date:
            # Handle string dates vs datetime dates
            if isinstance(last_date, str):
                try:
                    last_date_obj = datetime.strptime(last_date.split(' ')[0], '%Y-%m-%d').date()
                except ValueError:
                    # fallback if format is different
                    last_date_obj = today.date()
            else:
                last_date_obj = last_date
            
            days = (today.date() - last_date_obj).days if hasattr(last_date_obj, 'days') or type(last_date_obj) == type(today.date()) else (today - last_date_obj).days
            if isinstance(days, timedelta):
                days = days.days
        else:
            days = 999999
            
        unsold_list.append({
            'id': prod.id,
            'code': prod.code,
            'name': prod.name,
            'stock': prod.stock,
            'unit': prod.unit,
            'multiplier': prod.multiplier,
            'cost_price': prod.cost_price,
            'last_sold_date': last_date if isinstance(last_date, str) else (last_date.strftime('%Y-%m-%d') if last_date else None),
            'days_unsold': days,
            'total_value': (prod.stock or 0) * (prod.cost_price or 0)
        })
        
    return jsonify({
        'items': unsold_list,
        'total': total,
        'pages': (total + limit - 1) // limit if limit and limit > 0 else 1
    })


@app.route('/api/reports/products', methods=['GET'])
def report_products():
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    quarter = request.args.get('quarter')
    search = request.args.get('search', '').lower()
    brand = request.args.get('brand')
    sort_by = request.args.get('sort_by', 'revenue')
    sort_order = request.args.get('sort_order', 'desc')
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)
    
    query = OrderDetail.query.join(Order).outerjoin(Product).options(joinedload(OrderDetail.product)).filter(Order.type == 'Sale')
    # Date filtering
    if year:
        query = query.filter(db.func.strftime('%Y', Order.date) == str(year))
    if month:
        query = query.filter(db.func.strftime('%m', Order.date) == str(month).zfill(2))
    if quarter:
        quarters = {'1': ('01','03'), '2': ('04','06'), '3': ('07','09'), '4': ('10','12')}
        start, end = quarters.get(quarter, ('01','12'))
        query = query.filter(db.func.strftime('%m', Order.date).between(start, end))
    if day:
        query = query.filter(db.func.strftime('%d', Order.date) == str(day).zfill(2))
    
    if brand:
        query = query.filter(Product.brand == brand)
        
    details = query.all()
    report = {}
    for d in details:
        pid = d.product_id or f"custom_{d.product_name_override}"
        if pid not in report:
            p_name = d.product_name_override or (d.product.name if d.product else 'Sản phẩm đã xóa')
            p_unit = d.product.unit if d.product else 'ĐV'
            report[pid] = {
                'id': pid,
                'name': p_name,
                'unit': p_unit,
                'quantity': 0,
                'revenue': 0,
                'cost': 0,
                'profit': 0
            }
        report[pid]['quantity'] += d.quantity
        report[pid]['revenue'] += d.quantity * d.price
        
        # Calculate cost using locked price if available
        # Fallback values for old data
        fallback_cost = 0
        if d.product:
            if d.product.is_combo:
                fallback_cost = sum(ci.quantity * ci.product.cost_price for ci in d.product.combo_items if ci.product)
            else:
                fallback_cost = d.product.cost_price or 0
        else:
            fallback_cost = d.price
            
        unit_cost = d.cost_price if d.cost_price is not None else fallback_cost
        report[pid]['cost'] += d.quantity * unit_cost
        report[pid]['profit'] = report[pid]['revenue'] - report[pid]['cost']
    
    report_list = list(report.values())
    
    # Search filtering
    if search:
        s_norm = remove_accents(search)
        report_list = [item for item in report_list if s_norm in remove_accents(item['name'])]
    
    # Sorting
    reverse = (sort_order == 'desc')
    def get_sort_key(x):
        val = x.get(sort_by, 0)
        if val is None: return 0
        if isinstance(val, str): return val.lower()
        return val
        
    report_list.sort(key=get_sort_key, reverse=reverse)
    
    total = len(report_list)
    if page and limit:
        start = (page - 1) * limit
        end = start + limit
        items = report_list[start:end]
        return jsonify({
            'items': items,
            'total': total,
            'pages': (total + limit - 1) // limit,
            'current_page': page
        })
    
    return jsonify(report_list)

@app.route('/api/reports/partners', methods=['GET'])
def report_partners():
    p_type = request.args.get('type', 'Customer')
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    quarter = request.args.get('quarter')
    search = request.args.get('search', '').lower()
    sort_by = request.args.get('sort_by', 'total_amount')
    sort_order = request.args.get('sort_order', 'desc')
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)
    
    o_type = 'Sale' if p_type == 'Customer' else 'Purchase'
    query = Order.query.filter(Order.type == o_type, Order.display_id.notin_(['NODAU', '#NODAU']))
    # Date filtering
    if year:
        query = query.filter(db.func.strftime('%Y', Order.date) == str(year))
    if month:
        query = query.filter(db.func.strftime('%m', Order.date) == str(month).zfill(2))
    if quarter:
        quarters = {'1': ('01','03'), '2': ('04','06'), '3': ('07','09'), '4': ('10','12')}
        start, end = quarters.get(quarter, ('01','12'))
        query = query.filter(db.func.strftime('%m', Order.date).between(start, end))
    if day:
        query = query.filter(db.func.strftime('%d', Order.date) == str(day).zfill(2))
        
    orders = query.all()
    report = {}
    for o in orders:
        pid = o.partner_id or 0 # 0 for retail
        pname = o.partner.name if o.partner else ('KHÁCH LẺ' if p_type == 'Customer' else 'NCC VÃNG LAI')
        
        if pid not in report:
            report[pid] = {
                'name': pname,
                'id': pid,
                'count': 0,
                'total_amount': 0,
                'profit': 0 # only relevant for sales
            }
        
        report[pid]['count'] += 1
        report[pid]['total_amount'] += o.total_amount
        
        if o_type == 'Sale':
            order_profit = 0
            for d in o.details:
                # Use locked cost_price
                fallback_cost = 0
                if d.product:
                    if d.product.is_combo:
                        fallback_cost = sum(ci.quantity * ci.product.cost_price for ci in d.product.combo_items if ci.product)
                    else:
                        fallback_cost = d.product.cost_price or 0
                else:
                    fallback_cost = d.price
                
                unit_cost = d.cost_price if d.cost_price is not None else fallback_cost
                order_profit += d.quantity * (d.price - unit_cost)
            report[pid]['profit'] += order_profit
            
    report_list = list(report.values())
    
    # Search filtering
    if search:
        s_norm = remove_accents(search)
        report_list = [item for item in report_list if s_norm in remove_accents(item['name'])]
    
    # Sorting
    reverse = (sort_order == 'desc')
    def get_sort_key(x):
        val = x.get(sort_by, 0)
        if val is None: return 0
        if isinstance(val, str): return val.lower()
        return val
        
    report_list.sort(key=get_sort_key, reverse=reverse)
    
    total = len(report_list)
    if page and limit:
        start = (page - 1) * limit
        end = start + limit
        items = report_list[start:end]
        return jsonify({
            'items': items,
            'total': total,
            'pages': (total + limit - 1) // limit,
            'current_page': page
        })
        
    return jsonify(report_list)


@app.route('/api/reports/product-movement', methods=['GET'])
def report_product_movement():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product_ids = request.args.get('product_ids') # Comma separated
    brand = request.args.get('brand')
    partner_id = request.args.get('partner_id')
    move_type_filter = request.args.get('type') # 'Sale' or 'Purchase'
    
    query = OrderDetail.query.join(Order).join(Product).outerjoin(Partner)
    
    # Date filtering
    if start_date:
        try:
            dt = datetime.fromisoformat(start_date)
            query = query.filter(Order.date >= dt)
        except: pass
    if end_date:
        try:
            dt = datetime.fromisoformat(end_date)
            if len(end_date) <= 10:
                dt = dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Order.date <= dt)
        except: pass
        
    # Product filtering
    if product_ids:
        # product_ids is a string like "1,2,3"
        try:
            ids = [int(i.strip()) for i in product_ids.split(',') if i.strip()]
            if ids:
                query = query.filter(OrderDetail.product_id.in_(ids))
        except: pass
        
    # Brand filtering
    if brand:
        query = query.filter(Product.brand == brand)
        
    # Partner filtering
    if partner_id:
        try:
            pid = int(partner_id)
            query = query.filter(Order.partner_id == pid)
        except: pass

    # Type filtering (Sale/Purchase)
    if move_type_filter:
        query = query.filter(Order.type == move_type_filter)
        
    # Order by date descending
    query = query.order_by(Order.date.desc())
    
    details = query.options(joinedload(OrderDetail.product), joinedload(OrderDetail.order).joinedload(Order.partner)).all()
    
    results = []
    for d in details:
        # Movement type: Sale -> Xuất, Purchase -> Nhập
        move_type = 'Xuất' if d.order.type == 'Sale' else 'Nhập'
        
        results.append({
            'date': d.order.date.isoformat(),
            'display_id': d.order.display_id,
            'order_id': d.order.id,
            'product_name': d.product_name_override or (d.product.name if d.product else 'Sản phẩm đã xóa'),
            'brand': d.product.brand if d.product else '',
            'type': move_type,
            'quantity': d.quantity,
            'price': d.price,
            'total': d.quantity * d.price,
            'partner_name': d.order.partner.name if d.order.partner else ('KHÁCH LẺ' if d.order.type == 'Sale' else 'NCC VÃNG LAI'),
            'unit': d.product.unit if d.product else ''
        })
        
    return jsonify(results)

@app.route('/api/reports/synthesis', methods=['GET'])
def report_synthesis():
    r_type = request.args.get('type', 'Sale')  # Sale, Purchase
    year = request.args.get('year')
    month = request.args.get('month')
    quarter = request.args.get('quarter')
    day = request.args.get('day')
    # Filter params
    partner_id = request.args.get('partner_id', type=int)
    product_id = request.args.get('product_id', type=int)
    brand = request.args.get('brand')
    
    # Optional Date Range (overrides y/m/d/q if provided)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    page = request.args.get('page', type=int)
    limit = request.args.get('limit', type=int)

    group_by_brand = request.args.get('group_by_brand') == 'true'

    # Base query
    # We want: Partner Name, Product Name, Brand, Unit, Sum(Qty), Sum(Amount)
    if group_by_brand:
        query = db.session.query(
            db.func.coalesce(Partner.name, 'KHÁCH LẺ' if r_type == 'Sale' else 'NCC VÃNG LAI').label('partner_name'),
            db.func.max(Product.brand).label('brand'),
            db.func.sum(OrderDetail.quantity).label('total_qty'),
            db.func.sum(OrderDetail.quantity * OrderDetail.price).label('total_val'),
            Order.partner_id
        ).join(Order, OrderDetail.order_id == Order.id)\
         .join(Product, OrderDetail.product_id == Product.id)\
         .outerjoin(Partner, Order.partner_id == Partner.id)
    else:
        query = db.session.query(
            db.func.coalesce(Partner.name, 'KHÁCH LẺ' if r_type == 'Sale' else 'NCC VÃNG LAI').label('partner_name'),
            Product.name.label('product_name'),
            Product.brand,
            Product.unit,
            db.func.sum(OrderDetail.quantity).label('total_qty'),
            db.func.sum(OrderDetail.quantity * OrderDetail.price).label('total_val'),
            Order.partner_id,
            OrderDetail.product_id
        ).join(Order, OrderDetail.order_id == Order.id)\
         .join(Product, OrderDetail.product_id == Product.id)\
         .outerjoin(Partner, Order.partner_id == Partner.id)

    # Filter by Order Type
    query = query.filter(Order.type == r_type)

    # Date Filtering
    if start_date and end_date:
         query = query.filter(Order.date >= datetime.fromisoformat(start_date))\
                      .filter(Order.date <= datetime.fromisoformat(end_date))
    else:
        if year:
            query = query.filter(db.func.strftime('%Y', Order.date) == str(year))
        if month:
            query = query.filter(db.func.strftime('%m', Order.date) == str(month).zfill(2))
        if day:
            query = query.filter(db.func.strftime('%d', Order.date) == str(day).zfill(2))
        if quarter:
            quarters = {'1': ('01','03'), '2': ('04','06'), '3': ('07','09'), '4': ('10','12')}
            start, end = quarters.get(quarter, ('01','12'))
            query = query.filter(db.func.strftime('%m', Order.date).between(start, end))

    # Detailed Filters
    if partner_id is not None:
        if partner_id == 0:
            query = query.filter(Order.partner_id == None)
        else:
            query = query.filter(Order.partner_id == partner_id)
    if product_id and not group_by_brand:
        query = query.filter(OrderDetail.product_id == product_id)
    if brand:
        query = query.filter(Product.brand == brand)

    # Group By
    if group_by_brand:
        query = query.group_by(
            Order.partner_id,
            db.func.coalesce(Partner.name, 'KHÁCH LẺ' if r_type == 'Sale' else 'NCC VÃNG LAI')
        )
    else:
        query = query.group_by(
            Order.partner_id,
            OrderDetail.product_id,
            db.func.coalesce(Partner.name, 'KHÁCH LẺ' if r_type == 'Sale' else 'NCC VÃNG LAI'),
            Product.name,
            Product.brand,
            Product.unit
        )

    # Fetch
    rows = query.all()

    results = []
    if group_by_brand:
        for p_name, p_brand, qty, val, pid in rows:
            results.append({
                'partner_name': p_name,
                'product_name': f'TỔNG DOANH THU HÃNG: {p_brand or brand or "KHO"}',
                'brand': p_brand or brand or '',
                'unit': '',
                'quantity': qty,
                'revenue': val,
                'partner_id': pid,
                'product_id': None
            })
    else:
        for p_name, prod_name, p_brand, unit, qty, val, pid, prod_id in rows:
            results.append({
                'partner_name': p_name,
                'product_name': prod_name,
                'brand': p_brand or '',
                'unit': unit,
                'quantity': qty,
                'revenue': val,
                'partner_id': pid,
                'product_id': prod_id
            })

    # NEW: Support flattening for combos
    if request.args.get('flatten') == 'true' and not group_by_brand:
        results = expand_synthesis_records(results)

    # Sorting
    sort_by = request.args.get('sort_by', 'revenue')
    sort_order = request.args.get('sort_order', 'desc')
    reverse = (sort_order == 'desc')
    def get_sort_key(x):
        val = x.get(sort_by, 0)
        if val is None: return 0
        if isinstance(val, str): return val.lower()
        return val
        
    results.sort(key=get_sort_key, reverse=reverse)

    total = len(results)

    if page and limit:
        start = (page - 1) * limit
        end = start + limit
        items = results[start:end]
        return jsonify({
            'items': items,
            'total': total,
            'pages': (total + limit - 1) // limit,
            'current_page': page
        })

    return jsonify(results)

def expand_synthesis_records(rows):
    """
    Helper function to flatten combo products into their individual components.
    Used for reporting when inventory movement at component level is required.
    """
    memo = {} # Key: (partner_id, product_id) to re-aggregate
    
    for row in rows:
        p_id = row.get('product_id')
        if not p_id:
            key = (row.get('partner_id'), 'none')
            memo[key] = row.copy()
            continue

        p = Product.query.get(p_id)
        if p and p.is_combo:
            for ci in p.combo_items:
                child = ci.product
                if not child: continue
                
                child_qty = row['quantity'] * ci.quantity
                # Since this is a synthesis report, we pro-rate or just show estimated value.
                # Here we calculate based on child's sale price if available.
                child_val = child_qty * (child.sale_price or 0)
                
                key = (row.get('partner_id'), child.id)
                if key in memo:
                    memo[key]['quantity'] += child_qty
                    memo[key]['revenue'] += child_val
                else:
                    memo[key] = {
                        'partner_name': row['partner_name'],
                        'partner_phone': row.get('partner_phone', ''),
                        'product_name': child.name,
                        'brand': child.brand or '',
                        'unit': child.unit or '',
                        'quantity': child_qty,
                        'revenue': child_val,
                        'partner_id': row.get('partner_id'),
                        'product_id': child.id,
                        'is_from_combo': True,
                        'original_combo': p.name
                    }
        else:
            key = (row.get('partner_id'), p_id)
            if key in memo:
                memo[key]['quantity'] += row['quantity']
                memo[key]['revenue'] += row['revenue']
            else:
                memo[key] = row.copy()
                
    final_results = list(memo.values())
    final_results.sort(key=lambda x: x['revenue'], reverse=True)
    return final_results

@app.route('/api/reports/synthesis/export', methods=['GET'])
def export_synthesis_report():
    r_type = request.args.get('type', 'Sale')
    year = request.args.get('year')
    month = request.args.get('month')
    day = request.args.get('day')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    partner_id = request.args.get('partner_id', type=int)
    product_id = request.args.get('product_id', type=int)
    brand = request.args.get('brand')
    
    group_by_brand = request.args.get('group_by_brand') == 'true'

    # Base query
    # We want: Partner Name, Product Name, Brand, Unit, Sum(Qty), Sum(Amount)
    if group_by_brand:
        query = db.session.query(
            db.func.coalesce(Partner.name, 'KHÁCH LẺ' if r_type == 'Sale' else 'NCC VÃNG LAI').label('partner_name'),
            Partner.phone.label('partner_phone'),
            db.func.max(Product.brand).label('brand'),
            db.func.sum(OrderDetail.quantity).label('total_qty'),
            db.func.sum(OrderDetail.quantity * OrderDetail.price).label('total_val'),
            Order.partner_id
        ).join(Order, OrderDetail.order_id == Order.id)\
         .join(Product, OrderDetail.product_id == Product.id)\
         .outerjoin(Partner, Order.partner_id == Partner.id)
    else:
        query = db.session.query(
            db.func.coalesce(Partner.name, 'KHÁCH LẺ' if r_type == 'Sale' else 'NCC VÃNG LAI').label('partner_name'),
            Partner.phone.label('partner_phone'),
            Product.name.label('product_name'),
            Product.brand,
            Product.unit,
            db.func.sum(OrderDetail.quantity).label('total_qty'),
            db.func.sum(OrderDetail.quantity * OrderDetail.price).label('total_val'),
            Order.partner_id,
            OrderDetail.product_id
        ).join(Order, OrderDetail.order_id == Order.id)\
         .join(Product, OrderDetail.product_id == Product.id)\
         .outerjoin(Partner, Order.partner_id == Partner.id)

    query = query.filter(Order.type == r_type)

    if start_date and end_date:
         query = query.filter(Order.date >= datetime.fromisoformat(start_date))\
                      .filter(Order.date <= datetime.fromisoformat(end_date))
    else:
        if year: query = query.filter(db.func.strftime('%Y', Order.date) == str(year))
        if month: query = query.filter(db.func.strftime('%m', Order.date) == str(month).zfill(2))
        if day: query = query.filter(db.func.strftime('%d', Order.date) == str(day).zfill(2))

    if partner_id is not None:
         if partner_id == 0: query = query.filter(Order.partner_id == None)
         else: query = query.filter(Order.partner_id == partner_id)
    if product_id and not group_by_brand: query = query.filter(OrderDetail.product_id == product_id)
    if brand: query = query.filter(Product.brand == brand)

    # Group By
    if group_by_brand:
        rows = query.group_by(
            Order.partner_id,
            db.func.coalesce(Partner.name, 'KHÁCH LẺ' if r_type == 'Sale' else 'NCC VÃNG LAI'),
            Partner.phone
        ).all()
    else:
        rows = query.group_by(
            Order.partner_id,
            OrderDetail.product_id,
            db.func.coalesce(Partner.name, 'KHÁCH LẺ' if r_type == 'Sale' else 'NCC VÃNG LAI'),
            Partner.phone,
            Product.name,
            Product.brand,
            Product.unit
        ).all()

    results = []
    if group_by_brand:
        for p_name, p_phone, p_brand, qty, val, pid in rows:
            results.append({
                'partner_name': p_name,
                'partner_phone': p_phone or '',
                'product_name': f'TỔNG DOANH THU HÃNG: {p_brand or brand or "KHO"}',
                'brand': p_brand or brand or '',
                'unit': '',
                'quantity': qty,
                'revenue': val,
                'partner_id': pid,
                'product_id': None
            })
    else:
        for p_name, p_phone, prod_name, p_brand, unit, qty, val, pid, prod_id in rows:
            results.append({
                'partner_name': p_name,
                'partner_phone': p_phone or '',
                'product_name': prod_name,
                'brand': p_brand or '',
                'unit': unit,
                'quantity': qty,
                'revenue': val,
                'partner_id': pid,
                'product_id': prod_id
            })

    # Flatten for export
    if group_by_brand:
        flattened_results = results
    else:
        flattened_results = expand_synthesis_records(results)

    wb = Workbook()
    ws = wb.active
    ws.title = "BaoCaoTongHop"
    
    headers = ['Đối Tác', 'Số Điện Thoại', 'Sản Phẩm', 'Hãng', 'ĐVT', 'Số Lượng', 'Doanh Thu', 'Ghi Chú']
    ws.append(headers)
    for h in range(1, len(headers) + 1):
        ws.cell(row=1, column=h).font = Font(bold=True)

    for r in flattened_results:
        ws.append([
            r['partner_name'],
            r.get('partner_phone') or '',
            r['product_name'],
            r['brand'],
            r['unit'],
            r['quantity'],
            r['revenue'],
            f"Từ Combo: {r['original_combo']}" if r.get('is_from_combo') else ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"bao_cao_tong_hop_{r_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/reports/product-movement/export', methods=['GET'])
def export_product_movement():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product_ids = request.args.get('product_ids')
    brand = request.args.get('brand')
    partner_id = request.args.get('partner_id')
    move_type_filter = request.args.get('type')
    
    query = OrderDetail.query.join(Order).join(Product).outerjoin(Partner)
    
    if start_date:
        try: query = query.filter(Order.date >= datetime.fromisoformat(start_date))
        except: pass
    if end_date:
        try: 
            dt = datetime.fromisoformat(end_date)
            if len(end_date) <= 10: dt = dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Order.date <= dt)
        except: pass
        
    if product_ids:
        try:
            ids = [int(i.strip()) for i in product_ids.split(',') if i.strip()]
            if ids: query = query.filter(OrderDetail.product_id.in_(ids))
        except: pass
        
    if brand:
        query = query.filter(Product.brand == brand)
        
    if partner_id:
        try:
            pid = int(partner_id)
            query = query.filter(Order.partner_id == pid)
        except: pass

    if move_type_filter:
        query = query.filter(Order.type == move_type_filter)

    query = query.order_by(Order.date.desc())
    details = query.options(joinedload(OrderDetail.order).joinedload(Order.partner), joinedload(OrderDetail.product)).all()
    
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "BienDongHangHoa"
    
    headers = ['Ngày', 'Mã Đơn', 'Sản Phẩm', 'Hãng', 'Loại', 'Đối Tác', 'Số Lượng', 'Đơn Vị', 'Đơn Giá', 'Thành Tiền']
    ws.append(headers)
    
    for d in details:
        move_type = 'Xuất' if d.order.type == 'Sale' else 'Nhập'
        ws.append([
            d.order.date.strftime('%d/%m/%Y %H:%M'),
            d.order.display_id,
            d.product_name_override or (d.product.name if d.product else ''),
            d.product.brand if d.product else '',
            move_type,
            d.order.partner.name if d.order.partner else ('Khách Lẻ' if d.order.type == 'Sale' else 'NCC Vãng Lai'),
            d.quantity,
            d.product.unit if d.product else '',
            d.price,
            d.quantity * d.price
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='bao_cao_bien_dong_hang_hoa.xlsx'
    )

@app.route('/api/backup', methods=['GET'])
def download_backup():
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        db_path = get_storage_path(os.path.join("instance", "easypos.db"))
        
        if not os.path.exists(db_path):
             # Fallback to root if not in instance
             db_path = get_storage_path("easypos.db")
        
        if not os.path.exists(db_path):
            return jsonify({'error': 'Local Database file not found'}), 404
        
        return send_file(
            db_path,
            as_attachment=True,
            download_name=f"easypos_local_backup_{timestamp}.db"
        )
            
    except Exception as e:
        app.logger.error(f"Backup Error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/restore', methods=['POST'])
def restore_backup():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    try:
        # SQLite Restore Logic
        db_path = get_storage_path(os.path.join("instance", "easypos.db"))
        db.session.remove()
        db.engine.dispose()
        file.save(db_path)
        run_migrations()
        
        return jsonify({'message': 'Dữ liệu đã được khôi phục thành công! Hãy khởi động lại ứng dụng.'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error restoring backup: {e}")
        return jsonify({'error': str(e)}), 500

# --- Security / Reset ---

# --- Security / Reset ---
@app.route('/api/reset-database', methods=['POST'])
def reset_database():
    data = request.json
    password = data.get('password')
    
    # Simple password check
    if password != 'admin.reset':
        return jsonify({'error': 'Sai mật khẩu!'}), 403
        
    try:
        # Strictly SQLite Reset
        # Delete in strict reverse dependency order
        # 1. Transactions & Vouchers
        BankTransaction.query.delete()
        CashVoucher.query.delete()
        
        # 2. Details
        OrderDetail.query.delete()
        
        # 3. Orders
        Order.query.delete()
        
        # 4. Product dependencies
        ComboItem.query.delete()
        CustomerPrice.query.delete()
        
        # 5. Core Entities
        Product.query.delete()
        Partner.query.delete()
        BankAccount.query.delete()
        PrintTemplate.query.delete()
        
        db.session.commit()
        return jsonify({'message': 'Đã xóa toàn bộ dữ liệu thành công!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/optimize-db', methods=['POST'])
def optimize_db():
    try:
        with db.engine.begin() as conn:
            conn.execute(db.text('PRAGMA wal_checkpoint(TRUNCATE)'))
            conn.execute(db.text('PRAGMA incremental_vacuum'))
            conn.execute(db.text('VACUUM'))
            conn.execute(db.text('PRAGMA optimize'))
        return jsonify({'message': 'Tối ưu hóa dữ liệu thành công!'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clean-ram', methods=['POST'])
def clean_ram():
    try:
        import gc
        import ctypes
        
        # 1. Clear SQLAlchemy thread-local sessions (identity maps/objects cache)
        try:
            db.session.remove()
        except Exception as sess_err:
            print(f"Error removing db session: {sess_err}")

        # 2. Shrink SQLite database memory cache and dispose engine pool
        try:
            with db.engine.begin() as conn:
                conn.execute(db.text('PRAGMA shrink_memory'))
            # Dispose engine to close all connection pools (frees SQLite cache memory)
            db.engine.dispose()
        except Exception as db_err:
            print(f"Error shrinking/disposing SQLite memory: {db_err}")
            
        # 3. Force Python garbage collection aggressively
        gc.collect()
        gc.collect(0)
        gc.collect(1)
        gc.collect(2)
        
        # 4. Empty working set and set small working set size for current process (on Windows)
        if sys.platform == 'win32':
            try:
                from ctypes import wintypes
                kernel32 = ctypes.windll.kernel32
                psapi = ctypes.windll.psapi
                
                kernel32.GetCurrentProcess.restype = wintypes.HANDLE
                psapi.EmptyWorkingSet.argtypes = [wintypes.HANDLE]
                psapi.EmptyWorkingSet.restype = wintypes.BOOL
                
                kernel32.SetProcessWorkingSetSize.argtypes = [wintypes.HANDLE, ctypes.c_size_t, ctypes.c_size_t]
                kernel32.SetProcessWorkingSetSize.restype = wintypes.BOOL
                
                handle = kernel32.GetCurrentProcess()
                psapi.EmptyWorkingSet(handle)
                
                size_t_max = ctypes.c_size_t(-1).value
                kernel32.SetProcessWorkingSetSize(handle, size_t_max, size_t_max)
            except Exception as win_err:
                print(f"Error calling EmptyWorkingSet / SetProcessWorkingSetSize on backend: {win_err}")
                
        return jsonify({'message': 'Dọn dẹp RAM backend thành công!'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-stats', methods=['GET'])
def db_stats():
    try:
        stats = {
            'orders': Order.query.count(),
            'order_details': OrderDetail.query.count(),
            'products': Product.query.count(),
            'vouchers': CashVoucher.query.count(),
            'partners': Partner.query.count()
        }
        # Get file size
        db_path = get_storage_path(os.path.join("instance", "easypos.db"))
        if os.path.exists(db_path):
            size_mb = os.path.getsize(db_path) / (1024 * 1024)
            stats['db_size_mb'] = round(size_mb, 2)
        else:
            stats['db_size_mb'] = 0
            
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'API Not Found'}), 404
    
    # Safety check for static folder
    if not app.static_folder or not os.path.exists(os.path.join(app.static_folder, 'index.html')):
        return f"Frontend build not found in {app.static_folder}. Please ensure 'frontend/dist' exists.", 500
        
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/api/cash-vouchers', methods=['GET'])
def get_cash_vouchers():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search_partner = request.args.get('search_partner', '')
    voucher_type = request.args.get('type') # Receipt / Payment
    partner_id = request.args.get('partner_id', type=int)
    search_id = request.args.get('search_id', '')

    query = CashVoucher.query
    if search_partner:
        s_norm = remove_accents(search_partner)
        query = query.outerjoin(Partner).filter(db.func.remove_accents(db.func.coalesce(Partner.name, 'Hệ thống')).ilike(f'%{s_norm}%'))

    if start_date:
        try:
            dt = datetime.fromisoformat(start_date)
            query = query.filter(CashVoucher.date >= dt)
        except: pass
    if end_date:
        try:
            dt = datetime.fromisoformat(end_date)
            if len(end_date) <= 10:
                dt = dt.replace(hour=23, minute=59, second=59)
            query = query.filter(CashVoucher.date <= dt)
        except: pass
    
    if partner_id is not None:
        if partner_id == 0:
            query = query.filter(CashVoucher.partner_id == None)
        else:
            query = query.filter(CashVoucher.partner_id == partner_id)

    if search_id:
        # Support searching PT/PC IDs
        query = query.filter(db.cast(CashVoucher.id, db.String).ilike(f'%{search_id}%'))

    if voucher_type:
        query = query.filter(CashVoucher.type == voucher_type)
        
    vouchers = query.order_by(CashVoucher.date.desc()).all()
    return jsonify([v.to_dict() for v in vouchers])

# --- Customer Care Endpoints ---

@app.route('/api/events', methods=['GET'])
def get_events():
    events = Event.query.filter_by(is_active=True).order_by(Event.date.desc()).all()
    return jsonify([e.to_dict() for e in events])

@app.route('/api/events', methods=['POST'])
def create_event():
    data = request.json
    new_event = Event(
        name=data['name'],
        gift_types=data.get('gift_types'),
        icon=data.get('icon')
    )
    db.session.add(new_event)
    db.session.commit()
    return jsonify(new_event.to_dict()), 201

@app.route('/api/events/<int:id>', methods=['PUT'])
def update_event(id):
    event = Event.query.get_or_404(id)
    data = request.json
    if 'name' in data:
        event.name = data['name']
    if 'is_active' in data:
        event.is_active = data['is_active']
    if 'gift_types' in data:
        event.gift_types = data['gift_types']
    if 'icon' in data:
        event.icon = data['icon']
    db.session.commit()
    return jsonify(event.to_dict())

@app.route('/api/events/<int:id>', methods=['DELETE'])
def delete_event(id):
    event = Event.query.get_or_404(id)
    db.session.delete(event)
    db.session.commit()
    return jsonify({'message': 'Deleted successfully'})

@app.route('/api/event-logs', methods=['GET'])
def get_event_logs():
    # Optional filtering
    event_id = request.args.get('event_id')
    partner_id = request.args.get('partner_id')
    
    query = EventLog.query
    if event_id: query = query.filter_by(event_id=event_id)
    if partner_id: query = query.filter_by(partner_id=partner_id)
    
    logs = query.all()
    return jsonify([l.to_dict() for l in logs])

@app.route('/api/event-logs/toggle', methods=['POST'])
def toggle_event_log():
    data = request.json
    event_id = data.get('event_id')
    partner_id = data.get('partner_id')
    gift_type = data.get('gift_type') # Optional: specific gift type
    
    if not event_id or not partner_id:
        return jsonify({'error': 'Missing event_id or partner_id'}), 400
        
    log = EventLog.query.filter_by(event_id=event_id, partner_id=partner_id).first()
    
    if log:
        # If toggling OFF or switching type behavior
        # Logic: If submitting same type -> toggle OFF. If different type -> Update type.
        if gift_type and log.gift_type != gift_type:
            log.gift_type = gift_type
            status = True
        else:
            db.session.delete(log)
            status = False
    else:
        # Tick (Create log)
        new_log = EventLog(
            event_id=event_id, 
            partner_id=partner_id,
            gift_type=gift_type
        )
        db.session.add(new_log)
        status = True
        
    db.session.commit()
    return jsonify({'status': status, 'message': 'Toggled successfully'})

# --- Active Devices LAN Tracking System ---
import threading

active_devices_lock = threading.Lock()
active_devices_registry = {}  # ip -> { "last_seen": str, "user_agent": str, "label": str }

@app.before_request
def track_active_device():
    if request.path.startswith('/api/'):
        ip = request.remote_addr
        forwarded_for = request.headers.get('X-Forwarded-For')
        if forwarded_for:
            ip = forwarded_for.split(',')[0].strip()
            
        if request.path == '/api/active-devices':
            return
            
        user_agent = request.headers.get('User-Agent', '')
        
        # Determine human-friendly device type
        device_label = "Thiết bị Máy trạm"
        if "Tauri" in user_agent:
            device_label = "Ứng dụng Tauri (Client)"
        elif "Windows" in user_agent:
            device_label = "Máy tính Windows"
        elif "Android" in user_agent:
            device_label = "Điện thoại Android"
        elif "iPhone" in user_agent or "iPad" in user_agent:
            device_label = "Thiết bị iOS"
        elif "Macintosh" in user_agent:
            device_label = "Máy tính Mac"
        
        with active_devices_lock:
            active_devices_registry[ip] = {
                "last_seen": datetime.now().isoformat(),
                "user_agent": user_agent,
                "label": device_label
            }

@app.route('/api/active-devices', methods=['GET'])
def get_active_devices():
    now = datetime.now()
    active_list = []
    server_ip = get_local_ip()
    
    with active_devices_lock:
        inactive_ips = []
        for ip, info in active_devices_registry.items():
            try:
                last_seen_dt = datetime.fromisoformat(info["last_seen"])
                # 2 minutes of inactivity is standard for heartbeats
                if now - last_seen_dt > timedelta(minutes=2):
                    inactive_ips.append(ip)
                else:
                    is_host = ip in ('127.0.0.1', 'localhost', 'tauri.localhost') or ip == server_ip
                    active_list.append({
                        "ip": ip,
                        "last_seen": info["last_seen"],
                        "label": "Máy chủ chính (Host)" if is_host else info["label"],
                        "is_host": is_host
                    })
            except Exception:
                pass
                
        for ip in inactive_ips:
            del active_devices_registry[ip]
            
    # Guarantee host is in list
    has_host = any(item["is_host"] for item in active_list)
    if not has_host:
        active_list.insert(0, {
            "ip": server_ip or "127.0.0.1",
            "last_seen": datetime.now().isoformat(),
            "label": "Máy chủ chính (Host)",
            "is_host": True
        })
        
    return jsonify(active_list)

@app.route('/api/tauri/save-and-open', methods=['POST'])
def tauri_save_and_open():
    import base64
    import os
    import subprocess
    import socket
    
    # Check if request comes from local machine (host) or remote client
    server_ips = {'127.0.0.1', '::1', 'localhost'}
    try:
        hostname = socket.gethostname()
        for ip in socket.gethostbyname_ex(hostname)[2]:
            server_ips.add(ip)
    except Exception:
        pass
        
    client_ip = request.headers.get('X-Forwarded-For', request.headers.get('X-Real-IP', request.remote_addr))
    if client_ip and ',' in client_ip:
        client_ip = client_ip.split(',')[0].strip()
        
    if client_ip not in server_ips:
        return jsonify({
            'success': False,
            'error': 'not_host',
            'message': 'Cannot open natively on server for remote client'
        }), 200

    data = request.json or {}
    filename = data.get('filename', 'export.xlsx')
    base64_data = data.get('base64_data')
    
    if not base64_data:
        return jsonify({'error': 'No data provided'}), 400
        
    try:
        file_bytes = base64.b64decode(base64_data)
        
        # Save to Downloads folder
        downloads_path = ""
        try:
            if os.name == 'nt':
                import winreg
                sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
                with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
                    downloads_path = winreg.QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]
            else:
                downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
        except:
            downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
            
        if not os.path.exists(downloads_path):
            os.makedirs(downloads_path, exist_ok=True)
            
        file_path = os.path.join(downloads_path, filename)
        
        # Avoid overwriting existing files
        base, ext = os.path.splitext(file_path)
        counter = 1
        while os.path.exists(file_path):
            file_path = f"{base}_{counter}{ext}"
            counter += 1
            
        with open(file_path, 'wb') as f:
            f.write(file_bytes)
            
        # Open natively
        try:
            if os.name == 'nt':
                os.startfile(file_path)
            else:
                try:
                    if os.uname().sysname == 'Darwin':
                        subprocess.call(('open', file_path))
                    else:
                        subprocess.call(('xdg-open', file_path))
                except:
                    subprocess.call(('xdg-open', file_path))
        except Exception as open_err:
            print("Error opening file natively:", open_err)
            
        return jsonify({
            'success': True,
            'path': file_path
        })
    except Exception as e:
        print("Tauri save-and-open error:", e)
        return jsonify({'error': str(e)}), 500

@app.route('/api/ping', methods=['GET'])
def api_ping():
    return jsonify({'status': 'ok', 'message': 'pong'})

@app.route('/api/ip', methods=['GET'])
def get_ip():
    import socket
    return jsonify({
        'ip': get_local_ip(),
        'port': CURRENT_PORT,
        'hostname': socket.gethostname()
    })

@app.route('/api/network/unlock-firewall', methods=['POST'])
def unlock_firewall():
    import os
    if os.name != 'nt':
        return jsonify({'error': 'Tính năng này chỉ hỗ trợ hệ điều hành Windows.'}), 400
    try:
        import subprocess
        import base64
        port = CURRENT_PORT
        rule_name = f"LyangPOS LAN Access ({port})"
        
        # The powershell script to add the firewall rule
        script = f'netsh advfirewall firewall delete rule name="{rule_name}" 2>$null; netsh advfirewall firewall add rule name="{rule_name}" dir=in action=allow protocol=TCP localport={port}'
        encoded_script = base64.b64encode(script.encode('utf-16le')).decode('utf-8')
        
        # Command to run powershell elevated
        ps_cmd = f'Start-Process powershell -ArgumentList "-NoProfile -ExecutionPolicy Bypass -EncodedCommand {encoded_script}" -Verb RunAs -WindowStyle Hidden'
        
        # Run using subprocess
        subprocess.run(["powershell", "-NoProfile", "-Command", ps_cmd], check=True)
        
        return jsonify({
            'success': True,
            'message': 'Đã gửi yêu cầu mở Tường lửa. Vui lòng nhấn "Yes" (Đồng ý) nếu xuất hiện hộp thoại UAC (quyền Administrator) của Windows.'
        })
    except Exception as e:
        app.logger.error(f"Firewall Unlock Error: {e}")
        return jsonify({'error': f'Lỗi hệ thống khi mở khóa: {str(e)}'}), 500


def get_local_ip():
    import socket
    try:
        # Prioritize 192.168 block
        all_ips = socket.gethostbyname_ex(socket.gethostname())[2]
        for ip in all_ips:
            if ip.startswith("192.168."):
                return ip
        
        # Method 2: Connected socket (accurate for internet-connected machines)
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0.2)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        try:
            # Fallback to any non-loopback
            all_ips = socket.gethostbyname_ex(socket.gethostname())[2]
            for ip in all_ips:
                if not ip.startswith("127."):
                    return ip
            return all_ips[0] if all_ips else "127.0.0.1"
        except:
            return "127.0.0.1"

# --- Serve Frontend ---

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve(path):
    if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    else:
        return send_from_directory(app.static_folder, 'index.html')

# --- Integrity Check ---


@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    app.logger.info("Manual shutdown requested. Exiting...")
    # Use a thread to exit so we can return the response first
    threading.Timer(1.0, lambda: os._exit(0)).start()
    return jsonify({'message': 'Server is shutting down...'})

@app.route('/api/settings/repair-backend', methods=['POST'])
def repair_backend():
    try:
        from migrations_manager import link_orphaned_stock_batches
        with db.engine.begin() as conn:
            link_orphaned_stock_batches(conn)
        
        all_prods = Product.query.all()
        for p in all_prods:
            recalculate_product_cost_price(p.id)
            
        return jsonify({'message': 'Sửa lỗi và vá dữ liệu thành công!'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Run startup repair check
with app.app_context():
    try:
        repair_setting = AppSetting.query.filter_by(setting_key='repair_on_startup').first()
        if repair_setting and repair_setting.setting_value == 'true':
            app.logger.info("Startup Repair: 'repair_on_startup' is enabled. Starting repair...")
            from migrations_manager import link_orphaned_stock_batches
            with db.engine.begin() as conn:
                link_orphaned_stock_batches(conn)
            
            # Recalculate average costs for all products
            all_prods = Product.query.all()
            app.logger.info(f"Startup Repair: Recalculating cost for {len(all_prods)} products...")
            for p in all_prods:
                recalculate_product_cost_price(p.id)
                
            repair_setting.setting_value = 'false'
            db.session.commit()
            app.logger.info("Startup Repair: Completed successfully.")
    except Exception as e:
        app.logger.error(f"Startup repair error: {e}")

@app.route('/api/remote-scans', methods=['POST'])
def add_remote_scan():
    data = request.json
    barcode = data.get('barcode')
    if not barcode:
        return jsonify({'error': 'Barcode is required'}), 400
        
    scan = RemoteScanQueue(barcode=barcode)
    db.session.add(scan)
    db.session.commit()
    return jsonify(scan.to_dict()), 201

@app.route('/api/remote-scans/pop', methods=['GET'])
def pop_remote_scan():
    # Lấy mã vạch cũ nhất chưa xử lý
    scan = RemoteScanQueue.query.filter_by(is_processed=False).order_by(RemoteScanQueue.created_at.asc()).first()
    if scan:
        scan.is_processed = True
        db.session.commit()
        return jsonify(scan.to_dict())
@app.route('/api/tts', methods=['GET'])
def get_tts():
    import hashlib
    text = request.args.get('text', '')
    voice_type = request.args.get('voice', 'edge-vi-female')
    rate = request.args.get('rate', '1.0')
    
    if not text:
        return jsonify({"error": "Text parameter is required"}), 400
        
    voice_map = {
        'edge-vi-female': 'vi-VN-HoaiMyNeural',
        'edge-vi-male': 'vi-VN-NamMinhNeural',
        'google': 'vi-VN-HoaiMyNeural'
    }
    
    edge_voice = voice_map.get(voice_type, 'vi-VN-HoaiMyNeural')
    
    try:
        rate_val = float(rate)
        pct = int((rate_val - 1.0) * 100)
        rate_str = f"{'+' if pct >= 0 else ''}{pct}%"
    except:
        rate_str = "+0%"
        
    temp_dir = tempfile.gettempdir()
    hash_name = hashlib.md5(f"{text}_{edge_voice}_{rate_str}".encode('utf-8')).hexdigest()
    output_path = os.path.join(temp_dir, f"tts_{hash_name}.mp3")
    
    if not os.path.exists(output_path):
        try:
            async def run_edge_tts():
                communicate = edge_tts.Communicate(text, edge_voice, rate=rate_str)
                await communicate.save(output_path)
            
            try:
                loop = asyncio.get_event_loop()
            except RuntimeError:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
            
            loop.run_until_complete(run_edge_tts())
        except Exception as e:
            return jsonify({"error": str(e)}), 500
            
    return send_file(output_path, mimetype="audio/mpeg")

# --- Daily & Pending Invoice Tracking (Theo Dõi Xuất Hóa Đơn Khách Hàng) ---

@app.route('/api/accounting/daily-invoices', methods=['GET'])
def get_daily_invoices():
    try:
        scope = request.args.get('scope', 'daily') # 'daily', 'pending', 'completed'
        date_str = request.args.get('date') # YYYY-MM-DD
        if not date_str:
            date_str = get_vn_time().strftime('%Y-%m-%d')
            
        search = (request.args.get('search') or '').strip().lower()
        filter_status = request.args.get('status', 'all') # 'all', 'invoiced', 'uninvoiced'
        
        query = Order.query.filter(
            Order.type == 'Sale',
            Order.display_id.notin_(['NODAU', '#NODAU'])
        )
        
        if scope == 'daily':
            start_dt = datetime.strptime(date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
            end_dt = datetime.strptime(date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(Order.date >= start_dt, Order.date <= end_dt)
        elif scope == 'pending':
            # Get all uninvoiced orders (nợ HĐ từ trước tới nay)
            query = query.filter(Order.is_invoiced == False)
        elif scope == 'completed':
            query = query.filter(Order.is_invoiced == True)
            if date_str:
                start_dt = datetime.strptime(date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
                end_dt = datetime.strptime(date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
                query = query.filter(Order.date >= start_dt, Order.date <= end_dt)
                
        orders = query.order_by(Order.date.desc()).all()
        
        # Group by partner
        partners_map = {}
        total_sales_amount = 0
        invoiced_amount = 0
        uninvoiced_amount = 0
        total_orders_count = len(orders)
        invoiced_orders_count = 0
        uninvoiced_orders_count = 0
        
        for o in orders:
            p_id = o.partner_id or 0
            p_name = o.partner.name if o.partner else 'Khách Lẻ'
            p_phone = o.partner.phone if o.partner else ''
            p_address = o.partner.address if o.partner else ''
            p_debt = o.partner.debt_balance if o.partner else 0
            
            is_inv = bool(o.is_invoiced)
            total_sales_amount += (o.total_amount or 0)
            if is_inv:
                invoiced_amount += (o.total_amount or 0)
                invoiced_orders_count += 1
            else:
                uninvoiced_amount += (o.total_amount or 0)
                uninvoiced_orders_count += 1
                
            if p_id not in partners_map:
                partners_map[p_id] = {
                    'partner_id': p_id,
                    'partner_name': p_name,
                    'partner_phone': p_phone,
                    'partner_address': p_address,
                    'partner_debt': p_debt,
                    'total_amount': 0,
                    'amount_paid': 0,
                    'invoiced_orders_count': 0,
                    'uninvoiced_orders_count': 0,
                    'total_items_count': 0,
                    'invoiced_items_count': 0,
                    'pending_items_count': 0,
                    'orders': [],
                    'items': [],
                    'invoice_numbers': []
                }
                
            partners_map[p_id]['total_amount'] += (o.total_amount or 0)
            partners_map[p_id]['amount_paid'] += (o.amount_paid or 0)
            if is_inv:
                partners_map[p_id]['invoiced_orders_count'] += 1
                if o.invoice_no and o.invoice_no not in partners_map[p_id]['invoice_numbers']:
                    partners_map[p_id]['invoice_numbers'].append(o.invoice_no)
            else:
                partners_map[p_id]['uninvoiced_orders_count'] += 1
                
            order_items = []
            for d in o.details:
                d_dict = d.to_dict()
                d_is_inv = bool(d.is_invoiced) if d.is_invoiced is not None else is_inv
                d_dict['is_invoiced'] = d_is_inv
                d_dict['order_id'] = o.id
                d_dict['order_display_id'] = o.display_id or f"HD{o.id}"
                d_dict['order_date'] = o.date.isoformat() if o.date else None
                order_items.append(d_dict)
                
                partners_map[p_id]['items'].append(d_dict)
                partners_map[p_id]['total_items_count'] += 1
                if d_is_inv:
                    partners_map[p_id]['invoiced_items_count'] += 1
                else:
                    partners_map[p_id]['pending_items_count'] += 1
                
            partners_map[p_id]['orders'].append({
                'id': o.id,
                'display_id': o.display_id or f"HD{o.id}",
                'date': o.date.isoformat(),
                'total_amount': o.total_amount or 0,
                'amount_paid': o.amount_paid or 0,
                'payment_method': o.payment_method or '',
                'is_invoiced': is_inv,
                'invoice_no': o.invoice_no or '',
                'invoice_date': o.invoice_date.isoformat() if o.invoice_date else None,
                'invoice_note': o.invoice_note or '',
                'details_count': len(order_items),
                'details': order_items
            })
            
        partners_list = []
        for p in partners_map.values():
            p['total_orders_count'] = len(p['orders'])
            p['is_fully_invoiced'] = (p['pending_items_count'] == 0 and p['uninvoiced_orders_count'] == 0)
            
            # Filter search
            if search:
                name_match = search in p['partner_name'].lower()
                phone_match = search in p['partner_phone'].lower()
                order_match = any(search in (ord['display_id'] or '').lower() or search in (ord['invoice_no'] or '').lower() for ord in p['orders'])
                item_match = any(search in (item.get('product_name') or '').lower() or search in (item.get('product_code') or '').lower() for item in p['items'])
                if not (name_match or phone_match or order_match or item_match):
                    continue
                    
            # Filter status
            if filter_status == 'invoiced' and not p['is_fully_invoiced']:
                continue
            if filter_status == 'uninvoiced' and p['is_fully_invoiced']:
                continue
                
            # In 'daily' scope: hide already invoiced partners (they move to 'completed' tab)
            if scope == 'daily' and filter_status != 'invoiced' and p['is_fully_invoiced']:
                continue
                
            # In 'pending' scope: only show partners needing additional invoices
            if scope == 'pending' and p['is_fully_invoiced']:
                continue
                
            # In 'completed' scope: only show fully invoiced partners
            if scope == 'completed' and not p['is_fully_invoiced']:
                continue
                
            partners_list.append(p)
            
        # Sort: pending/uninvoiced first, then by total_amount desc
        partners_list.sort(key=lambda x: (x['is_fully_invoiced'], -x['total_amount']))
        
        invoiced_partners = sum(1 for p in partners_map.values() if p['is_fully_invoiced'])
        uninvoiced_partners = len(partners_map) - invoiced_partners
        
        return jsonify({
            'scope': scope,
            'date': date_str,
            'summary': {
                'total_partners_count': len(partners_map),
                'invoiced_partners_count': invoiced_partners,
                'uninvoiced_partners_count': uninvoiced_partners,
                'total_sales_amount': total_sales_amount,
                'invoiced_amount': invoiced_amount,
                'uninvoiced_amount': uninvoiced_amount,
                'total_orders_count': total_orders_count,
                'invoiced_orders_count': invoiced_orders_count,
                'uninvoiced_orders_count': uninvoiced_orders_count
            },
            'partners': partners_list
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/accounting/order-details/<int:detail_id>/invoice-status', methods=['POST'])
def update_order_detail_invoice_status(detail_id):
    try:
        detail = OrderDetail.query.get_or_404(detail_id)
        data = request.json or {}
        
        is_invoiced = data.get('is_invoiced', True)
        detail.is_invoiced = is_invoiced
        if is_invoiced:
            detail.invoiced_quantity = data.get('invoiced_quantity', detail.quantity)
            detail.invoice_no = data.get('invoice_no', detail.invoice_no or '')
        else:
            detail.invoiced_quantity = 0.0
            detail.invoice_no = ''
            
        # Check parent order
        order = detail.order_id and Order.query.get(detail.order_id)
        if order:
            all_details = OrderDetail.query.filter_by(order_id=order.id).all()
            all_invoiced = all(d.is_invoiced for d in all_details)
            order.is_invoiced = all_invoiced
            if all_invoiced and not order.invoice_date:
                order.invoice_date = get_vn_time()
            elif not all_invoiced and not any(d.is_invoiced for d in all_details):
                order.invoice_date = None
                
        db.session.commit()
        return jsonify(detail.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/accounting/orders/<int:order_id>/invoice-status', methods=['POST'])
def update_order_invoice_status(order_id):
    try:
        order = Order.query.get_or_404(order_id)
        data = request.json or {}
        
        is_invoiced = data.get('is_invoiced', True)
        order.is_invoiced = is_invoiced
        if is_invoiced:
            order.invoice_no = data.get('invoice_no', order.invoice_no or '')
            order.invoice_note = data.get('invoice_note', order.invoice_note or '')
            order.invoice_date = get_vn_time()
        else:
            order.invoice_no = ''
            order.invoice_note = ''
            order.invoice_date = None
            
        # Update all details of this order
        for d in order.details:
            d.is_invoiced = is_invoiced
            if is_invoiced:
                d.invoiced_quantity = d.quantity
                if order.invoice_no: d.invoice_no = order.invoice_no
            else:
                d.invoiced_quantity = 0.0
                d.invoice_no = ''
                
        db.session.commit()
        return jsonify(order.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/accounting/partners/<int:partner_id>/batch-invoice', methods=['POST'])
def batch_invoice_partner(partner_id):
    try:
        data = request.json or {}
        date_str = data.get('date')
        is_invoiced = data.get('is_invoiced', True)
        invoice_no = data.get('invoice_no', '')
        invoice_note = data.get('invoice_note', '')
        
        p_filter = (Order.partner_id == partner_id) if partner_id > 0 else (Order.partner_id.is_(None))
        
        query = Order.query.filter(
            Order.type == 'Sale',
            Order.display_id.notin_(['NODAU', '#NODAU']),
            p_filter
        )
        
        if date_str:
            start_dt = datetime.strptime(date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
            end_dt = datetime.strptime(date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(Order.date >= start_dt, Order.date <= end_dt)
            
        orders = query.all()
        
        for o in orders:
            o.is_invoiced = is_invoiced
            if is_invoiced:
                if invoice_no: o.invoice_no = invoice_no
                if invoice_note: o.invoice_note = invoice_note
                o.invoice_date = get_vn_time()
            else:
                o.invoice_no = ''
                o.invoice_note = ''
                o.invoice_date = None
                
            for d in o.details:
                d.is_invoiced = is_invoiced
                if is_invoiced:
                    d.invoiced_quantity = d.quantity
                    if invoice_no: d.invoice_no = invoice_no
                else:
                    d.invoiced_quantity = 0.0
                    d.invoice_no = ''
                    
        db.session.commit()
        return jsonify({'message': f'Đã cập nhật {len(orders)} đơn hàng thành công!', 'updated_count': len(orders)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/accounting/partners/bulk-batch-invoice', methods=['POST'])
def bulk_batch_invoice_partners():
    try:
        data = request.json or {}
        partner_ids = data.get('partner_ids', [])
        date_str = data.get('date')
        is_invoiced = data.get('is_invoiced', True)
        invoice_no = data.get('invoice_no', '')
        invoice_note = data.get('invoice_note', '')
        
        if not partner_ids:
            return jsonify({'message': 'Không có khách hàng nào được chọn', 'updated_count': 0})
            
        parsed_ids = []
        has_null_partner = False
        for pid in partner_ids:
            if pid is None or pid == 0 or pid == '0' or str(pid).lower() == 'null':
                has_null_partner = True
            else:
                try:
                    val = int(pid)
                    if val > 0:
                        parsed_ids.append(val)
                    else:
                        has_null_partner = True
                except (ValueError, TypeError):
                    has_null_partner = True
        
        conditions = []
        if parsed_ids:
            conditions.append(Order.partner_id.in_(parsed_ids))
        if has_null_partner:
            conditions.append(Order.partner_id.is_(None))
            
        if not conditions:
            return jsonify({'message': 'Không có đối tác hợp lệ', 'updated_count': 0})
            
        filter_cond = conditions[0] if len(conditions) == 1 else or_(*conditions)
        query = Order.query.filter(
            Order.type == 'Sale',
            Order.display_id.notin_(['NODAU', '#NODAU']),
            filter_cond
        )
        
        if date_str:
            start_dt = datetime.strptime(date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
            end_dt = datetime.strptime(date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(Order.date >= start_dt, Order.date <= end_dt)
            
        orders = query.all()
        
        for o in orders:
            o.is_invoiced = is_invoiced
            if is_invoiced:
                if invoice_no: o.invoice_no = invoice_no
                if invoice_note: o.invoice_note = invoice_note
                o.invoice_date = get_vn_time()
            else:
                o.invoice_no = ''
                o.invoice_note = ''
                o.invoice_date = None
                
            for d in o.details:
                d.is_invoiced = is_invoiced
                if is_invoiced:
                    d.invoiced_quantity = d.quantity
                    if invoice_no: d.invoice_no = invoice_no
                else:
                    d.invoiced_quantity = 0.0
                    d.invoice_no = ''
                    
        db.session.commit()
        return jsonify({'message': f'Đã xuất đủ hóa đơn cho {len(partner_ids)} khách hàng ({len(orders)} đơn hàng)!', 'updated_count': len(orders)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/accounting/partners/<int:partner_id>/update-items-invoice', methods=['POST'])
def update_partner_items_invoice(partner_id):
    try:
        data = request.json or {}
        items = data.get('items', []) # [{ detail_id, is_invoiced, invoiced_quantity, invoice_no }]
        invoice_no_default = data.get('invoice_no', '')
        
        updated_orders = set()
        
        for item in items:
            d_id = item.get('detail_id')
            if not d_id: continue
            detail = OrderDetail.query.get(d_id)
            if not detail: continue
            
            is_inv = item.get('is_invoiced', True)
            detail.is_invoiced = is_inv
            detail.invoiced_quantity = item.get('invoiced_quantity', detail.quantity if is_inv else 0.0)
            detail.invoice_no = item.get('invoice_no') or (invoice_no_default if is_inv else '')
            
            if detail.order_id:
                updated_orders.add(detail.order_id)
                
        # Re-evaluate all affected orders
        for order_id in updated_orders:
            order = Order.query.get(order_id)
            if order:
                all_details = OrderDetail.query.filter_by(order_id=order.id).all()
                all_invoiced = len(all_details) > 0 and all(d.is_invoiced for d in all_details)
                order.is_invoiced = all_invoiced
                if all_invoiced:
                    if invoice_no_default and not order.invoice_no:
                        order.invoice_no = invoice_no_default
                    if not order.invoice_date:
                        order.invoice_date = get_vn_time()
                elif not any(d.is_invoiced for d in all_details):
                    order.invoice_no = ''
                    order.invoice_date = None
                    
        db.session.commit()
        return jsonify({'message': 'Đã cập nhật trạng thái các món thành công!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        ensure_schema(db.engine)
        initialize_stock_batches(db.engine)
    is_bundle = getattr(sys, 'frozen', False)
    local_ip = get_local_ip()

    try:
        print("\n" + "="*50)
        print("LYANG POS IS RUNNING!")
        print(f"Local access: http://localhost:{CURRENT_PORT}")
        print(f"Network access (Wifi): http://{local_ip}:{CURRENT_PORT}")
        print("="*50 + "\n")
    except:
        pass

    # Start Flask in a background thread
    # Daemon=False so the app stays alive after Splash Screen (Tkinter) closes
    flask_thread = threading.Thread(target=lambda: app.run(host="0.0.0.0", port=CURRENT_PORT, debug=False, use_reloader=False), daemon=False)
    flask_thread.start()
    
    IS_TAURI = ('--tauri' in sys.argv) or (os.environ.get('TAURI_ENV') == 'true')
    
    if IS_TAURI:
        app.logger.info("Running in Tauri mode. Splash screen and browser opening are disabled.")
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            pass
    elif not os.environ.get('NO_GUI') and not os.environ.get('HEADLESS'):
        try:
            splash = SplashScreen(CURRENT_PORT)
            splash.run()
        except Exception as e:
            print(f"Could not start Splash Screen: {e}")
            # Keep main thread alive if splash fails or is skipped
            try:
                while True: time.sleep(1)
            except KeyboardInterrupt:
                pass
    else:
        # Headless mode: Keep main thread alive
        try:
            while True: time.sleep(1)
        except KeyboardInterrupt:
            pass

